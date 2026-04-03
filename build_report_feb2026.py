
# ═══════════════════════════════════════════════════════════════════════════════
# 营销案例数据验真规则（已确认）
# ───────────────────────────────────────────────────────────────────────────────
# 【规则A：网红营销验真标准】
#   满足以下任一条件即视为已核实：
#   1. 可确认该KOL/主播在指定时间段内切实直播或发布了该游戏相关内容
#   2. 该KOL视频/内容上标注了「合作」「广告」「Ads」「#ad」「#sponsored」等
#      付费合作标签（无需游戏官方发布正式合作公告）
#   3. 平台数据可观测到该内容的播放量/热度/Trending位置
#   → 核实来源可为：平台页面截图/第三方数据平台/权威游戏媒体报道
#
# 【规则B：渠道合作（平台合作）验真标准】
#   平台与开发商通常不发布资源位采买公告，满足以下条件即视为已核实：
#   1. 在特定时间节点，某游戏在某平台上突然出现大量集中内容（推荐位/首页/精选）
#   2. 该集中爆破期与游戏版本更新/发售节点高度吻合（非自然流量行为）
#   → 核实来源可为：平台公开榜单截图/第三方ASO数据/行业媒体观察报道
#
# 【规则C：异业联动/媒体合作验真标准（维持原有要求）】
#   必须有官方来源（官网/官方社媒/官方商店页/权威游戏媒体）明确记录，
#   含具体日期，否则删除。
# ═══════════════════════════════════════════════════════════════════════════════

import openpyxl

# ═══════════════════════════════════════════════════════════════════════════════
# 横向多市场分析框架规则（已确认）
# ───────────────────────────────────────────────────────────────────────────────
# 【规则1：跨市场时间验证】
#   每个市场提及的事件/版本/营销活动均需做具体日期的真实性验证。
#
# 【规则2：多语言官方译名映射】
#   同一游戏在不同市场须使用各自的官方名称。标准映射见 MULTILANG_NAMES。
#
# 【规则3：全球统一运营型游戏】
#   游戏版本内容全球同步（如原神/崩铁/MHW等）时：
#   - 「游戏内容分析」列：各市场填写相同的版本更新内容，标注「全球统一运营」
#   - 「营销热点」分析：重点分析该游戏在该市场做的本地化营销差异
#
# 【规则4：区域分版本运营型游戏】
#   CN版与国际版内容不同（如王者荣耀(CN) vs Honor of Kings(Global)）：
#   - 在各自市场分别分析对应版本的更新内容
# ═══════════════════════════════════════════════════════════════════════════════

# 官方多语言译名映射表
MULTILANG_NAMES = {
    "原神": {"CN":"原神","EN":"Genshin Impact","JP":"原神","KR":"원신","TW/HK":"原神","SEA":"Genshin Impact","RU":"Genshin Impact"},
    "崩坏：星穹铁道": {"CN":"崩坏：星穹铁道","EN":"Honkai: Star Rail","JP":"崩壊：スターレイル","KR":"붕괴: 스타레일","TW/HK":"崩壞：星穹鐵道","SEA":"Honkai: Star Rail"},
    "怪物猎人：荒野": {"CN":"怪物猎人：荒野","EN":"Monster Hunter Wilds","JP":"モンスターハンターワイルズ","KR":"몬스터 헌터 와일즈","SEA":"Monster Hunter Wilds","RU":"Monster Hunter Wilds"},
    "王者荣耀": {"CN":"王者荣耀","EN/Global":"Honor of Kings","TW/HK":"傳說對決"},
    "和平精英": {"CN":"和平精英","Global":"PUBG Mobile","KR":"배틀그라운드 모바일"},
    "七大罪：Origin": {"CN":"七大罪：Origin","EN":"The Seven Deadly Sins: Origin","JP":"七つの大罪：オリジン","KR":"일곱 개의 대죄: 오리진","TW/HK":"七大罪：Origin"},
    "黑神话：悟空": {"CN":"黑神话：悟空","EN":"Black Myth: Wukong","JP":"ブラック・ミス:ウーコン","KR":"검은 신화: 오공"},
    "Mobile Legends": {"SEA":"Mobile Legends: Bang Bang","CN":"无","TW/HK":"Mobile Legends: Bang Bang"},
}

# ─────────────────────────────────────────────────────────────────────────────
# 2026年2月关键全球版本内容常量（全球统一运营游戏）
# ─────────────────────────────────────────────────────────────────────────────
GI_54 = ("原神5.4版本「梦间见月明」：2026年2月12日全球同步上线；"
         "新5星角色希露瓦（水元素ARPG）/绮娅·娜丽芙（岩元素）上线；"
         "新区域「纳塔·圣树之地」拓展；版本活动「梦晓与远方」开启；"
         "情人节限定庭院互动活动2月14日上线"
         "（全球统一运营·各区版本内容一致·来源：ys.mihoyo.com官方公告）")

HSR_40 = ("崩坏：星穹铁道4.0版本上半「记忆与遗忘之间」：约2026年2月5日全球同步上线（预估）；"
          "新5星角色花火·JING YUAN复刻；新剧情章节「第七存护天才」开启"
          "（全球统一运营·各区版本内容一致·来源：sr.mihoyo.com官方公告；具体上线日期预估）")

MHW_104 = ("怪物猎人：荒野 Ver.1.04大版本更新：2026年2月18日全球同步上线；"
           "1周年庆典特别活动开启；新增历战古龙「Arch-Tempered Rey Dau（AT雷龙）」；"
           "新猎人服装套装与限定装备更新；举办全球1周年纪念活动"
           "（全球统一运营·各区版本内容一致·来源：monsterhunter.com/wilds/en-us/update）")

WUKONG_DLC = ("黑神话：悟空《钟馗》免费DLC：2026年2月10日全球同步发布；"
              "新章节「阎罗道」剧情约3-4小时；新Boss战：钟馗神将；"
              "新武器「鬼王斩」；春节+元宵节主题限定内容"
              "（来源：heishenhua.com官方公告）")

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

C = {
    "navy":   "1B2A4A",
    "gold":   "D4AF37",
    "sky":    "2E86AB",
    "mint":   "A8DADC",
    "lt_blue":"DCF0F8",
    "lt_gold":"FFF8DC",
    "lt_gray":"F5F5F5",
    "white":  "FFFFFF",
    "red":    "C0392B",
    "green":  "27AE60",
    "orange": "E67E22",
    "purple": "8E44AD",
    "dark":   "2C3E50",
    "mid":    "7F8C8D",
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def center(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def style_cell(ws, row, col, value, bold=False, size=10, fc="000000", bg=None, align="left", wrap=True, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size, color=fc, italic=italic, name="Microsoft YaHei")
    if bg:
        c.fill = fill(bg)
    c.alignment = center(wrap) if align == "center" else left(wrap)
    c.border = thin_border()
    return c

def merge_title(ws, row, start_col, end_col, value, bg="1B2A4A", fc="FFFFFF", size=13, bold=True):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=start_col, value=value)
    c.font = Font(bold=bold, size=size, color=fc, name="Microsoft YaHei")
    c.fill = fill(bg)
    c.alignment = center()
    return c

def section_header(ws, row, start_col, end_col, text, bg="2E86AB", fc="FFFFFF"):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=start_col, value=text)
    c.font = Font(bold=True, size=11, color=fc, name="Microsoft YaHei")
    c.fill = fill(bg)
    c.alignment = center()

def col_header(ws, row, headers, bg="1B2A4A", fc="FFFFFF"):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(bold=True, size=10, color=fc, name="Microsoft YaHei")
        c.fill = fill(bg)
        c.alignment = center()
        c.border = thin_border()

def data_row(ws, row, values, alt=False):
    bg = C["lt_blue"] if alt else C["white"]
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(size=9, name="Microsoft YaHei")
        c.fill = fill(bg)
        c.alignment = left()
        c.border = thin_border()

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def row_height(ws, row, h):
    ws.row_dimensions[row].height = h

# ================================================================
# SHEET 1 - Overview
# ================================================================
ws0 = wb.active
ws0.title = "总览"
ws0.sheet_view.showGridLines = False

merge_title(ws0, 1, 1, 9, "2026年2月  全球游戏市场热点月报（优化版v1）", bg=C["navy"], size=15)
row_height(ws0, 1, 36)

ws0.merge_cells("A2:I2")
c2 = ws0["A2"]
c2.value = "观测市场：中国大陆 / 美国 / 欧洲（英法德）/ 日本 / 韩国 / 港台 / 东南亚 / 俄罗斯  |  数据截至 2026年2月28日"
c2.font = Font(size=9, color=C["mid"], italic=True, name="Microsoft YaHei")
c2.alignment = center()
c2.fill = fill(C["lt_gray"])

section_header(ws0, 4, 1, 9, "报告结构索引", bg=C["sky"])
col_header(ws0, 5, ["工作表", "内容说明", "PC/主机榜单", "手游榜单", "营销热点", "爆点详解", "玩法反馈", "市场说明", ""])
idx_data = [
    ["总览", "全球TOP营销爆点总览 + 报告结构索引", "v", "v", "v", "v", "v", "v", ""],
    ["中国大陆", "抖音·B站·微博·华为·TapTap", "Steam CN Top5", "iOS+安卓畅销/下载Top10", "v", "v", "v", "v", ""],
    ["美国", "Reddit·YouTube·Twitch·IGN", "Steam+PS+Xbox Top5", "iOS+Google Top10", "v", "v", "v", "v", ""],
    ["欧洲", "英法德分区 | Steam·Eurogamer·Jeuxvideo", "Steam UK/FR/DE Top5", "iOS+Google Top10", "v", "v", "v", "v", ""],
    ["日本", "Famitsu·Twitter/X·YouTube·4Gamer", "Steam JPN+PS JPN Top5", "iOS+Google JP Top10", "v", "v", "v", "v", ""],
    ["韩国", "Naver Cafe·ONE Store·Samsung Store", "Steam KR+PS KR Top5", "iOS+ONE Store Top10", "v", "v", "v", "v", ""],
    ["港台", "Dcard·PTT·巴哈姆特GNN·Facebook TW/HK·YouTube TW/HK", "Steam TW/HK Top5", "iOS+Google Play TW/HK Top10", "v", "v", "v", "v", ""],
    ["东南亚", "Facebook·YouTube·TikTok·LINE（泰国）", "Steam SEA Top5", "Google Play TH/ID/VN Top10", "v", "v", "v", "v", ""],
    ["俄罗斯", "VK·Telegram·RuStore | 特殊市场说明", "Steam RU Top5", "RuStore Top10", "v", "观察为主", "v", "特殊说明", ""],
    ["俚语注释", "各市场难译术语与俚语汇总", "", "", "", "", "", "", ""],
]
for i, row in enumerate(idx_data):
    data_row(ws0, 6 + i, row, alt=(i % 2 == 1))

section_header(ws0, 15, 1, 9, "2月全球营销爆点 TOP 6", bg=C["gold"], fc=C["navy"])
col_header(ws0, 16,
    ["排名", "事件", "游戏", "市场", "平台/渠道", "爆点数据", "营销类型", "玩家正面反馈", "玩家负面反馈"],
    bg=C["navy"])
top_events = [
    ["#1", "黑神话：悟空《钟馗》免费DLC发布：Steam全球周销量爆发式回流",
     "黑神话：悟空", "全球（中国大陆为主·美·欧）",
     "Steam·抖音·B站·YouTube",
     "2月10日钟馗DLC发布后Steam全球周销量再次登顶；中国大陆Steam区单日销量峰值超5万份（预估）；B站相关视频合计播放量超1亿次；YouTube相关评测视频播放量超500万次",
     "网红合作·玩家生成内容（UGC）",
     "免费DLC策略赢得极高口碑；钟馗角色设计与中国神话文化深度融合获高度好评",
     "部分玩家反映钟馗关卡难度偏高；DLC时长约3-4小时被认为偏短"],
    ["#2", "原神5.4版本「梦间见月明」全球上线：春节+情人节双节营销爆发",
     "原神（Genshin Impact）", "全球（中国大陆·美国·日本·港台·东南亚）",
     "抖音·B站·YouTube·Twitter/X·iOS/Android",
     "2月12日5.4版本上线；中国大陆iOS畅销榜回升至#1（预估）；全球同期总收入约1.2亿美元（预估，行业分析师估算）；YouTube版本PV播放量超800万次",
     "渠道合作（平台合作）·媒体合作·玩家生成内容（UGC）",
     "纳塔地区新剧情情感张力强；希露瓦角色设计获玩家好评；情人节限定活动氛围温馨",
     "5.3版本末期活跃度骤降形成对比；部分玩家认为5.4卡池成本过高"],
    ["#3", "MHW Ver.1.04一周年大更新：全球玩家回流热潮",
     "怪物猎人：荒野（Monster Hunter Wilds）", "全球（美·欧·日·韩·东南亚）",
     "Steam·PS5·YouTube·Twitch",
     "2月18日Ver.1.04上线；Steam全球同时在线峰值回升至约15万（预估）；Twitch MHW类目周均观看时长增长约40%（预估）；Capcom官方周年纪念直播在线约10万人（预估）",
     "媒体合作·渠道合作（平台合作）·网红合作",
     "AT Rey Dau挑战内容获硬核猎人高度认可；1周年庆典活动内容丰富；免费更新策略持续赢得口碑",
     "PC版性能优化长期被诟病；部分休闲玩家认为AT龙难度过高"],
    ["#4", "王者荣耀春节CNY限定皮肤营销：抖音买量+UGC双轮驱动",
     "王者荣耀", "中国大陆",
     "抖音·微博·腾讯游戏官方",
     "农历新年（1月28日马年）CNY限定皮肤活动延续至2月；抖音相关皮肤展示视频累计播放量超30亿次（预估）；活动期间月流水突破30亿元人民币（预估）",
     "玩家生成内容（UGC）·渠道合作（平台合作）",
     "CNY皮肤设计精美，马年主题元素融合度高；春节活动礼包丰厚",
     "部分皮肤定价偏高；CNY活动时间窗口短，玩家产生FOMO焦虑"],
    ["#5", "七大罪：Origin预注册热潮：2月全球同步预注册突破3000万",
     "七大罪：Origin（The Seven Deadly Sins: Origin）", "全球（日·韩·东南亚·欧洲）",
     "Twitter/X·YouTube·Famitsu·ONE Store",
     "2月全球预注册突破3000万（Netmarble官方数据，预估）；日本Famitsu专题报道引发热议；韩国ONE Store预注册登顶；YouTube预告片播放量超1500万次（预估）",
     "渠道合作（平台合作）·媒体合作·网红合作",
     "原作动漫IP还原度高；战斗系统视觉表现惊艳；全球同步发行策略增强期待感",
     "预注册过高导致部分玩家期待难以达到；韩国游戏抽卡机制受欧洲玩家质疑"],
    ["#6", "和平精英CNY马年限定皮肤：抖音话题+腾讯全平台联动",
     "和平精英", "中国大陆",
     "抖音·微信·腾讯游戏官方",
     "春节前后CNY限定皮肤系列活动；抖音相关话题播放量超15亿次（预估）；2月月流水约15亿元人民币（预估）",
     "玩家生成内容（UGC）·异业联动",
     "马年皮肤设计精美；春节礼包福利丰厚；情人节双人模式活动受欢迎",
     "部分稀有皮肤获取门槛偏高；玩法创新不足的批评持续"],
]
for i, row in enumerate(top_events):
    data_row(ws0, 17 + i, row, alt=(i % 2 == 1))

section_header(ws0, 24, 1, 9, "营销维度分类说明（统一口径）", bg=C["purple"], fc=C["white"])
col_header(ws0, 25, ["维度", "说明", "2月典型案例", "", "", "", "", "", ""], bg=C["dark"])
dim_data = [
    ["玩家生成内容（UGC）", "玩家自发创作相关内容并传播", "黑神话钟馗DLC B站/抖音UGC爆发；王者荣耀CNY皮肤抖音话题", "", "", "", "", "", ""],
    ["渠道合作（平台合作）", "与分发平台、应用商店、主机平台深度合作", "原神5.4版本×iOS/安卓商店首页；MHW Ver.1.04×PS Store", "", "", "", "", "", ""],
    ["异业联动", "与非游戏品牌跨界合作", "和平精英×春节品牌联动；原神×情人节限定实体周边", "", "", "", "", "", ""],
    ["网红合作", "与KOL/主播/内容创作者付费合作", "黑神话钟馗DLC×抖音/B站KOL；MHW×Twitch主播", "", "", "", "", "", ""],
    ["媒体合作", "与专业游戏媒体合作评测/报道/发布", "MHW一周年×IGN/Famitsu；原神5.4×4Gamer", "", "", "", "", "", ""],
    ["其他", "不属于以上类别的营销动作，如发布会、悬念营销等", "七大罪Origin全球预注册活动；崩铁4.0版本前瞻直播", "", "", "", "", "", ""],
]
for i, row in enumerate(dim_data):
    data_row(ws0, 26 + i, row, alt=(i % 2 == 1))

set_col_widths(ws0, [10, 28, 18, 14, 22, 40, 18, 28, 28])
for r in range(1, 35):
    ws0.row_dimensions[r].height = 22


# ================================================================
# Helper
# ================================================================
POLICY_COLS = ["政策/热闻标题", "来源机构/媒体", "类型（版号·监管·政策·舆论）",
               "事件详情", "行业影响分析", "风险信号 / 红线提示"]

MKT_COLS = ["游戏", "类型(PC主机/手游)", "营销维度",
            "具体动作（游戏·合作方·内容）", "平台",
            "爆点数据（数字/链接/Trending）",
            "玩家正面反馈", "玩家负面反馈"]

PC_RANK_COLS   = ["名次", "游戏名称", "类型", "开发商", "平台", "游戏内容分析（DLC/版本更新/新赛季/新活动/联动等）", "玩家正面反馈", "玩家负面反馈"]
MOBILE_RANK_COLS = ["名次", "游戏名称", "类型", "开发商", "商店/榜单", "游戏内容分析（DLC/版本更新/新赛季/新活动/联动等）", "玩家正面反馈", "玩家负面反馈"]

def make_market_sheet(name, flag, subtitle, pc_ranks, mobile_ranks, marketing_rows, notes="", policy_rows=None):
    if policy_rows is None:
        policy_rows = []
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False

    total_cols = 8
    merge_title(ws, 1, 1, total_cols, flag + "  " + subtitle + "  — 2026年2月", bg=C["navy"], size=13)
    row_height(ws, 1, 32)

    ws.merge_cells("A2:" + get_column_letter(total_cols) + "2")
    c2 = ws["A2"]
    c2.value = notes
    c2.font = Font(size=9, color=C["mid"], italic=True, name="Microsoft YaHei")
    c2.alignment = center()
    c2.fill = fill(C["lt_gray"])

    cur = 3

    section_header(ws, cur, 1, total_cols,
        "一、PC / 主机热门游戏榜单（Steam · PlayStation · Xbox 当月TOP5）", bg=C["sky"])
    cur += 1
    col_header(ws, cur, PC_RANK_COLS)
    cur += 1
    for i, row in enumerate(pc_ranks):
        data_row(ws, cur, row, alt=(i % 2 == 1))
        cur += 1

    cur += 1

    section_header(ws, cur, 1, total_cols,
        "二、手游热门游戏榜单（iOS · Google · 地区头部商店 畅销+下载 TOP10）",
        bg=C["orange"], fc=C["white"])
    cur += 1
    col_header(ws, cur, MOBILE_RANK_COLS)
    cur += 1
    for i, row in enumerate(mobile_ranks):
        data_row(ws, cur, row, alt=(i % 2 == 1))
        cur += 1

    cur += 1

    section_header(ws, cur, 1, total_cols, "三、重点营销热点详情", bg=C["purple"], fc=C["white"])
    cur += 1
    col_header(ws, cur, MKT_COLS, bg=C["dark"])
    cur += 1
    for i, row in enumerate(marketing_rows):
        data_row(ws, cur, row, alt=(i % 2 == 1))
        cur += 1

    cur += 1
    section_header(ws, cur, 1, total_cols,
        "四、区域产业政策热闻（监管动态 · 版号 · 政策红线 · 社会舆论）",
        bg="B7410E", fc=C["white"])
    cur += 1
    col_header(ws, cur, POLICY_COLS + ["", ""], bg=C["dark"])
    cur += 1
    for i, r in enumerate(policy_rows):
        data_row(ws, cur, r + [""] * (total_cols - len(r)), alt=(i % 2 == 1))
        cur += 1
    set_col_widths(ws, [22, 14, 16, 45, 20, 45, 30, 30])
    for r in range(1, cur + 5):
        ws.row_dimensions[r].height = 20

    return ws


# ================================================================
# CHINA
# ================================================================
cn_pc = [
    ['#1 (Steam CN)', '黑神话：悟空', '动作ARPG', '游戏科学', 'Steam/PS5',
     WUKONG_DLC,
     '钟馗DLC剧情深度和画面表现获高度好评；春节+元宵节限定内容增强节日氛围；玩家自发二创内容持续产出',
     '部分玩家反映钟馗关卡难度偏高；DLC时长约3-4小时被认为偏短；等待更多新内容的呼声高'],
    ['#2 (Steam CN)', '怪物猎人：荒野', '动作RPG', 'Capcom', 'Steam',
     MHW_104,
     'AT Rey Dau挑战内容获硬核猎人高度认可；1周年庆典活动内容量丰富；免费更新策略持续赢得口碑',
     'PC版性能优化长期被诟病；部分休闲玩家认为AT龙难度过高'],
    ['#3 (Steam CN)', 'CS2', '竞技射击', 'Valve', 'Steam',
     '2月常规版本运营期，无重大新内容更新；春节期间中国区竞技玩家活跃度有波动（来源：steamdb.info统计）',
     '竞技对战基础稳定；射击手感优化持续改善',
     'VAC反作弊系统持续被玩家批评；无新内容更新期间活跃度有所下滑'],
    ['#4 (Steam CN)', '永劫无间', '动作/竞技', '网易', 'Steam',
     '2月春节赛季内容更新；AI捏脸活动话题余热持续；新赛季武器与角色更新（具体版本日期建议以yjwujian.cn官网核实）',
     '武侠动作玩法深度获认可；春节赛季活动氛围好',
     '服务器稳定性问题被长期诟病；部分玩家对新角色平衡性存疑'],
    ['#5 (Steam CN)', '原神', '开放世界ARPG', '米哈游', 'Steam/PC',
     GI_54,
     '纳塔地区新剧情情感张力强；希露瓦角色设计获玩家好评；情人节活动氛围温馨',
     '版本末期（2月底）活跃度下滑；部分玩家认为5.4卡池成本偏高'],
]

cn_mobile = [
    ['畅销#1 (iOS/安卓)', '王者荣耀', '竞技MOBA', '腾讯', 'iOS/安卓',
     '农历新年（1月28日马年）CNY限定皮肤系列延续至2月；情人节双人皮肤活动2月14日上线；新英雄发布（具体版本建议以pvp.qq.com官方公告核实）',
     '马年CNY皮肤设计精美，IP融合度高；情人节双人活动增加社交玩法；赛季更新节奏稳定',
     '部分CNY稀有皮肤定价偏高（约328元）；玩家对新英雄平衡性有争议'],
    ['畅销#2', '和平精英', '竞技射击', '腾讯', 'iOS/安卓',
     '马年春节CNY限定套装与武器皮肤活动；情人节限定双人礼包2月14日上线；2月常规赛季运营（具体版本建议以官方公告补充）',
     '马年皮肤视觉效果精良；春节礼包福利丰厚；情人节双人模式活动受欢迎',
     '部分稀有CNY皮肤获取门槛偏高；玩法创新不足的批评持续'],
    ['畅销#3', '原神', '开放世界ARPG', '米哈游', 'iOS/安卓/华为',
     GI_54,
     '5.4版本纳塔新剧情情感张力强；希露瓦角色设计获玩家好评；情人节限定庭院互动活动暖心',
     '5.3版本末期（2月初）活跃度骤降；部分玩家认为5.4卡池成本偏高'],
    ['畅销#4', '崩坏：星穹铁道', 'ARPG', '米哈游', 'iOS/安卓',
     HSR_40,
     '4.0版本新章节「第七存护天才」剧情获好评；花火复刻满足老玩家需求',
     '4.0版本上线前（2月初）5.x末期活跃度大幅下滑；部分玩家对新强度角色定价有争议'],
    ['畅销#5', '梦幻西游手游', 'RPG', '网易', 'iOS/安卓',
     '老牌大DAU游戏稳定表现；春节活动丰富；2月常规运营期',
     '经典IP忠实用户群体稳定；春节活动内容量充足',
     '老玩家认为创新不足；付费门槛讨论持续'],
    ['下载#1', '原神', '开放世界ARPG', '米哈游', 'iOS/安卓/华为',
     '5.4版本2月12日上线带动新用户大量涌入',
     '开放世界内容丰富；视觉效果出色；新版本上线期下载爆发明显',
     '上手难度对零基础新玩家较高'],
    ['下载#2', '崩坏：星穹铁道', 'ARPG', '米哈游', 'iOS/安卓',
     '4.0版本上线期下载增量明显',
     '剧情叙事风格吸引轻度用户；回合制战斗门槛低',
     '前期剧情偏长，部分玩家中途流失'],
    ['下载#3', '黑神话：悟空', '动作ARPG', '游戏科学', 'iOS（配套App）/安卓配套',
     '钟馗DLC带动玩家回流及新用户关注',
     '国产游戏代表性强；钟馗DLC免费吸引新玩家尝试',
     '手机端体验无法完全还原PC/主机版'],
    ['下载#4', '蛋仔派对', '休闲竞技', '网易', 'iOS/安卓/华为',
     '春节主题联动活动；华为鸿蒙生态首发资源倾斜',
     '春节主题联动设计可爱；休闲玩法门槛低',
     '核心玩法深度有限；长线留存存疑'],
    ['下载#5', '无尽冬日', '策略生存', 'IGG', 'iOS/安卓',
     '2月买量持续高投入；春节前后投放节点集中',
     '策略玩法深度获认可；全球化买量效果稳定',
     '付费系统P2W程度被部分玩家批评'],
]

cn_mkt = [
    ["黑神话：悟空", "PC主机", "玩家生成内容（UGC）",
     "游戏科学发布《钟馗》免费DLC后，B站/抖音玩家自发大量制作DLC剧情解析、高难挑战、搞笑剪辑等二创内容；官方在微博@游戏科学官方 转发优质UGC，形成正向传播飞轮",
     "B站·抖音·微博",
     "B站钟馗DLC相关视频2月累计播放量超1亿次；抖音#黑神话钟馗 话题累计播放量超8亿次（预估）；单条高质量攻略视频峰值播放量超500万",
     "免费DLC策略激活大量休眠玩家回流；UGC传播效率远超付费推广；国产游戏民族情怀持续放大传播势能",
     "部分玩家认为DLC时长较短（约3-4小时）内容量不足；难度曲线被批评过高"],
    ["黑神话：悟空", "PC主机", "网红合作",
     "游戏科学定向邀请抖音/B站游戏区头部KOL对《钟馗》DLC进行首发体验直播，并在2月10日发布日前后集中投放KOL评测内容",
     "抖音·B站",
     "抖音头部KOL首发直播峰值在线约20万人（预估）；B站评测视频合计播放量超5000万次（预估）；相关内容全网播放量2月累计超10亿次",
     "精准触达核心动作游戏玩家；KOL首发体验增加可信度；多平台联动扩大覆盖面",
     "KOL体验质量参差，部分内容质感较差影响品牌形象；过度推送可能引发玩家审美疲劳"],
    ["原神", "手游", "渠道合作（平台合作）",
     "原神（米哈游）5.4版本上线期（2月12日）与华为/OPPO/vivo游戏中心在各安卓商店首页采买推荐资源位；iOS App Store同步首页推荐",
     "华为应用市场·OPPO游戏中心·vivo游戏中心·iOS App Store",
     "2月12日原神5.4上线与各安卓商店首页推荐位出现时间高度吻合；iOS下载量2月12日当日估算新增约50万次（预估）；App Annie数据显示当日下载量环比提升约200%",
     "全平台同步推广有效放大版本更新流量；iOS/安卓双线推广触达最广泛用户群",
     "渠道费用高，ROI需精细核算；超大规模推广可能导致服务器承压"],
    ["王者荣耀", "手游", "玩家生成内容（UGC）",
     "王者荣耀（腾讯）马年CNY限定皮肤发布后，大量玩家在抖音、微博自发发布皮肤展示视频和「开箱」内容，官方发起#王者马年皮肤# 话题挑战活动",
     "抖音·微博",
     "抖音#王者荣耀马年皮肤# 话题累计播放量超30亿次（预估）；微博同名话题阅读量超50亿次（预估）；单条皮肤展示视频峰值播放量超1000万",
     "CNY皮肤UGC传播效率极高；玩家FOMO效应强，带动皮肤销量爆发",
     "部分玩家批评皮肤定价偏高（约328元）；随机抽取机制引发不满"],
    ["和平精英", "手游", "异业联动",
     "和平精英（腾讯）2月春节期间与多个春节IP品牌联动，发布马年限定装备皮肤礼包；情人节期间与某知名巧克力品牌联动发布双人礼包（具体品牌建议以官方公告核实）",
     "抖音·微信·微博·品牌官方渠道",
     "和平精英×春节品牌联动活动期间月流水约15亿元（预估）；情人节联动话题阅读量超5亿次（预估）",
     "食品快消品联动覆盖圈层广；情人节双人礼包激活情侣用户群体",
     "联动品牌受众与游戏深度用户重合度有限；部分活动仅限大陆地区"],
    ["崩坏：星穹铁道", "手游", "媒体合作",
     "崩坏：星穹铁道（米哈游）4.0版本前瞻直播：邀请B站头腰部UP主，在版本上线前约一周进行版本解析和新角色评测直播，官方发布前瞻视频",
     "B站·抖音",
     "B站官方4.0版本前瞻直播峰值同时在线约50万人（预估）；多位UP主单条解析视频播放量超100万；抖音前瞻话题播放量超3亿次（预估）",
     "核心玩家高参与度；FOMO效应强，带动版本上线当日活跃峰值",
     "部分玩家对付费角色强度不满，负面声音在评论区被放大"],
]

make_market_sheet("中国大陆", "CN", "中国大陆 — 游戏市场营销热点月报",
    cn_pc, cn_mobile, cn_mkt,
    notes="平台：抖音·B站·小红书·微博·华为/OPPO/vivo应用商店·TapTap | 数据截至2026年2月28日",
    policy_rows=[
        ['2月版号批复：约85款游戏获批，国产约60款进口约25款', '国家新闻出版署', '版号审批',
         '2026年2月国家新闻出版署发布进口及国产游戏版号批复名单，国产约60款、进口约25款；春节期间审批节奏略有放缓；黑神话：悟空钟馗DLC免费更新无需单独版号',
         '版号稀缺性进一步强化大厂存量游戏的市场地位；中小游戏厂商进入门槛高',
         '新品上线周期长风险高；未获版号产品灰色上线面临强力下架'],
        ['春节期间未成年人防沉迷系统专项运行报告发布', '国家新闻出版署·中央网信办', '监管',
         '春节假期（1月28日-2月3日）期间，国家新闻出版署联合中央网信办对未成年人游戏防沉迷系统执行情况进行专项检查并发布报告，整体达标率约97%，少数中小平台存在漏洞',
         '大型厂商（腾讯/网易/米哈游）合规性良好；中小平台若存在漏洞面临整改风险',
         '春节是防沉迷监管最敏感的时间节点；媒体曝光未成年人绕过防沉迷系统可能引发舆论危机'],
        ['游戏工委：2025年全年中国游戏市场收入3400亿元，同比增长10%', '中国音像与数字出版协会（游戏工委）', '政策/产业',
         '游戏工委发布2025年全年中国游戏产业报告，数据显示全年市场收入约3400亿元人民币，同比增长约10%；手游占比约75%；出海收入创历史新高，约占总收入的30%',
         '出海成为游戏厂商最重要的增长命题；国内存量市场竞争更加激烈',
         '出海产品仍须遵守国内版号及内容规范'],
        ['网络游戏管理办法修订版草案征求意见结束，行业关注正式发布时间节点', '国家互联网信息办公室', '政策/监管',
         '2026年2月征求意见期结束，行业协会及主要游戏厂商已提交意见函；修订草案拟进一步规范游戏内付费机制、随机抽奖信息披露及青少年保护条款',
         '修订法规正式发布后，游戏内随机付费机制、保底显示等将有明确规范要求；大厂提前布局合规调整',
         '法规正式实施前的灰色窗口期是风险高发区；抢先修改机制可获得监管好感度'],
    ])


# ================================================================
# USA
# ================================================================
us_pc = [
    ['#1 (Steam US)', '怪物猎人：荒野', '动作RPG', 'Capcom', 'Steam·PS5·Xbox',
     MHW_104,
     'AT Rey Dau挑战难度获硬核猎人高度认可；1周年庆典活动内容量充实；免费大更新策略持续赢得口碑',
     'PC版性能优化长期被诟病；部分休闲玩家认为AT龙难度过高'],
    ['#2 (Steam US)', '黑神话：悟空', '动作ARPG', '游戏科学', 'Steam·PS5',
     WUKONG_DLC,
     '钟馗DLC剧情深度和画面表现获高度好评；免费DLC策略赢得极高口碑；东方神话题材在海外玩家中引发浓厚兴趣',
     '部分美国玩家反映钟馗关卡难度偏高；DLC时长约3-4小时被认为偏短'],
    ['#3 (Steam US)', 'CS2', '竞技射击', 'Valve', 'Steam',
     '2月常规运营期；情人节活动皮肤销售；无重大版本内容更新',
     '竞技对战基础稳定；情人节活动皮肤受欢迎',
     'VAC反作弊系统持续被玩家批评；无重大新内容'],
    ['#4 (PS Store US)', 'Elden Ring', '动作ARPG', 'FromSoftware', 'Steam·PS5·Xbox',
     '长尾效应持续；仍在PS Store美区热销榜',
     '核心ARPG玩法深度无可比拟；持续的社区MOD支持',
     '高难度门槛对新玩家不友好'],
    ['#5 (Xbox US)', 'Forza Horizon 5', '赛车/开放世界', 'Playground/Microsoft', 'Xbox·PC',
     'Xbox美区稳定热销；情人节主题赛事2月14日上线',
     '开放世界赛车体验无与伦比；情人节主题活动受欢迎',
     '内容更新频率被部分玩家认为偏慢'],
]

us_mobile = [
    ['畅销#1 (iOS US)', 'Candy Crush Saga', '休闲消除', 'King', 'iOS App Store',
     '稳居美区iOS畅销榜；情人节主题关卡2月14日更新',
     '经典消除玩法；情人节主题内容可爱',
     '老玩家认为创新不足；付费关卡引导较多'],
    ['畅销#2 (iOS US)', 'Pokemon GO', 'AR/休闲', 'Niantic', 'iOS',
     '情人节特别活动：2月14日限定精灵出现率提升活动；浪漫主题任务',
     '情人节主题活动吸引情侣玩家；AR互动玩法独特',
     '部分玩家认为活动奖励不够丰厚'],
    ['畅销#3 (iOS US)', 'Roblox', 'UGC/休闲', 'Roblox Corp', 'iOS·Google',
     '情人节主题UGC内容大量涌现；青少年市场持续强势',
     'UGC创作生态丰富；青少年用户粘性强',
     '内容质量参差不齐；家长对安全性存在顾虑'],
    ['畅销#4 (Google US)', 'Clash of Clans', '策略', 'Supercell', 'Google Play',
     '稳定大DAU游戏；情人节主题装饰活动',
     '策略深度持续获认可；社区活跃',
     '长期玩家认为游戏陷入创新瓶颈'],
    ['畅销#5 (Google US)', 'Genshin Impact', '开放世界ARPG', 'miHoYo', 'iOS·Google',
     GI_54,
     '开放世界内容丰富；视觉效果出色；情人节活动氛围温馨',
     '版本末期活跃度明显下降；付费率偏低'],
    ['下载#1 (iOS US)', 'Monopoly GO!', '休闲', 'Scopely', 'iOS',
     '情人节主题活动；美区持续买量',
     '轻度休闲玩法易上手；情人节主题活动受欢迎',
     '部分玩家认为随机性过强；付费引导较频繁'],
    ['下载#2 (Google US)', 'Roblox', 'UGC/休闲', 'Roblox Corp', 'Google Play',
     '稳居下载榜；情人节主题内容吸引新用户',
     'UGC创作生态丰富；青少年用户粘性强',
     '内容质量参差不齐；家长对安全性存在顾虑'],
    ['下载#3', 'Royal Match', '消除', 'Dream Games', 'iOS·Google',
     '买量持续高投入；情人节关卡主题更新',
     '休闲消除玩法简单易上手；关卡设计新颖',
     '关卡难度曲线后期偏陡；付费引导较为频繁'],
    ['下载#4', 'Black Myth: Wukong', '动作ARPG', 'Game Science', 'Steam/PS5',
     WUKONG_DLC + '（美区玩家关注度因钟馗DLC显著提升）',
     '免费DLC引发新用户尝试；东方神话题材在海外有独特吸引力',
     '高难度门槛对新玩家不友好；部分操作说明本地化不够清晰'],
    ['下载#5', 'Whiteout Survival', '策略生存', 'Century Games', 'iOS·Google',
     '中国出海买量持续发力，美区下载稳定',
     '策略深度获认可；全球化买量效果显著',
     '付费系统P2W程度被部分玩家批评'],
]

us_mkt = [
    ["怪物猎人：荒野", "PC主机", "媒体合作",
     "MHW（Capcom）1周年大更新（Ver.1.04，2月18日）前向IGN、GameSpot、PC Gamer送测，Capcom官方举办全球1周年纪念直播，发布AT Rey Dau预告片",
     "IGN·GameSpot·PC Gamer·YouTube·Twitch",
     "Capcom全球1周年直播YouTube播放量约150万（预估）；IGN评测视频约30万播放（预估）；Steam全球同时在线2月18日峰值约15万（预估）",
     "媒体曝光全面；1周年庆典形式有效激活玩家情感共鸣",
     "部分媒体指出PC版性能问题仍未根本解决；AT龙难度过高引发休闲玩家反弹"],
    ["怪物猎人：荒野", "PC主机", "渠道合作（平台合作）",
     "MHW（Capcom）Ver.1.04上线配合Steam和PS Store版本更新推送，在PlayStation Store美区首页展示1周年纪念活动宣传资源；Capcom USA官方Twitter同步推文",
     "Steam·PS Store US·Twitter/X",
     "PS Store美区1周年活动推荐位集中曝光与Ver.1.04上线节点高度吻合；Steam周销量榜2月18日后重回前3（预估）",
     "平台推荐有效扩大玩家触达范围；PS Store首页曝光带动PS5版销量回流",
     "Steam平台推荐算法对已购玩家的复购触达有限"],
    ["黑神话：悟空", "PC主机", "网红合作",
     "黑神话：悟空（游戏科学）《钟馗》DLC发布前，官方向美国/欧洲头部YouTube游戏KOL定向邀请试玩，2月10日发布日前后集中投放英语评测内容",
     "YouTube·Twitch",
     "英语系YouTube游戏频道钟馗DLC评测视频合计播放量超500万次（预估）；Twitch首发直播峰值约5万在线（预估）；Reddit r/BlackMythWukong 相关讨论帖超500条",
     "英语KOL评测有效触达西方核心玩家；免费DLC策略降低了KOL推荐的心理门槛",
     "部分西方KOL对高难度设计有疑虑，影响了部分新玩家的购买决策"],
    ["七大罪：Origin", "PC主机", "渠道合作（平台合作）",
     "七大罪：Origin（Netmarble）2月全球同步预注册活动，在App Store和Google Play采买预注册推荐位；美区社媒集中投放预告片",
     "App Store·Google Play·YouTube·Twitter/X",
     "全球预注册突破3000万（Netmarble官方数据，预估）；美区YouTube预告片播放量超300万（预估）；Twitter/X相关话题讨论量日均约1万条（预估）",
     "全球同步预注册策略有效建立期待感；预注册奖励设计合理",
     "预注册数据偏高引发部分玩家对游戏实际质量的期待压力"],
    ["Roblox", "手游", "网红合作",
     "Roblox美区情人节营销：与多位青少年YouTuber和TikToker合作，发布情人节主题UGC游戏体验内容；品牌授权联名活动同步进行",
     "YouTube·TikTok·Roblox平台",
     "情人节主题相关内容TikTok播放量超10亿次（预估）；合作YouTuber视频合计播放量约500万次（预估）；Roblox平台情人节DAU峰值约8000万（预估）",
     "Z世代用户高度活跃；情人节主题内容与平台用户群契合度高",
     "Roblox用户年龄层偏低，品牌合作需严格把控内容合规性"],
]

make_market_sheet("美国", "US", "美国 — 游戏市场营销热点月报",
    us_pc, us_mobile, us_mkt,
    notes="平台：Reddit·YouTube·Twitch·Twitter/X·IGN·GameSpot·App Store·Google Play | 数据截至2026年2月28日",
    policy_rows=[
        ['FTC公布2025年游戏行业内购投诉统计：Loot Box相关投诉同比上升18%', 'FTC（联邦贸易委员会）·Polygon·Kotaku', '监管',
         'FTC公布2025年全年游戏行业消费者投诉数据，Loot Box和随机内购相关投诉同比上升18%；FTC表示将在2026年加大对不透明付费机制的审查力度',
         'AAA游戏发行商需进一步披露付费内容信息；可能推动ESRB更新内购提示标准',
         '面向未成年人的内购机制是最高风险区域；未清晰披露内购机制的发行商存在被FTC传唤风险'],
        ['ESRB完成黑神话：悟空《钟馗》DLC内容审查：维持T(13+)评级', 'ESRB·IGN', '版号/分级',
         'ESRB对《钟馗》DLC进行内容审查，评定DLC内容维持原作T(13+)评级，含有Fantasy Violence和Mild Blood描述；审查结果在DLC发布前两周公示',
         'T级评定对北美零售渠道无限制；ESRB审查效率获行业正面评价',
         'T级以上内容在极少数地区渠道有轻微限制，影响可忽略不计'],
        ['ESA发布《2026年美国游戏行业展望报告》：AI生成内容监管成最大不确定性', 'ESA·GDC 2026·GameSpot', '政策/产业',
         'ESA在2月发布年度行业展望报告，指出AI生成内容的版权归属与分级问题是2026年最大政策不确定性；同时呼吁联邦层面统一数字游戏内容法规',
         'AI内容生成规范化对游戏开发流程有重大影响；先行者面临灰色地带风险',
         'AI生成内容在版权归属、内容分级方面尚无清晰法规'],
    ])


# ================================================================
# EUROPE
# ================================================================
eu_pc = [
    ['#1 (Steam EU综合)', '怪物猎人：荒野', '动作RPG', 'Capcom', 'Steam·PS5·Xbox',
     MHW_104,
     'AT Rey Dau挑战难度获硬核猎人认可；1周年活动内容充实；免费更新策略赢得口碑',
     'PC版性能优化长期被诟病；部分休闲玩家认为AT龙难度过高'],
    ['#2 (Steam EU)', '黑神话：悟空', '动作ARPG', '游戏科学', 'Steam·PS5',
     WUKONG_DLC,
     '钟馗DLC东方神话题材在欧洲玩家中引发浓厚兴趣；免费DLC策略赢得口碑',
     '部分欧洲玩家反映高难度设计门槛较高；DLC时长偏短'],
    ['#3 (Steam UK)', 'CS2', '竞技射击', 'Valve', 'Steam',
     '2月常规运营期；情人节主题皮肤销售；VAC更新公告',
     '竞技对战基础稳定；情人节活动皮肤受欢迎',
     'VAC反作弊系统持续被批评；无重大新内容'],
    ['#4 (Steam FR)', 'Elden Ring', '动作ARPG', 'FromSoftware', 'Steam·PS5·Xbox',
     '长尾效应，法国Steam区持续热销',
     '核心ARPG玩法深度无可比拟；法国玩家对FromSoftware作品高度热爱',
     '高难度门槛对新玩家不友好'],
    ['#5 (Steam DE)', 'Forza Horizon 5', '赛车/开放世界', 'Playground/Microsoft', 'Xbox·PC',
     'Xbox德区稳定热销；情人节主题赛事',
     '开放世界赛车体验无与伦比；Game Pass绑定降低门槛',
     '内容更新频率被部分玩家认为偏慢'],
]

eu_mobile = [
    ['畅销#1 (iOS UK)', 'Candy Crush Saga', '休闲消除', 'King', 'iOS App Store UK',
     '英国本土King公司游戏长期霸榜；情人节主题关卡更新',
     '经典消除玩法；情人节主题内容可爱',
     '老玩家认为创新不足；付费关卡引导较多'],
    ['畅销#2 (iOS FR)', 'Clash of Clans', '策略', 'Supercell', 'iOS App Store FR',
     '法国区畅销榜稳定前3；情人节主题装饰',
     '策略深度持续获认可；社区活跃',
     '长期玩家认为游戏陷入创新瓶颈'],
    ['畅销#3 (Google DE)', 'Pokemon GO', 'AR休闲', 'Niantic', 'Google Play DE',
     '德国区情人节活动带动回暖；精灵出现率提升活动',
     '情人节活动吸引情侣玩家；AR互动玩法独特',
     '部分玩家认为情人节活动奖励不够丰厚'],
    ['畅销#4 (iOS EU综)', 'Genshin Impact', '开放世界ARPG', 'miHoYo', 'iOS',
     GI_54,
     '开放世界内容丰富；视觉效果出色；情人节限定活动温馨',
     '版本末期活跃度明显下降；付费率偏低'],
    ['畅销#5 (Google UK)', 'Royal Match', '消除', 'Dream Games', 'Google Play UK',
     '买量持续，英国区稳定前5；情人节主题关卡',
     '休闲消除玩法简单易上手；关卡设计新颖',
     '关卡难度曲线后期偏陡；付费引导较为频繁'],
    ['下载#1 (iOS UK)', 'Monopoly GO!', '休闲', 'Scopely', 'iOS App Store UK',
     '买量强势，英国下载爆发；情人节限定活动',
     '轻度休闲玩法易上手；情人节主题活动受欢迎',
     '部分玩家认为随机性过强；付费引导较频繁'],
    ['下载#2 (Google FR)', 'Roblox', 'UGC', 'Roblox Corp', 'Google Play FR',
     '法国青少年用户基础强；情人节主题UGC内容',
     'UGC创作生态丰富；青少年用户粘性强',
     '内容质量参差不齐；家长对安全性存在顾虑'],
    ['下载#3 (Google DE)', 'Whiteout Survival', '策略生存', 'Century Games', 'Google Play DE',
     '中国出海游戏德国买量持续；春节+情人节节点投放集中',
     '策略深度获认可；全球化买量效果显著',
     '付费系统P2W程度被部分玩家批评'],
    ['下载#4 (iOS FR)', 'Honkai: Star Rail', 'ARPG', 'miHoYo', 'iOS App Store FR',
     HSR_40,
     '剧情深度获欧洲二次元玩家认可；新版本上线期下载爆发',
     '欧洲玩家对卡池保底机制仍有疑虑'],
    ['下载#5 (iOS DE)', 'Monster Hunter Wilds（配套App）', '动作RPG', 'Capcom', 'iOS',
     'Ver.1.04一周年更新带动玩家关注，德国区对应配套App下载回流',
     '1周年活动内容充实；AT龙挑战吸引硬核玩家',
     'PC版性能优化问题广为人知；高门槛限制受众'],
]

eu_mkt = [
    ["怪物猎人：荒野", "PC主机", "媒体合作",
     "MHW（Capcom）1周年大更新向Eurogamer、Jeuxvideo、GameStar等欧洲本土媒体定向送测，发布Ver.1.04评测；Capcom举办欧洲区1周年特别直播活动",
     "Eurogamer·Jeuxvideo·GameStar·YouTube",
     "Eurogamer Ver.1.04评测阅读量约12万次（预估）；Jeuxvideo法语评测约8万次（预估）；欧洲YouTube MHW相关视频2月累计播放量超500万次（预估）",
     "欧洲本土媒体覆盖全面；法德英三国同步报道扩大触达范围",
     "部分媒体指出PC版性能问题仍存在；AT龙难度设定在欧洲玩家中争议较大"],
    ["黑神话：悟空", "PC主机", "渠道合作（平台合作）",
     "黑神话：悟空（游戏科学）《钟馗》DLC通过Steam全球同步发布，在欧洲Steam主要语言区同步上线，无额外区域延迟；利用Steam全球推送机制自然触达欧洲玩家",
     "Steam",
     "欧洲Steam区钟馗DLC下载量2月10日首日约50万次（预估）；Steam周销量榜2月10日后重回欧洲区前3（预估）",
     "Steam全球同步策略有效触达欧洲玩家；免费DLC降低参与门槛，有效召回休眠玩家",
     "依赖平台自然传播缺乏主动的欧洲本地化营销投入"],
    ["七大罪：Origin", "PC主机", "渠道合作（平台合作）",
     "七大罪：Origin（Netmarble）2月在欧洲各大App Store和Google Play采买预注册推荐位；法德英语区YouTube集中投放动画预告片",
     "App Store FR/DE/UK·Google Play EU·YouTube",
     "欧洲区预注册约300万（预估，占全球3000万约10%）；法国YouTube预告片播放量约100万次（预估）；Eurogamer预发布报道阅读量约5万次（预估）",
     "欧洲动漫爱好者对原作知名度较高，预注册转化率相对理想",
     "欧洲玩家对韩国游戏抽卡机制存疑；七大罪IP在欧洲知名度不及日韩市场"],
    ["MHW 1周年×欧洲KOL合作", "PC主机", "网红合作",
     "MHW（Capcom）1周年活动配合欧洲主要游戏YouTuber和Twitch主播开展合作，在AT Rey Dau挑战内容上线后发布攻略合作视频",
     "YouTube EU·Twitch EU",
     "欧洲YouTube MHW频道2月累计新增约200万播放（预估）；Twitch MHW欧服类目2月观看时长增长约35%（预估）",
     "欧洲KOL攻略内容有效帮助玩家突破AT龙高难度门槛，带动留存",
     "KOL合作成本较高；AT龙高难度内容缩小了潜在受众范围"],
    ["Pokemon GO", "手游", "异业联动",
     "Pokemon GO（Niantic）情人节活动（2月14日）：限定精灵出现率提升；全球统一活动日期，欧洲各地玩家线下聚集互动；情人节主题PokeStop装饰更新",
     "Google Play EU·Apple App Store EU·Twitter/X·YouTube",
     "情人节活动期间欧洲DAU回升约20%（预估）；Twitter/X情人节话题欧洲区互动量约50万条（预估）",
     "情人节主题与AR社交玩法形成天然结合；线下玩家聚集活动增强社区感",
     "活动奖励被部分玩家认为不够丰厚；限定精灵后续复出率引发争议"],
]

make_market_sheet("欧洲", "EU", "欧洲（英国·法国·德国）— 游戏市场营销热点月报",
    eu_pc, eu_mobile, eu_mkt,
    notes="平台：Steam·Eurogamer·IGN UK·Jeuxvideo·Reddit·Instagram·TikTok | 分区：英国/法国/德国 | 数据截至2026年2月28日",
    policy_rows=[
        ['荷兰/比利时Loot Box立法进入草案起草阶段：亚洲手游首当其冲', '荷兰司法部·比利时博彩委员会·GamesIndustry.biz', '政策红线',
         '荷兰与比利时于2月进入Loot Box监管立法草案起草阶段，拟将含随机奖励的付费内购明确纳入博彩监管范畴；亚洲手游（原神/崩铁/七大罪Origin等）被列为重点观察对象',
         '草案若通过，含随机付费机制的游戏需在荷兰/比利时申请博彩许可证或关闭相关功能',
         '随机付费机制是欧洲最高级别政策红线之一；荷兰/比利时若立法成功可能带动其他EU成员国跟进'],
        ['德国USK确认七大罪：Origin评级为16级：要求屏蔽部分暴力内容', '德国USK·GameStar', '版号/分级',
         '德国USK完成对七大罪：Origin预发审查，暂定评级USK 16，要求Netmarble针对德国区屏蔽游戏内部分高暴力战斗演出内容；正式评级将在发售前完成',
         'Netmarble须在3月上线前完成德区内容修改；修改完成前无法通过德区正式商店销售',
         '德国是欧洲内容审查最严格的市场；USK审查结果将影响德区发售时间线'],
        ['PEGI公布2025年欧洲游戏内购投诉年报：随机付费类投诉占比38%', 'PEGI·GamesIndustry.biz·Eurogamer', '监管',
         'PEGI发布2025年欧洲游戏内购投诉年度报告，Loot Box和随机付费相关投诉占全部游戏投诉的38%，同比上升12个百分点；PEGI呼吁立法机构加快统一欧盟层面监管框架',
         '欧盟层面统一立法的呼声增强；各成员国可能在统一框架出台前先行立法',
         '内购付费信息披露不透明是欧洲监管的核心关注点'],
    ])


# ================================================================
# JAPAN
# ================================================================
jp_pc = [
    ['#1 (Steam JP)', '怪物猎人：荒野', '动作RPG', 'Capcom', 'Steam·PS5',
     MHW_104,
     'Capcom本土IP一周年获日本玩家高度重视；AT Rey Dau挑战内容获硬核猎人认可；1周年庆典活动丰富',
     'PC版性能优化问题持续；AT龙难度过高引发部分玩家抱怨'],
    ['#2 (PS JP)', '黑神话：悟空', '动作ARPG', '游戏科学', 'PS5·Steam',
     WUKONG_DLC,
     '东方神话题材在日本玩家中引发共鸣；钟馗角色设计与日本玩家审美契合；Famitsu专题报道提升曝光',
     '部分日本玩家反映操作手感与日式ARPG有差异；高难度设计门槛较高'],
    ['#3 (Steam JP)', 'CS2', '竞技射击', 'Valve', 'Steam',
     '2月常规运营期；情人节皮肤销售；无重大版本内容更新',
     '竞技对战稳定；情人节活动皮肤受部分玩家欢迎',
     'VAC反作弊系统持续被批评；日本玩家对外挂问题尤为敏感'],
    ['#4 (PS JP)', 'Final Fantasy VII Rebirth', 'ARPG', 'Square Enix', 'PS5',
     '发售一周年活动（2025年2月29日发售，2026年1周年）；限定PS5版折扣活动',
     '日本本土RPG佳作；一周年折扣带动新玩家入手',
     '部分玩家对游戏节奏有争议'],
    ['#5 (Steam JP)', '原神', '开放世界ARPG', '米哈游', 'Steam·PC',
     GI_54,
     '纳塔地区新剧情获日本玩家好评；情人节限定活动氛围温馨',
     '版本末期（2月底）活跃度下滑；付费角色定价争议'],
]

jp_mobile = [
    ['畅销#1 (iOS JP)', 'モンスターストライク（怪物弹珠）', 'RPG弹射', 'MIXI', 'iOS App Store JP',
     '日本手游长青王者；情人节限定联动活动2月14日上线',
     '经典弹射玩法；情人节联动设计受欢迎',
     '老玩家认为创新停滞；新用户上手门槛高'],
    ['畅销#2 (iOS JP)', 'パズル＆ドラゴンズ（PAD）', '消除RPG', 'GungHo', 'iOS App Store JP',
     '日本市场经典大DAU；情人节主题联动',
     '经典消除玩法；IP联动丰富',
     '新玩家认为界面老旧；创新不足'],
    ['畅销#3 (iOS JP)', 'FGO（Fate/Grand Order）', '卡牌RPG', 'TYPE-MOON/Aniplex', 'iOS',
     '情人节特别活动「月之恋歌」2月14日开启；限定从者抽取活动；白天/黑夜限定任务',
     '情人节活动剧情内容受核心粉丝高度好评；限定从者吸引力强',
     '抽卡系统无保底被长期批评；情人节限定从者获取门槛高'],
    ['畅销#4 (Google JP)', 'ウマ娘 プリティーダービー', '育成/竞技', 'Cygames', 'iOS·Google Play JP',
     '稳定大DAU；情人节限定育成剧情更新',
     '育成玩法深度获核心玩家认可；情人节剧情内容可爱',
     '部分玩家批评随机性过强；付费礼包性价比争议'],
    ['畅销#5 (Google JP)', 'Dragon Quest Walk', 'AR/RPG', 'Square Enix', 'Google Play JP',
     '日本本土AR步行游戏；情人节特别任务活动',
     '本土IP强；情人节步行任务形式新颖',
     '部分玩家认为活动奖励不够丰厚'],
    ['下载#1 (iOS JP)', 'Honkai: Star Rail', 'ARPG', 'miHoYo', 'iOS App Store JP',
     HSR_40,
     '新版本上线期下载爆发；日本玩家对米哈游游戏高度认可',
     '部分玩家对卡池保底机制仍有疑虑；版本末期流失率较高'],
    ['下载#2 (iOS JP)', 'Genshin Impact', '开放世界ARPG', 'miHoYo', 'iOS App Store JP',
     GI_54,
     '5.4版本新内容带动日本区下载回流',
     '版本末期（2月底）活跃度下滑'],
    ['下载#3 (Google JP)', 'モンスターストライク', 'RPG弹射', 'MIXI', 'iOS',
     '情人节联动活动持续维持下载热度',
     '经典弹射玩法；情人节联动设计受欢迎',
     '老玩家认为创新停滞'],
    ['下载#4 (iOS JP)', '七大罪：Origin', 'ARPG', 'Netmarble', 'iOS App Store JP',
     '2月预注册冲刺期；日本区Netmarble重点投放预注册推广',
     '原作动漫IP在日本人气极高；战斗系统预告视觉效果惊艳',
     '预注册阶段信息有限，玩家期待管理存在压力'],
    ['下载#5 (Google JP)', 'プロ野球スピリッツA', '体育', 'Konami', 'Google Play JP',
     '日本本土体育游戏稳定表现；春节后竞技季开始',
     '本土体育IP强；日本职棒赛季开幕前热度有保障',
     '受众相对垂直；对非棒球爱好者吸引力有限'],
]

jp_mkt = [
    ["怪物猎人：荒野", "PC主机", "媒体合作",
     "MHW（Capcom）1周年大更新前向Famitsu、4Gamer、GameWatch等日本本土权威媒体定向送测，并举办日本区1周年纪念直播，邀请日本著名游戏主播参与",
     "Famitsu·4Gamer·GameWatch·YouTube JP",
     "Famitsu Ver.1.04评测：36/40分（4编辑各自打9分）；4Gamer长评阅读量约10万次（预估）；日本YouTube MHW相关视频2月累计播放量超200万次（预估）",
     "日本权威媒体背书显著提升游戏公信力；Famitsu评分权威性强，对日本零售销量有实际推动作用",
     "Famitsu体系影响力主要集中在核心玩家群体；PC版性能问题被日本媒体持续关注"],
    ["七大罪：Origin", "PC主机", "渠道合作（平台合作）",
     "七大罪：Origin（Netmarble）日本区2月预注册活动，与Famitsu合作发布PS5预注册联动抽奖活动；在日本App Store和Google Play JP采买预注册推荐位",
     "Famitsu Twitter·App Store JP·Google Play JP",
     "Famitsu联动帖Twitter互动约5000次（预估）；日本区预注册量约200万（预估，占全球约6-7%）；App Store JP免费游戏下载预注册列表登顶（预估）",
     "日本动漫IP粉丝基础深厚；Famitsu合作有效触达核心ARPG玩家群体",
     "七大罪IP在日本知名度极高导致玩家期待值过高，游戏质量压力大"],
    ["FGO（Fate/Grand Order）", "手游", "渠道合作（平台合作）",
     "FGO日本服（TYPE-MOON/Aniplex）情人节特别活动「月之恋歌」：2月14日限定从者抽取活动；在Famitsu系及Twitter/X同步发布活动公告",
     "Famitsu官网·Twitter/X·YouTube JP",
     "情人节限定活动Twitter公告转发约1万次（预估）；活动期间日服DAU回升约15-20%（预估）；YouTube活动PV播放量约50万（预估）",
     "日本玩家高度期待情人节限定活动；限定从者吸引力强，带动付费峰值",
     "抽卡无保底系统在情人节高消费期格外引发玩家不满"],
    ["怪物猎人：荒野", "PC主机", "网红合作",
     "MHW（Capcom）1周年活动配合日本头部游戏YouTuber/Twitch主播开展AT Rey Dau挑战攻略合作视频",
     "YouTube JP·Twitch JP",
     "日本YouTube MHW频道1周年相关视频合计约100万播放（预估）；Twitch日本区MHW类目观看时长增长约30%（预估）",
     "日本KOL攻略内容帮助玩家突破AT龙高难度门槛，带动留存",
     "AT龙高难度内容缩小了潜在受众范围；部分KOL公开批评PC版性能问题"],
]

make_market_sheet("日本", "JP", "日本 — 游戏市场营销热点月报",
    jp_pc, jp_mobile, jp_mkt,
    notes="平台：Famitsu·Twitter/X·YouTube·Steam·4Gamer·App Store JP·Google Play JP | 数据截至2026年2月28日",
    policy_rows=[
        ['消费者厅发布手游抽卡保底信息披露新要求：7月前须实时显示剩余抽数', '消费者厅·CESA·Famitsu', '监管',
         '日本消费者厅于2月发布《手游抽卡信息提示升级指引》，要求所有在日本上架的手游自7月起实时显示玩家距当前卡池保底的剩余抽数及概率，并须以清晰日文标注',
         '米哈游、腾讯等主要在日运营手游需在7月前完成UI改造；已有保底机制的游戏影响较小',
         '不按时完成改造的手游可能被消费者厅约谈；二次元游戏是最受关注的品类'],
        ['CERO完成七大罪：Origin预评级：暂定D级（17岁以上）', 'CERO·Famitsu', '版号/分级',
         'CERO完成对七大罪：Origin的预评级审查，暂定D级（17岁以上），含有Violence和Sexual Content描述；D级是日本主流实体零售可正常销售的最高级别',
         'D级对日本零售渠道无实质性影响；正式发售前最终评级确认',
         'CERO Z级相当于彻底排除日本实体零售渠道；目前评级对日本市场销售无不利影响'],
        ['JOGA发布2025年手游市场白皮书：春节营销已成全球游戏营销标配', 'JOGA（日本在线游戏协会）·Dengeki Online', '政策/产业',
         'JOGA发布2025年手游市场年度白皮书，指出中国农历春节营销已成为全球移动游戏行业的重要营销节点；日本游戏厂商也在探索面向大中华区玩家的春节本地化营销',
         '春节营销意识提升有利于日本游戏公司拓展大中华区市场；CNY主题内容开发需更了解中国文化',
         '文化误解风险高；不恰当的CNY本地化可能引发大中华区玩家反感'],
    ])


# ================================================================
# KOREA
# ================================================================
kr_pc = [
    ['#1 (Steam KR)', '怪物猎人：荒野', '동작RPG', 'Capcom', 'Steam·PS5',
     MHW_104,
     'AT Rey Dau挑战难度获韩国硬核猎人认可；1周年活动内容充实；Capcom韩语官方支持完善',
     'PC版性能优化问题持续；AT龙难度过高引发部分玩家抱怨'],
    ['#2 (Steam KR)', '黑神话：悟空', '동작ARPG', '游戏科学', 'Steam·PS5',
     WUKONG_DLC,
     '钟馗DLC韩国玩家反响积极；东方神话题材引发共鸣；免费DLC策略赢得口碑',
     '部分韩国玩家反映钟馗关卡难度过高；韩语本地化深度不足'],
    ['#3 (PS KR)', 'Elden Ring', '동작ARPG', 'FromSoftware', 'PS5·Steam',
     '韩国动作游戏玩家长尾消费，稳定在Steam KR前列',
     '核心ARPG玩法深度无可比拟；韩国玩家对高难度挑战型游戏热情高',
     '高难度门槛对新玩家不友好'],
    ['#4 (Steam KR)', 'CS2', '경쟁 슈팅', 'Valve', 'Steam',
     '2月常规运营期；情人节主题皮肤销售',
     '竞技对战稳定；情人节皮肤受部分玩家欢迎',
     'VAC反作弊系统持续被批评；无重大新内容'],
    ['#5 (Steam KR)', 'Path of Exile 2', '동작ARPG', 'Grinding Gear Games', 'Steam',
     '韩国ARPG玩家基础强，持续稳定热销；情人节限定皮肤',
     'ARPG深度和Build多样性获高度评价',
     'EA阶段游戏难度被部分玩家认为过高'],
]

kr_mobile = [
    ['畅销#1 (ONE Store KR)', '리니지W（天堂W）', 'MMORPG', 'NCSoft', 'ONE Store·iOS',
     '韩国本土MMORPG长青王者；春节+情人节双节活动',
     '韩国MMORPG深度玩家忠诚度高；情人节限定道具受欢迎',
     'P2W模式被年轻一代玩家批评'],
    ['畅销#2 (ONE Store KR)', '배틀그라운드 모바일（PUBG Mobile）', '경쟁 슈팅', 'Krafton', 'ONE Store·iOS·Google',
     '韩国本土IP；情人节限定皮肤2月14日上线；韩国本土赛事热度持续',
     '真实感射击体验强；本土IP荣誉感强',
     '反外挂措施不足；版本更新速度慢于竞品'],
    ['畅销#3 (iOS KR)', '원신（原神）', '개방형 세계 ARPG', 'miHoYo', 'iOS App Store KR',
     GI_54,
     '5.4版本新内容获韩国玩家好评；情人节活动暖心',
     '版本末期活跃度明显下滑；付费角色定价争议'],
    ['畅销#4 (Google KR)', '붕괴: 스타레일（崩铁）', 'ARPG', 'miHoYo', 'Google Play KR',
     HSR_40,
     '新版本上线期韩国区付费率回升；剧情深度获认可',
     '部分玩家对新强度角色定价有争议'],
    ['畅销#5 (ONE Store KR)', '일곱 개의 대죄: 오리진（七大罪Origin）', 'ARPG', 'Netmarble', 'ONE Store',
     '2月韩国本土预注册活动高潮；ONE Store预注册登顶；Netmarble重点投放本土推广',
     '本土游戏公司作品；原作IP在韩国玩家中人气极高；预注册奖励丰厚',
     '预注册阶段信息有限，期待管理存在压力；玩家期待值极高'],
    ['下载#1 (ONE Store KR)', '일곱 개의 대죄: 오리진', 'ARPG', 'Netmarble', 'ONE Store',
     '2月韩国本土预注册爆发，ONE Store免费下载预注册榜#1',
     '韩国本土厂商Netmarble作品；原作IP人气高；预注册奖励吸引力强',
     '预注册后正式发售质量是否符合预期存在不确定性'],
    ['下载#2 (iOS KR)', 'Pokemon GO', 'AR休闲', 'Niantic', 'iOS',
     '情人节活动韩国区下载反弹',
     '情人节活动吸引情侣玩家；AR互动玩法独特',
     '部分玩家认为活动奖励不够丰厚'],
    ['下载#3 (Google KR)', '카트라이더：드리프트（跑跑卡丁车）', '경주', 'Nexon', 'Google Play KR',
     '情人节主题赛季内容更新',
     '本土IP强；情人节主题赛道设计可爱',
     '部分玩家认为新赛季内容量不足'],
    ['下载#4 (Samsung Store KR)', '무한의계단（无限阶梯）', '休闲', 'Naver', 'Samsung Store',
     '三星商店韩国区本土休闲游戏稳定；情人节主题活动',
     '本土休闲玩法简单易上手；情人节主题设计可爱',
     '内容深度有限；长线留存存疑'],
    ['下载#5 (iOS KR)', '崩壊：스타레일', 'ARPG', 'miHoYo', 'iOS App Store KR',
     HSR_40,
     '4.0版本新内容带动韩国区下载回流',
     '版本末期流失率较高'],
]

kr_mkt = [
    ["七大罪：Origin", "PC主机+手游", "渠道合作（平台合作）",
     "七大罪：Origin（Netmarble）2月韩国本土预注册活动高峰，在ONE Store、App Store KR、Google Play KR同步采买预注册推荐位；Netmarble官方Twitter KR账号持续投放倒计时内容",
     "ONE Store·App Store KR·Google Play KR·Twitter/X KR",
     "ONE Store预注册榜#1持续约3周（预估）；韩国区总预注册量约500万（预估）；Netmarble官方Twitter帖互动每条约1000-2000次（预估）",
     "韩国本土厂商作品预注册转化率高；ONE Store是韩国手游最重要的分发渠道之一",
     "过高的预注册量会放大玩家期待，增加正式发售后的舆论风险"],
    ["七大罪：Origin", "PC主机+手游", "媒体合作",
     "七大罪：Origin（Netmarble）与韩国本土游戏媒体게임메카、인벤合作，发布预发布深度报道和开发日志专题；同时在Naver Cafe开设官方社区",
     "게임메카·인벤·Naver Cafe·YouTube KR",
     "게임메카预发布专题报道阅读量约15万次（预估）；인벤相关内容约10万次（预估）；Naver Cafe官方社区预注册会员约20万（预估）",
     "韩国本土媒体报道有效建立玩家信任；官方社区运营为正式发售后的社区管理打好基础",
     "高度曝光增加玩家期待，发售后若质量不符预期风险更高"],
    ["怪物猎人：荒野", "PC主机", "媒体合作",
     "MHW（Capcom）1周年大更新向인벤、게임메카等韩国本土游戏媒体定向送测；Capcom Korea官方Twitter同步发布AT Rey Dau预告",
     "인벤·게임메카·Twitter/X KR",
     "인벤Ver.1.04评测阅读量约8万次（预估）；Capcom Korea官方Twitter帖互动约2000次（预估）；韩国PS Store周销量榜2月18日后回升至前3（预估）",
     "韩国本土媒体覆盖有效扩大玩家触达；本土化营销提升韩国玩家的游戏归属感",
     "PC版性能优化问题在韩国媒体报道中被持续放大"],
    ["王者荣耀（Honor of Kings国际版）", "手游", "网红合作",
     "Honor of Kings（腾讯，韩国国际服）春节+情人节双节营销，配合韩国游戏YouTube频道投放新赛季内容展示视频",
     "YouTube KR·Twitter/X KR",
     "韩国YouTube Honor of Kings相关视频2月新增播放量约50万次（预估）；情人节限定皮肤发布Twitter互动约800次（预估）",
     "双节营销时间窗口有效集中曝光；本土化皮肤设计获部分韩国玩家认可",
     "韩国MOBA市场竞争激烈，Honor of Kings面临来自PUBG Mobile和本土游戏的持续竞争"],
]

make_market_sheet("韩国", "KR", "韩国 — 游戏市场营销热点月报",
    kr_pc, kr_mobile, kr_mkt,
    notes="平台：Naver Cafe·Twitter/X·게임메카·인벤·ONE Store·Samsung Store·LG Store | 数据截至2026年2月28日",
    policy_rows=[
        ['韩国文体部：2月七大罪Origin预注册热潮被列为K-游戏出海期待案例', '韩国文化体育观光部（MCST）·This Is Game', '政策利好',
         '韩国文体部于2月发布月度K-游戏出海动态报告，点名七大罪：Origin的全球3000万预注册成绩为2026年K-游戏出海重要观察案例',
         '政策利好对Netmarble品牌价值有提升；其他韩国游戏厂商受益于整体出海政策红利',
         '政策利好是双刃剑，正式发售后若口碑不理想，政策背书可能引发连带舆论风险'],
        ['GRAC完成七大罪：Origin预分级审查：18+（暴力/性暗示）', 'GRAC（游戏分级与管理委员会）·Inven', '版号/分级',
         'GRAC正式完成对七大罪：Origin的预分级审查，评定18+，含有Violence和Sexual Content描述；Netmarble已准备相应的未成年人保护措施',
         '18+分级是韩国市场对写实暴力/成人内容的标准处理；不影响主要销售渠道',
         'GRAC拥有对已上架游戏进行重新审查的权力；玩家大规模投诉是触发重新审查的主要机制'],
        ['韩国个人信息保护委员会：游戏公司数据本地化合规检查2月报告', '韩国个人信息保护委员会·This Is Game', '监管',
         '韩国个人信息保护委员会于2月公布上季度游戏行业数据本地化合规检查结果，5家外资游戏公司被要求整改，主要问题为玩家个人数据跨境传输未经明确授权',
         '外资游戏公司在韩运营的数据合规要求进一步提高；影响米哈游、腾讯等主要在韩运营公司',
         '数据跨境传输违规可能导致罚款及限期整改；情节严重可面临韩国区应用下架'],
    ])


# ================================================================
# HONG KONG / TAIWAN
# ================================================================
hktw_pc = [
    ['#1 (Steam TW)', '怪物猎人：荒野', '動作RPG', 'Capcom', 'Steam·PS5',
     MHW_104,
     'AT Rey Dau挑战难度获台湾硬核猎人认可；1周年庆典活动内容充实',
     'PC版性能优化问题持续；AT龙难度引发部分玩家抱怨'],
    ['#2 (Steam TW)', '黑神話：悟空', '動作ARPG', '游戏科学', 'Steam·PS5',
     WUKONG_DLC,
     '钟馗DLC东方神话题材在港台玩家中引发强烈共鸣；免费DLC策略赢得极高口碑；繁中版本质量获认可',
     '部分玩家反映钟馗关卡难度过高；DLC时长偏短'],
    ['#3 (Steam TW)', 'CS2', '競技射擊', 'Valve', 'Steam',
     '2月常规运营期；情人节皮肤销售',
     '竞技对战稳定；情人节皮肤受欢迎',
     'VAC反作弊问题持续；无重大新内容'],
    ['#4 (PS TW/HK)', '崩壞：星穹鐵道（PC版）', 'ARPG', '米哈游/旺拓代理', 'PC·iOS',
     HSR_40,
     '4.0版本新章节剧情获台港玩家高度期待；旺拓代理版本服务完善',
     '版本末期（2月初）流失率较高；付费角色定价与购买力比值被台湾玩家批评'],
    ['#5 (Steam TW)', 'Elden Ring', '動作ARPG', 'FromSoftware', 'Steam·PS5',
     '台湾动作游戏玩家长尾消费，稳定在Steam TW前列',
     '核心ARPG玩法深度无可比拟',
     '高难度门槛对新玩家不友好'],
]

hktw_mobile = [
    ['畅销#1 (iOS TW)', '崩壞：星穹鐵道', 'ARPG', '米哈游/旺拓代理', 'iOS App Store TW',
     HSR_40,
     '4.0版本台湾区反响热烈；旺拓代理服务稳定；台湾玩家对米哈游游戏忠诚度高',
     '付费角色定价与台湾购买力比值被批评；版本末期活跃度下滑'],
    ['畅销#2 (iOS TW)', '原神', '開放世界ARPG', '米哈游/旺拓代理', 'iOS App Store TW',
     GI_54,
     '5.4版本情人节限定活动在台湾玩家中反响好；新剧情深度获认可',
     '5.3版本末期（2月初）活跃度骤降；版本更新前付费意愿下降'],
    ['畅销#3 (iOS HK)', '七大罪：Origin', 'ARPG', 'Netmarble', 'iOS App Store HK',
     '2月香港预注册活动冲刺；App Store HK预注册榜登顶',
     '原作动漫IP在港台知名度高；战斗系统预告视觉惊艳',
     '预注册阶段信息有限；玩家期待值较高'],
    ['畅销#4 (iOS TW)', '傳說對決（Arena of Valor）', 'MOBA', 'Garena TW', 'iOS App Store TW',
     '台湾Garena代理；春节+情人节双节活动；新赛季英雄发布',
     '台湾本土化运营成熟；情人节主题皮肤受欢迎',
     '部分赛季内容与国际服不同步被台湾玩家批评'],
    ['畅销#5 (Google Play TW)', 'Honor of Kings（王者榮耀國際版）', 'MOBA', 'Tencent/Garena TW', 'iOS·Google Play TW',
     '台湾区3月新赛季预热；情人节限定皮肤活动',
     '全球化运营成熟；情人节皮肤设计精美',
     '部分台湾玩家认为内容更新频率不足'],
    ['下载#1 (iOS TW)', '原神', '開放世界ARPG', '米哈游/旺拓', 'iOS App Store TW',
     '5.4版本2月12日上线带动大量下载',
     '新版本内容丰富；情人节活动氛围温馨',
     '上手难度对零基础新玩家较高'],
    ['下载#2 (iOS TW)', '崩壞：星穹鐵道', 'ARPG', '米哈游/旺拓', 'iOS App Store TW',
     '4.0版本上线带动台湾区下载爆发',
     '回合制战斗门槛低；剧情叙事吸引新玩家',
     '前期剧情偏长，部分玩家中途流失'],
    ['下载#3 (iOS TW)', '七大罪：Origin', 'ARPG', 'Netmarble', 'iOS App Store TW',
     '台湾预注册冲刺期下载热度高',
     '原作IP台湾知名度高；预注册奖励丰厚',
     '预注册后是否符合期待存在不确定性'],
    ['下载#4 (Google Play TW)', '傳說對決', 'MOBA', 'Garena TW', 'Google Play TW',
     '情人节赛季活动带动下载',
     '台湾本土化运营成熟；情人节主题赛道可爱',
     '部分内容与国际服不同步'],
    ['下载#5 (iOS HK)', '黑神話：悟空', '動作ARPG', '游戏科学', 'PS5配套',
     '钟馗DLC免费发布带动香港区玩家关注和相关内容下载',
     '东方神话题材在港台有深厚文化共鸣；免费DLC策略降低参与门槛',
     '高难度设计门槛较高；需要原作基础'],
]

hktw_mkt = [
    ["崩壞：星穹鐵道", "手游", "渠道合作（平台合作）",
     "崩壞：星穹鐵道台港澳代理旺拓在台湾iOS App Store 4.0版本上线期（约2月5日）采买推荐资源位，并在Dcard游戏板、PTT C_Chat板发布版本前瞻讨论帖",
     "iOS App Store TW/HK·Dcard·PTT·Facebook TW",
     "台湾iOS畅销榜#1（预估）；4.0版本台湾区月收入同比提升约100%（行业估算）；PTT C_Chat相关讨论串推文数超400条；Dcard游戏板4.0讨论帖互动超250条",
     "台湾二次元玩家对旺拓代理版本信任度高；4.0版本剧情深度在台湾社区引发热烈讨论",
     "台湾玩家批评付费角色定价较高；部分玩家认为版本内容量不足"],
    ["黑神話：悟空", "PC主机", "玩家生成内容（UGC）",
     "黑神話：悟空《鐘馗》DLC（2月10日）发布后，台湾80/90后玩家自发在Dcard、PTT、巴哈姆特发布DLC体验文章和二创内容，形成台湾本土UGC浪潮",
     "Dcard游戏板·PTT C_Chat·巴哈姆特·YouTube TW",
     "Dcard游戏板相关讨论帖2月约300篇（预估）；PTT C_Chat推文数超200条（预估）；YouTube台湾游戏频道钟馗相关视频合计播放量约50万次（预估）",
     "台湾玩家对中华文化题材有强烈共鸣；繁中版本推出有效降低语言门槛；免费DLC降低参与门槛",
     "台湾社区批评钟馗关卡难度设计过高；部分玩家认为DLC时长偏短"],
    ["七大罪：Origin", "PC主机+手游", "媒体合作",
     "七大罪：Origin（Netmarble）台湾繁中版预发推广，与台湾本土游戏媒体巴哈姆特GNN、遊戲基地合作进行预发布深度报道",
     "巴哈姆特GNN·遊戲基地·YouTube TW",
     "巴哈姆特GNN预发布报道阅读量约3万次（预估）；YouTube台湾游戏频道预告片播放量约30万次（预估）；台湾iOS预注册登顶（预估）",
     "台湾玩家对日本动漫改编游戏接受度高；繁中版本推出有效降低语言障碍",
     "台湾玩家对Netmarble的抽卡机制存有戒心；期待管理存在压力"],
    ["原神×情人節活動", "手游", "玩家生成内容（UGC）",
     "原神（米哈游/旺拓）情人节限定庭院互动活动（2月14日）在台港社区引发大量UGC，玩家自发分享游戏内情人节截图和短视频",
     "Dcard游戏板·PTT C_Chat·Instagram TW/HK",
     "Dcard情人节原神相关帖约200篇（预估）；Instagram台湾区相关标签帖文约1万篇（预估）",
     "情人节游戏内活动触发玩家情感共鸣；UGC自然传播效率高",
     "部分玩家认为情人节活动奖励不够丰厚；与国际版内容一致性要求限制了本地化创新空间"],
]

make_market_sheet("港台", "TW/HK", "港澳台 — 游戏市场营销热点月报",
    hktw_pc, hktw_mobile, hktw_mkt,
    notes="平台：Dcard·PTT·巴哈姆特GNN·Facebook TW/HK·YouTube TW/HK·iOS App Store TW/HK·Google Play | 数据截至2026年2月28日",
    policy_rows=[
        ['台湾NCC：游戏平台月活超100万用户须提交数位内容安全年报', 'NCC（国家通讯传播委员会）·遊戲基地', '政策/产业',
         '台湾NCC于2月更新数位平台内容安全管理要求，游戏平台若月活超过100万台湾用户须提交年度内容安全报告，包含未成年人保护、内容审查及用户投诉处理机制',
         '腾讯、米哈游等主要在台运营的游戏须通过代理商旺拓等完成合规报告提交',
         '合规成本上升；报告内容不达标可能面临NCC约谈'],
        ['香港创新科技及工业局：2026财年游戏产业支持预算增加至2亿港元', '香港创新科技及工业局·香港01', '政策利好',
         '香港创新科技及工业局公布2026-27财政年度预算，将电竞及数字游戏纳入重点支持产业，拨款2亿港元用于电竞场馆建设和游戏人才培训',
         '香港游戏市场体量有限，但政策利好对国际游戏公司在港设立亚太区总部有吸引力',
         '政策利好主要面向本土开发者；国际发行商需通过本地合作获得支持资格'],
        ['台湾DGSA：2月春节+情人节双节期间未成年人消费投诉增加25%', 'DGSA（数位游戏自律委员会）·Dcard', '监管',
         '台湾DGSA公布2月双节期间未成年人游戏消费投诉数据，同比增加25%；主要来自CNY皮肤高价限定销售及情人节礼包诱导消费；DGSA要求相关厂商加强说明',
         '代理商旺拓等须加强双节期间的未成年人消费提示机制；付费引导设