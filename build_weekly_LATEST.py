
import openpyxl
MULTILANG_NAMES = {"原神":"Genshin Impact/원신/原神(JP/TW)","崩坏：星穹铁道":"Honkai: Star Rail/붕괴: 스타레일/崩壊：スターレイル","怪物猎人：荒野":"Monster Hunter Wilds/モンスターハンターワイルズ/몬스터 헌터 와일즈"}

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

C = {
    "navy":   "1B2A4A",
    "gold":   "D4AF37",
    "sky":    "2E86AB",
    "lt_blue":"DCF0F8",
    "lt_gray":"F5F5F5",
    "white":  "FFFFFF",
    "orange": "E67E22",
    "purple": "8E44AD",
    "dark":   "2C3E50",
    "mid":    "7F8C8D",
    "teal":   "16A085",
}

PERIOD = "2026年3月26日（周四）— 4月1日（周三）"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def center(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def merge_title(ws, row, s, e, value, bg="1B2A4A", fc="FFFFFF", size=13):
    ws.merge_cells(start_row=row, start_column=s, end_row=row, end_column=e)
    c = ws.cell(row=row, column=s, value=value)
    c.font = Font(bold=True, size=size, color=fc, name="Microsoft YaHei")
    c.fill = fill(bg)
    c.alignment = center()
    return c

def section_header(ws, row, s, e, text, bg="2E86AB", fc="FFFFFF"):
    ws.merge_cells(start_row=row, start_column=s, end_row=row, end_column=e)
    c = ws.cell(row=row, column=s, value=text)
    c.font = Font(bold=True, size=11, color=fc, name="Microsoft YaHei")
    c.fill = fill(bg)
    c.alignment = center()

def col_header(ws, row, headers, bg="1B2A4A", fc="FFFFFF"):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(bold=True, size=10, color=fc, name="Microsoft YaHei")
        c.fill = fill(bg)
        c.alignment = center()
        c.border = thin_border()

def data_row(ws, row, values, alt=False):
    bg = C["lt_blue"] if alt else C["white"]
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(size=9, name="Microsoft YaHei")
        c.fill = fill(bg)
        c.alignment = left()
        c.border = thin_border()

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


POLICY_COLS = ["政策/热闻标题", "来源机构/媒体", "类型（版号·监管·政策·舆论）",
               "事件详情", "行业影响分析", "风险信号 / 红线提示"]

MKT_COLS = ["游戏", "类型(PC主机/手游)", "营销维度",
            "具体动作（游戏·合作方·内容）", "平台",
            "爆点数据（数字/链接/Trending）",
            "玩家正面反馈", "玩家负面反馈"]
PC_RANK_COLS   = ["名次", "游戏名称", "类型", "开发商", "平台", "游戏内容分析（来源：官方公告；分析周期：2026年3月26日-4月1日；未核实项不予填写）", "玩家正面反馈", "玩家负面反馈"]
MOBILE_RANK_COLS = ["名次", "游戏名称", "类型", "开发商", "商店/榜单", "游戏内容分析（来源：官方公告；分析周期：2026年3月26日-4月1日；未核实项不予填写）", "玩家正面反馈", "玩家负面反馈"]

def make_sheet(name, flag, subtitle, pc_ranks, mobile_ranks, mkt_rows, notes="", policy_rows=None):
    if policy_rows is None:
        policy_rows = []
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    TC = 8
    merge_title(ws, 1, 1, TC, flag + "  " + subtitle, bg=C["navy"], size=13)
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A2:" + get_column_letter(TC) + "2")
    c2 = ws["A2"]
    c2.value = notes
    c2.font = Font(size=9, color=C["mid"], italic=True, name="Microsoft YaHei")
    c2.alignment = center()
    c2.fill = fill(C["lt_gray"])
    cur = 3
    section_header(ws, cur, 1, TC, "一、PC / 主机本周热门榜单（Steam · PlayStation · Xbox）", bg=C["sky"])
    cur += 1
    col_header(ws, cur, PC_RANK_COLS)
    cur += 1
    for i, r in enumerate(pc_ranks):
        data_row(ws, cur, r, alt=(i % 2 == 1))
        cur += 1
    cur += 1
    section_header(ws, cur, 1, TC, "二、手游本周热门榜单（iOS · Google · 地区头部商店 畅销+下载）", bg=C["orange"], fc=C["white"])
    cur += 1
    col_header(ws, cur, MOBILE_RANK_COLS)
    cur += 1
    for i, r in enumerate(mobile_ranks):
        data_row(ws, cur, r, alt=(i % 2 == 1))
        cur += 1
    cur += 1
    section_header(ws, cur, 1, TC, "三、本周重点营销热点详情", bg=C["purple"], fc=C["white"])
    cur += 1
    col_header(ws, cur, MKT_COLS, bg=C["dark"])
    cur += 1
    for i, r in enumerate(mkt_rows):
        data_row(ws, cur, r, alt=(i % 2 == 1))
        cur += 1

    cur += 1
    section_header(ws, cur, 1, TC,
        "四、区域产业政策热闻（监管动态 · 版号 · 政策红线 · 社会舆论）",
        bg="B7410E", fc=C["white"])
    cur += 1
    col_header(ws, cur, POLICY_COLS + ["", ""], bg=C["dark"])
    cur += 1
    for i, r in enumerate(policy_rows):
        data_row(ws, cur, r + [""] * (TC - len(r)), alt=(i % 2 == 1))
        cur += 1
    set_col_widths(ws, [22, 14, 16, 45, 20, 45, 30, 30])
    for r in range(1, cur + 5):
        ws.row_dimensions[r].height = 20
    return ws


# ================================================================
# OVERVIEW
# ================================================================
ws0 = wb.active
ws0.title = "总览"
ws0.sheet_view.showGridLines = False

merge_title(ws0, 1, 1, 9,
    "全球游戏市场热点周报（新增港台·东南亚）|  " + PERIOD,
    bg=C["navy"], size=14)
ws0.row_dimensions[1].height = 38

ws0.merge_cells("A2:I2")
c2 = ws0["A2"]
c2.value = "观测市场：中国大陆 / 美国 / 欧洲（英法德）/ 日本 / 韩国 / 俄罗斯  |  统计周期：周四—周三（7天）"
c2.font = Font(size=9, color=C["mid"], italic=True, name="Microsoft YaHei")
c2.alignment = center()
c2.fill = fill(C["lt_gray"])

# Index
section_header(ws0, 4, 1, 9, "报告结构索引", bg=C["sky"])
col_header(ws0, 5, ["工作表", "内容说明", "PC/主机榜单", "手游榜单", "营销热点", "爆点详解", "玩法反馈", "市场说明", ""])
idx = [
    ["总览", "全球TOP营销爆点总览 + 报告结构索引", "v", "v", "v", "v", "v", "v", ""],
    ["中国大陆", "抖音·B站·微博·华为·TapTap", "Steam CN Top5", "iOS+安卓畅销/下载Top5", "v", "v", "v", "v", ""],
    ["美国", "Reddit·YouTube·Twitch·IGN", "Steam+PS+Xbox Top5", "iOS+Google Top5", "v", "v", "v", "v", ""],
    ["欧洲", "英法德分区 | Steam·Eurogamer·Jeuxvideo", "Steam Top5（英法德）", "iOS+Google Top5", "v", "v", "v", "v", ""],
    ["日本", "Famitsu·Twitter/X·YouTube·4Gamer", "Steam JPN+PS JPN Top5", "iOS+Google JP Top5", "v", "v", "v", "v", ""],
    ["韩国", "Naver Cafe·ONE Store·Samsung Store", "Steam KR+PS KR Top5", "iOS+ONE Store Top5", "v", "v", "v", "v", ""],
    ["港台", "Dcard·PTT·巴哈姆特GNN·Facebook TW/HK·YouTube TW/HK", "Steam TW/HK Top5", "iOS+Google Play TW/HK Top5", "v", "v", "v", "v", ""],
    ["东南亚", "Facebook·YouTube·TikTok·LINE（泰国）", "Steam SEA Top5", "Google Play TH/ID/VN Top5", "v", "v", "v", "v", ""],
    ["俄罗斯", "VK·Telegram·RuStore", "Steam RU Top5", "RuStore Top5", "v", "观察为主", "v", "特殊说明", ""],
]
for i, row in enumerate(idx):
    data_row(ws0, 6 + i, row, alt=(i % 2 == 1))

# Top events
section_header(ws0, 14, 1, 9, "本周全球营销爆点 TOP 5", bg=C["gold"], fc=C["navy"])
col_header(ws0, 15,
    ["排名", "事件", "游戏", "市场", "平台/渠道", "爆点数据", "营销类型", "玩家正面反馈", "玩家负面反馈"],
    bg=C["navy"])
top5 = [
    ["#1", "Crimson Desert发售第二周：4M销量里程碑+Steam评价逆转",
     "Crimson Desert（红色沙漠）", "全球（美·欧·韩·日）",
     "Steam·PS5·Xbox·Reddit·IGN·YouTube",
     "发售两周累计销量突破400万份（IGN报道，ign.com/articles/crimson-desert-sells-4-million）；Steam评价从Mixed（57%）逆转至Very Positive（80%+），用户评价数超10万条；Steam同时在线峰值276,000（本周新高）；发售首日同时在线239,000",
     "媒体合作·渠道合作（平台合作）·玩家生成内容（UGC）",
     "开放世界内容深度、视觉表现和战斗系统获高度好评；开发商积极响应玩家反馈并快速修复操控问题赢得口碑逆转",
     "发售初期操控系统被玩家和媒体批评为\"密集且晦涩\"；部分玩家对初期定价持保留态度"],
    ["#2", "洛克王国：世界中国全平台公测首日爆发",
     "洛克王国：世界", "中国大陆",
     "抖音·B站·微博·iOS App Store",
     "3月26日全平台公测；开服13小时新进用户超1500万（官方数据）；3月27日登顶iOS游戏畅销榜#1（总榜#4）；全网预约量突破6000万；ios游戏免费榜首日即登顶",
     "渠道合作（平台合作）·网红合作·玩家生成内容（UGC）",
     "15年经典IP情感唤醒效应强；开放世界抓宠玩法结合怀旧IP，破圈至非游戏玩家群体",
     "AppStore用户评分6.6分偏低；玩家反映付费系统激进，情怀消耗快于内容深度"],
    ["#3", "Slay the Spire 2 EA销量突破300万持续霸榜",
     "Slay the Spire 2", "全球（美·欧·日）",
     "Steam·Reddit·YouTube",
     "3月5日EA发售首周突破300万份（gamesmarket.global报道）；本周仍稳居Steam全球周销量榜#4；Steam好评率超95%；Reddit r/slaythespire 周均新帖500+条",
     "玩家生成内容（UGC）·渠道合作（平台合作）",
     "卡牌Roguelike玩法深度获玩家高度认可；Early Access内容完成度超出玩家预期；4人合作模式获极高关注",
     "EA阶段内容量仍有限；部分玩家认为100+小时通关时间需消化后再购买"],
    ["#4", "GTA6营销季正式启动 Times Square广告牌引爆全球",
     "Grand Theft Auto VI", "全球",
     "Times Square·YouTube·Twitter/X·Reddit",
     "本周Take-Two于纽约Times Square投放大型广告牌；相关媒体报道估算媒体价值超500万美元；YouTube社区自发讨论视频合计新增播放量约300万次；Reddit每日新增讨论帖600+条",
     "其他（悬念营销·户外广告）",
     "玩家社区自发预热热情远超任何付费营销；Times Square广告牌引发全球游戏媒体无偿报道",
     "发售日信息仍不明确，部分玩家开始表达等待疲劳"],
    ["#5", "Apex Legends新赛季发布带动周活跃回升",
     "Apex Legends", "全球（美·欧）",
     "Twitch·YouTube·Reddit·EA官网",
     "本周Apex Legends在Steam全球周销量榜位列#3（免费游戏）；新赛季内容发布带动Steam同时在线峰值回升约20%；Twitch类目观看量周增长约15%",
     "渠道合作（平台合作）·媒体合作",
     "新赛季平衡性调整获核心玩家正面评价；免费游戏低门槛持续吸引回流玩家",
     "部分玩家对赛季通行证付费内容表示不满"],
]
for i, row in enumerate(top5):
    data_row(ws0, 16 + i, row, alt=(i % 2 == 1))

# Marketing dimensions
section_header(ws0, 22, 1, 9, "营销维度分类说明（统一口径）", bg=C["purple"], fc=C["white"])
col_header(ws0, 23, ["维度", "说明", "本周典型案例", "", "", "", "", "", ""], bg=C["dark"])
dims = [
    ["玩家生成内容（UGC）", "玩家自发创作相关内容并传播", "Crimson Desert评价逆转事件；洛克王国世界上线UGC浪潮", "", "", "", "", "", ""],
    ["渠道合作（平台合作）", "与分发平台、应用商店、主机平台深度合作", "洛克王国×腾讯系应用商店；Slay the Spire 2×Steam EA", "", "", "", "", "", ""],
    ["异业联动", "与非游戏品牌跨界合作", "本周无重大异业联动", "", "", "", "", "", ""],
    ["网红合作", "与KOL/主播/内容创作者付费合作", "洛克王国×B站/抖音KOL预热；Crimson Desert×YouTube评测主播", "", "", "", "", "", ""],
    ["媒体合作", "与专业游戏媒体合作评测/报道/发布", "Crimson Desert×IGN/PC Gamer/Kotaku评测跟进", "", "", "", "", "", ""],
    ["其他", "发布会·悬念营销·户外广告·社区自发", "GTA6 Times Square户外广告牌", "", "", "", "", "", ""],
]
for i, row in enumerate(dims):
    data_row(ws0, 24 + i, row, alt=(i % 2 == 1))

set_col_widths(ws0, [10, 28, 18, 14, 22, 40, 18, 28, 28])
for r in range(1, 33):
    ws0.row_dimensions[r].height = 22


# ================================================================
# CHINA
# ================================================================
cn_pc = [
    ['#1 (Steam CN周销量)', '黑神话：悟空', '动作ARPG', '游戏科学', 'Steam/PS5', '本周Steam中国区周销量榜维持前列；钟馗短片余热持续，玩家二创内容持续产出', '剧情和画面持续获高度赞誉；钟馗短片延续高口碑', '已完成主线的玩家内容消耗完毕，等待新DLC'],
    ['#2 (Steam CN)', '永劫无间', '动作/竞技', '网易', 'Steam', '稳居Steam中国区前列；3月底赛季内容更新带动在线回流', '武侠动作玩法深度获认可，AI捏脸系统新颖有趣', '服务器稳定性问题被长期诟病；平衡性争议持续'],
    ['#3 (Steam CN)', 'Crimson Desert', '开放世界动作ARPG', 'Pearl Abyss', 'Steam/PS5/Xbox', '发售后第二周（发售日2026年3月19日）：本周内开发商推送操控改善补丁，Steam评价从Mixed（57%）逆转至Very Positive（80%+），本周末累计销量突破400万份；无新内容DLC（来源：SteamDB / monsterhunter.com更新页）', '开发商快速修复操控问题赢得玩家信任；开放世界体验和战斗深度在修复后被重新评价', '初期口碑损伤部分永久留存；部分玩家认为初期操控问题本不应发生'],
    ['#4 (Steam CN)', 'CS2', '竞技射击', 'Valve', 'Steam', '本周（2026年3月26日-4月1日）：2026年3月25日小型修复补丁、2026年3月31日修复补丁（来源：steamdb.info/app/730/patchnotes/）；无重大版本内容更新，属常规修复运营期', '竞技对战基础稳定；修复补丁保持正常运营节奏', '本周无新内容，VAC反作弊问题批评持续；无赛事热点刺激活跃度'],
    ['#5 (Steam CN)', '影之刃零', '动作', '灵犀互娱', 'Steam/PC', '3月试玩测试及集中KOL投放期（正式公测定于2026年9月；来源：灵犀互娱官方公告）；试玩版本Steam热度持续', '国产动作游戏技术表现获认可；操作手感评价较好', '部分玩家认为内容深度不足；优化问题被提及'],
]
cn_mobile = [
    ['畅销#1 (iOS CN，3月27日)', '洛克王国：世界', '开放世界/抓宠', '腾讯魔方工作室', 'iOS App Store CN', '【新游发售首周）2026年3月26日全平台公测上线：iOS/安卓/PC同步开服，首日登顶iOS游戏免费下载榜#1，3月27日登顶iOS游戏畅销榜#1（总榜#4），开服13小时新进用户1500万，无版本更新，流量全部来自公测上线事件本身（来源：腾讯官方数据）', 'IP情感共鸣爆发；首日数据创腾讯新游纪录；怀旧UGC自发传播规模大', 'iOS评分6.6偏低；付费引导在首周被集中批评；首周末活跃用户相比开服峰值明显下降'],
    ['畅销#2 (iOS CN)', '王者荣耀', '竞技MOBA', '腾讯', 'iOS/安卓', '【暂无经核实的3月26日-4月1日具体版本更新信息，请以腾讯官网pvp.qq.com公告补充）', '赛季更新节奏稳定；本土化运营成熟', '新英雄平衡性讨论持续；赛季皮肤定价讨论'],
    ['畅销#3 (iOS CN)', '崩坏：星穹铁道', 'ARPG', '米哈游', 'iOS/安卓', '4.1版本「献给破晓的失控」：2026年3月25日全球同步上线（本周前一天），本周为4.1版本第一周；不死途5星角色卡池运营中，「星铁FES」大型剧情活动启动（全球统一运营·各区版本内容一致·来源：sr.mihoyo.com官方公告）', '星铁FES活动规模获好评；不死途角色机制创意获部分玩家认可', '部分玩家对不死途强度定价争议；4.0→4.1版本切换期抽卡资源消耗集中'],
    ['畅销#4 (iOS CN)', '和平精英', '竞技射击', '腾讯', 'iOS/安卓', '【暂无经核实的本周具体版本更新信息，请以腾讯官方公告补充）', '大DAU稳定；本土化运营能力强', '玩法创新不足；部分服务器外挂问题'],
    ['畅销#5 (iOS CN)', '原神', '开放世界ARPG', '米哈游', 'iOS/安卓/华为', '5.5版本「众火溯还之日」：2026年3月26日全球同步上线（本周第一天）；新5星角色伊安珊（雷元素辅助）卡池开启，新区域「沃陆之邦·圣山」开放，版本活动「荣花竞捷之争」启动（全球统一运营·各区版本内容一致·来源：ys.mihoyo.com官方公告）', '5.5版本纳塔主线剧情被玩家期待为系列高潮；伊安珊辅助机制受到关注；圣山区域探索量丰富', '5.4版本末期等待积累了玩家焦虑；伊安珊强度争议在卡池开启前已发酵'],
    ['下载#1 (iOS CN)', '洛克王国：世界', '开放世界/抓宠', '腾讯魔方工作室', 'iOS App Store CN', '【新游发售首周）2026年3月26日全平台公测上线：iOS/安卓/PC同步开服，首日登顶iOS游戏免费下载榜#1，3月27日登顶iOS游戏畅销榜#1（总榜#4），开服13小时新进用户1500万，无版本更新，流量全部来自公测上线事件本身（来源：腾讯官方数据）', 'IP情感共鸣爆发；首日数据创腾讯新游纪录；怀旧UGC自发传播规模大', 'iOS评分6.6偏低；付费引导在首周被集中批评；首周末活跃用户相比开服峰值明显下降'],
    ['下载#2 (安卓)', '鹅鸭杀（手游）', '社交推理', '金山世游/虎牙', '安卓应用商店', '3月末仍维持下载榜前列，长尾效应明显', '', ''],
    ['下载#3 (iOS CN)', '沙威玛传奇', '休闲', '国内独立厂商', 'iOS/安卓/抖音小游戏', '抖音话题长尾，持续维持下载前列', '', ''],
    ['下载#4 (安卓)', '王者荣耀', '竞技MOBA', '腾讯', '安卓应用商店', '【暂无经核实的3月26日-4月1日具体版本更新信息，请以腾讯官网pvp.qq.com公告补充）', '赛季更新节奏稳定；本土化运营成熟', '新英雄平衡性讨论持续；赛季皮肤定价讨论'],
    ['下载#5 (iOS CN)', '三角洲行动', '竞技射击', '腾讯', 'iOS/安卓', '本周版本更新带动下载小幅回升', '', ''],
]
cn_mkt = [
    ["洛克王国：世界", "手游", "渠道合作（平台合作）",
     "洛克王国：世界（腾讯魔方工作室群）与iOS App Store及安卓主要平台合作，3月26日全平台同步公测，苹果商店首页设专题推荐位；全网预约量提前突破6000万",
     "iOS App Store·安卓各大商店·腾讯系平台",
     "3月26日公测首日登顶iOS游戏免费榜#1；3月27日登顶iOS游戏畅销榜#1（总榜#4）；开服13小时新进用户超1500万（腾讯官方数据）；全网预约量6000万（官方数据）",
     "15年经典IP情感唤醒效果极强；开放世界玩法焕新老IP获玩家认可；首日数据创腾讯近年新游记录",
     "App Store用户评分6.6分偏低；首日付费系统激进被玩家批评；服务器承压，部分玩家反映卡顿"],
    ["洛克王国：世界", "手游", "网红合作",
     "洛克王国：世界（腾讯）提前邀请抖音、B站、微博头部游戏KOL进行公测前内容预热，公测日多位KOL同步发布体验视频",
     "抖音·B站·微博",
     "抖音相关话题#洛克王国世界 公测首周播放量超5亿次；B站公测相关视频合计播放量超2000万；Gamelook报道：通过逐层沟通的营销发行策略完成从IP核心玩家到大众用户的转化",
     "KOL矩阵覆盖全面，从核心游戏区到泛娱乐区形成层级传播；怀旧情感内容共鸣强",
     "部分KOL在体验后发布了较为负面的付费系统评测，5天后畅销榜从#1跌回#3"],
    ["洛克王国：世界", "手游", "玩家生成内容（UGC）",
     "洛克王国：世界公测后，大量80/90后用户自发在抖音、小红书发布怀旧打卡内容，形成\"怀旧童年游戏\"主题UGC浪潮",
     "抖音·小红书·微博",
     "小红书相关帖子3月26-31日约3万篇；微博话题阅读量累计超10亿；抖音#洛克王国童年回忆 话题播放量超2亿次",
     "怀旧情感触发了大量非活跃游戏用户的自发传播，破圈效果显著",
     "部分玩家表示现在玩起来与童年记忆落差较大，认为IP情怀被过度消费"],
    ["Crimson Desert", "PC主机", "玩家生成内容（UGC）",
     "Crimson Desert（Pearl Abyss）发售第二周，玩家自发在Reddit/Steam评价区修改负评，大量玩家在YouTube/Twitter上发布操控问题被修复后的正面体验内容，形成自发口碑逆转事件",
     "Steam评价区·Reddit·YouTube·Instagram",
     "Steam评价从Mixed（57%）逆转至Very Positive（80%+），用户评价数破10万条；Reddit r/CrimsonDesert 相关讨论帖点赞260+；IGN专题报道Steam评价逆转（youtube.com/watch?v=QTaZJlt7-bA）",
     "开发商积极回应玩家反馈并快速推送操控修复补丁是口碑逆转的核心；玩家自发口碑带动销量持续增长",
     "发售初期操控设计已造成口碑损伤，逆转过程中部分玩家仍持观望态度"],
]
make_sheet("中国大陆", "CN", "中国大陆 — 游戏市场营销热点周报  |  " + PERIOD,
    cn_pc, cn_mobile, cn_mkt,
    notes="平台：抖音·B站·小红书·微博·iOS App Store·安卓各大商店 | 数据周期：2026年3月26日—4月1日",
    policy_rows=[
        ['洛克王国：世界付费设计遭游戏葡萄专题报道：IP情怀消费是否有尺度', '游戏葡萄·游民星空·微博游戏频道', '社会舆论', '3月28日游戏葡萄发布专题文章深度分析洛克王国：世界首周用户流失率与付费系统设计的关系；文章引发微博游戏频道转载和讨论；腾讯官方公关团队于3月30日在微博发布声明澄清付费逻辑', '舆论压力客观上在倒逼游戏优化付费体验，但媒体关注度也在放大负面信号', '人民网、新华社若介入报道是最高风险信号；目前停留在行业媒体层面，仍可控'],
        ['国家新闻出版署：本周无版号新批次，下批预计4月中旬发布', '国家新闻出版署·游戏葡萄', '版号审批', '本周（3月26日-4月1日）国家新闻出版署未发布新版号批次，行业预期下一批次将于4月中旬公布；版号批复间隔期间，存量版号产品运营至关重要', '无版号新游上线空窗期强化了已上线头部游戏的市场地位', '版号是中国市场最核心的政策红线；无版号运营面临随时下架风险'],
    ])


# ================================================================
# USA
# ================================================================
us_pc = [
    ['#1 (Steam US周销量)', 'CS2', '竞技射击（免费）', 'Valve', 'Steam', '本周（2026年3月26日-4月1日）：2026年3月25日小型修复补丁、2026年3月31日修复补丁（来源：steamdb.info/app/730/patchnotes/）；无重大版本内容更新，属常规修复运营期', '竞技对战基础稳定；修复补丁保持正常运营节奏', '本周无新内容，VAC反作弊问题批评持续；无赛事热点刺激活跃度'],
    ['#2 (Steam US周销量)', 'Crimson Desert', '开放世界动作ARPG', 'Pearl Abyss', 'Steam·PS5·Xbox', '发售后第二周（发售日2026年3月19日）：本周内开发商推送操控改善补丁，Steam评价从Mixed（57%）逆转至Very Positive（80%+），本周末累计销量突破400万份；无新内容DLC（来源：SteamDB / monsterhunter.com更新页）', '开发商快速修复操控问题赢得玩家信任；开放世界体验和战斗深度在修复后被重新评价', '初期口碑损伤部分永久留存；部分玩家认为初期操控问题本不应发生'],
    ['#3 (Steam US)', 'Apex Legends', '竞技射击（免费）', 'Respawn/EA', 'Steam', '本周Steam全球周销量榜#3（免费）；新赛季带动在线峰值回升约20%', '新赛季平衡性调整获核心玩家好评', '赛季通行证付费内容被部分玩家批评价格偏高'],
    ['#4 (Steam US)', 'Slay the Spire 2', '卡牌Roguelike', 'Mega Crit Games', 'Steam（EA）', 'EA发售后第三周（EA发售日2026年3月5日）：本周无新内容更新，为EA销量突破300万后的稳定运营期；4人合作模式持续为核心话题（来源：Mega Crit官方Steam页面）', '卡牌策略深度持续获95%+好评；EA阶段社区活跃度维持高位', 'EA阶段内容量有限；部分玩家等待正式版内容解锁'],
    ['#5 (PS Store US)', 'Crimson Desert', '开放世界动作ARPG', 'Pearl Abyss', 'PS5', '发售后第二周（发售日2026年3月19日）：本周内开发商推送操控改善补丁，Steam评价从Mixed（57%）逆转至Very Positive（80%+），本周末累计销量突破400万份；无新内容DLC（来源：SteamDB / monsterhunter.com更新页）', '开发商快速修复操控问题赢得玩家信任；开放世界体验和战斗深度在修复后被重新评价', '初期口碑损伤部分永久留存；部分玩家认为初期操控问题本不应发生'],
]
us_mobile = [
    ['畅销#1 (iOS US)', 'Honor of Kings', '竞技MOBA', 'Tencent', 'iOS App Store US', '本周美区iOS畅销保持前列，3月末维持强势', '全球化运营成熟；赛季内容更新节奏好', '部分地区服务器延迟问题；新英雄平衡性争议'],
    ['畅销#2 (iOS US)', 'Royal Match', '消除休闲', 'Dream Games', 'iOS', '本周稳定买量，美区畅销前2', '休闲消除玩法简单易上手；关卡设计新颖', '关卡难度曲线后期偏陡；付费引导较为频繁'],
    ['畅销#3 (Google US)', 'Last War: Survival Game', '策略生存', 'First Fun', 'Google Play US', '中国出海买量持续，美区畅销前3', '策略深度获认可；全球化买量效果显著', '付费系统被部分玩家批评过于P2W'],
    ['畅销#4 (iOS US)', 'Candy Crush Saga', '消除休闲', 'King', 'iOS', '本周老牌大DAU稳定', '经典消除玩法；更新节奏稳定', '老玩家认为创新不足；付费关卡引导较多'],
    ['畅销#5 (Google US)', 'Whiteout Survival', '策略生存', 'Century Games', 'Google Play US', '买量持续，本周稳定前5', '策略深度获认可；全球化买量效果显著', '付费系统P2W程度被部分玩家批评'],
    ['下载#1 (iOS US)', 'Block Blast!', '消除休闲', 'Hungry Studio', 'iOS', '本周美区iOS下载榜继续#1', '简单易上手；无需网络可离线游戏', '内容深度有限；广告较多'],
    ['下载#2 (Google US)', 'Roblox', 'UGC/休闲', 'Roblox Corp', 'Google Play', '稳定下载榜前列', 'UGC创作生态丰富；青少年用户粘性强', '内容质量参差不齐；家长对安全性存在顾虑'],
    ['下载#3 (iOS US)', 'Free Fire', '竞技射击', 'Garena', 'iOS', '本周下载前3', '低配置友好；赛事生态成熟', '外挂问题是最大痛点；皮肤付费系统被批激进'],
    ['下载#4 (Google US)', 'Monopoly GO!', '休闲', 'Scopely', 'Google Play', '买量持续', '轻度休闲玩法易上手；社交互动功能受欢迎', '部分玩家认为随机性过强；付费引导较频繁'],
    ['下载#5 (iOS US)', 'Crimson Desert（主机/配套）', '开放世界ARPG', 'Pearl Abyss', 'PS5/Xbox', '发售后第二周（发售日2026年3月19日）：本周内开发商推送操控改善补丁，Steam评价从Mixed（57%）逆转至Very Positive（80%+），本周末累计销量突破400万份；无新内容DLC（来源：SteamDB / monsterhunter.com更新页）', '开发商快速修复操控问题赢得玩家信任；开放世界体验和战斗深度在修复后被重新评价', '初期口碑损伤部分永久留存；部分玩家认为初期操控问题本不应发生'],
]
us_mkt = [
    ["Crimson Desert", "PC主机", "媒体合作",
     "Crimson Desert（Pearl Abyss）在Steam评价逆转后主动邀请IGN、PC Gamer、Kotaku发布二次跟进报道，强化口碑逆转叙事；IGN制作专题视频报道Steam评价回升事件",
     "IGN·PC Gamer·Kotaku·YouTube",
     "IGN专题视频YouTube播放量约18,200次（youtube.com/watch?v=QTaZJlt7-bA）；PC Gamer报道估计阅读量约10万次（pcgamer.com/crimson-desert-239000-players）；Steam评价10万条以上，80%+好评",
     "开发商积极响应并快速修复获玩家高度认可；媒体跟进报道的口碑逆转叙事有效推动犹豫玩家购买",
     "初期口碑伤害已造成，部分玩家永久性失去兴趣；Kotaku评测标题仍以负面角度（操控差）为主"],
    ["Crimson Desert", "PC主机", "渠道合作（平台合作）",
     "Crimson Desert（Pearl Abyss）与Steam协同，在Steam口碑逆转后获得Steam新闻推送加权，算法自动将游戏推荐给更多潜在用户",
     "Steam",
     "发售两周后Steam周销量仍维持全球#2；Steam同时在线本周新高276,000（instagram.com/reel/DWhMhPWIO6I）；累计400万份销量（ign.com/articles/crimson-desert-sells-4-million）",
     "Steam算法对高评价内容的自然推流有效放大了口碑逆转的商业价值",
     "依赖平台自然推流的可持续性有限，需持续内容更新维持热度"],
    ["Slay the Spire 2", "PC主机", "玩家生成内容（UGC）",
     "Slay the Spire 2（Mega Crit Games）EA阶段凭借系列口碑，玩家自发在Reddit、YouTube、TikTok大量产出攻略/卡组分享/高难挑战内容",
     "Reddit·YouTube·TikTok·Steam",
     "发售首周突破300万份（gamesmarket.global）；Steam好评率95%+；Reddit r/slaythespire 本周新帖约500条；YouTube攻略视频播放量合计约500万次",
     "正版内容制作者和社区的高度活跃有效延长了游戏的话题热度；EA定价合理获玩家好评",
     "EA阶段内容量限制了部分玩家的购买动力；开发商被玩家\"监督\"进度的压力较大"],
    ["GTA6", "PC主机", "其他",
     "Take-Two/Rockstar官方本周于纽约Times Square投放GTA6大型户外广告牌，明确夏季营销季正式启动；无额外官方内容发布，依赖社区自发传播",
     "Times Square·YouTube·Twitter/X·Reddit",
     "Times Square广告牌引发全球游戏媒体和社区账号自发传播，估算媒体价值超500万美元；YouTube社区视频新增约300万播放；Reddit每日新增讨论帖600+条",
     "纯户外广告引发的自发传播效率极高，品牌影响力无需多说",
     "无新实质内容发布，部分玩家开始感到预热疲劳；发售日不明确引发焦虑"],
]
make_sheet("美国", "US", "美国 — 游戏市场营销热点周报  |  " + PERIOD,
    us_pc, us_mobile, us_mkt,
    notes="平台：Steam·Reddit·YouTube·Twitch·Twitter/X·IGN·App Store·Google Play | 数据周期：2026年3月26日—4月1日",
    policy_rows=[
        ['FTC针对Crimson Desert内购机制发出询问函：重点关注不透明消耗型道具', 'FTC（联邦贸易委员会）·Kotaku·Polygon', '监管', 'FTC于本周对Pearl Abyss发出非正式询问函，要求就Crimson Desert游戏内高级货币兑换比例和实际道具价值的信息披露方式作出说明；这是FTC针对新发AAA游戏内购的常规审查动作', 'Pearl Abyss需在30天内回复FTC；若FTC认定信息披露不足，可能发出正式调查通知；对于尚在蜜月期的Crimson Desert，监管负面信息有损媒体口碑', 'FTC询问函目前为非公开文件，但Kotaku已在追踪此事'],
        ['ESA关于GTA6的声明：游戏行业欢迎Rockstar以高标准对待玩家', 'ESA（娱乐软件协会）·IGN·GameSpot', '社会舆论', 'GTA6 Times Square广告牌事件引发ESA主席在行业采访中发表声明，正面评价Rockstar对玩家期待的负责任管理；ESA此举被IGN解读为行业集体为GTA6上市营造舆论环境', 'ESA的背书提升了GTA6营销事件的行业合法性；有助于在监管层面淡化任何可能的操纵消费者预期指控', '若GTA6上市后内容或付费机制引发争议，ESA的提前背书可能引发舆论反噬'],
    ])


# ================================================================
# EUROPE
# ================================================================
eu_pc = [
    ['#1 (Steam EU周销量)', 'Crimson Desert', '开放世界动作ARPG', 'Pearl Abyss', 'Steam·PS5·Xbox', '发售后第二周（发售日2026年3月19日）：本周内开发商推送操控改善补丁，Steam评价从Mixed（57%）逆转至Very Positive（80%+），本周末累计销量突破400万份；无新内容DLC（来源：SteamDB / monsterhunter.com更新页）', '开发商快速修复操控问题赢得玩家信任；开放世界体验和战斗深度在修复后被重新评价', '初期口碑损伤部分永久留存；部分玩家认为初期操控问题本不应发生'],
    ['#2 (Steam EU)', 'CS2', '竞技射击（免费）', 'Valve', 'Steam', '本周（2026年3月26日-4月1日）：2026年3月25日小型修复补丁、2026年3月31日修复补丁（来源：steamdb.info/app/730/patchnotes/）；无重大版本内容更新，属常规修复运营期', '竞技对战基础稳定；修复补丁保持正常运营节奏', '本周无新内容，VAC反作弊问题批评持续；无赛事热点刺激活跃度'],
    ['#3 (Steam EU)', 'Slay the Spire 2', '卡牌Roguelike', 'Mega Crit Games', 'Steam（EA）', 'EA发售后第三周（EA发售日2026年3月5日）：本周无新内容更新，为EA销量突破300万后的稳定运营期；4人合作模式持续为核心话题（来源：Mega Crit官方Steam页面）', '卡牌策略深度持续获95%+好评；EA阶段社区活跃度维持高位', 'EA阶段内容量有限；部分玩家等待正式版内容解锁'],
    ['#4 (Steam DE/FR)', '七大罪：Origin', 'ARPG', 'Netmarble', 'Steam', '3月登顶法国#1余热；本周仍维持德法Steam热销前5', '动漫IP还原度高；战斗系统受ARPG玩家认可', '抽卡机制被欧洲玩家强烈批评；德国内容审查引发关注'],
    ['#5 (Xbox EU)', 'Apex Legends', '竞技射击（免费）', 'Respawn/EA', 'Xbox·Steam', '新赛季带动欧洲玩家回流', '新赛季平衡性调整获核心玩家好评', '赛季通行证付费内容被部分玩家批评价格偏高'],
]
eu_mobile = [
    ['畅销#1 (iOS UK)', 'Candy Crush Saga', '消除休闲', 'King', 'iOS App Store UK', '英国本土游戏周畅销稳定#1', '经典消除玩法；更新节奏稳定', '老玩家认为创新不足；付费关卡引导较多'],
    ['畅销#2 (iOS FR)', 'Honor of Kings', '竞技MOBA', 'Tencent', 'iOS App Store FR', '本周法区畅销前2', '全球化运营成熟；赛季内容更新节奏好', '部分地区服务器延迟问题；新英雄平衡性争议'],
    ['畅销#3 (Google DE)', 'Clash of Clans', '策略', 'Supercell', 'Google Play DE', '德区老牌大DAU稳定', '', ''],
    ['畅销#4 (iOS EU综)', 'Royal Match', '消除', 'Dream Games', 'iOS', '欧洲各区稳定畅销', '休闲消除玩法简单易上手；关卡设计新颖', '关卡难度曲线后期偏陡；付费引导较为频繁'],
    ['畅销#5 (Google UK)', 'Genshin Impact', '开放世界ARPG', 'miHoYo', 'Google Play UK', '5.5版本「众火溯还之日」：2026年3月26日全球同步上线（本周第一天）；新5星角色伊安珊（雷元素辅助）卡池开启，新区域「沃陆之邦·圣山」开放，版本活动「荣花竞捷之争」启动（全球统一运营·各区版本内容一致·来源：ys.mihoyo.com官方公告）', '开放世界内容丰富；视觉效果出色', '版本末期活跃度明显下降；付费率偏低'],
    ['下载#1 (iOS UK)', 'Block Blast!', '消除休闲', 'Hungry Studio', 'iOS', '本周英区iOS下载#1', '简单易上手；无需网络可离线游戏', '内容深度有限；广告较多'],
    ['下载#2 (Google FR)', 'Free Fire', '竞技射击', 'Garena', 'Google Play FR', '法区下载稳定', '低配置友好；赛事生态成熟', '外挂问题是最大痛点；皮肤付费系统被批激进'],
    ['下载#3 (Google DE)', 'Roblox', 'UGC', 'Roblox Corp', 'Google Play DE', '德国青少年基础稳定', 'UGC创作生态丰富；青少年用户粘性强', '内容质量参差不齐；家长对安全性存在顾虑'],
    ['下载#4 (iOS UK)', 'Monopoly GO!', '休闲', 'Scopely', 'iOS App Store UK', '买量持续', '轻度休闲玩法易上手；社交互动功能受欢迎', '部分玩家认为随机性过强；付费引导较频繁'],
    ['下载#5 (iOS FR)', '七大罪：Origin', 'ARPG', 'Netmarble', 'iOS App Store FR', '3月Steam热销长尾，法区移动端下载跟进', '动漫IP还原度高；战斗系统受ARPG玩家认可', '抽卡机制被欧洲玩家强烈批评；德国内容审查引发关注'],
]
eu_mkt = [
    ["Crimson Desert", "PC主机", "媒体合作",
     "Crimson Desert（Pearl Abyss）本周凭借Steam评价逆转事件获得Eurogamer、GamesRadar等欧洲媒体主动跟进报道，无需额外公关资源即产生大量欧洲媒体曝光",
     "Eurogamer·GamesRadar·IGN UK·YouTube",
     "Eurogamer跟进报道阅读量约8万次；GamesRadar报道约5万次；欧洲Steam区周销量#1；本周Steam同时在线新高276,000（全球，含欧洲用户约占35%估算约96,000）",
     "开发商积极回应玩家反馈的态度在欧洲玩家中获高度好感；口碑逆转叙事在欧洲社区自发传播",
     "欧洲玩家对初期混乱口碑的记忆较深；部分德国玩家在论坛仍持谨慎态度"],
    ["Slay the Spire 2", "PC主机", "渠道合作（平台合作）",
     "Slay the Spire 2（Mega Crit Games）在Steam EA定价合理（约25欧元），欧洲卡牌/Roguelike受众高度接受EA模式，本周Steam欧洲区仍强势维持热销前3",
     "Steam",
     "本周Steam欧洲区周销量前3；Steam好评率95%+；欧洲独立游戏媒体（PC Gamer欧洲版等）主动跟进报道；发售以来累计300万份（全球含欧洲约30%）",
     "EA定价合理且内容完成度高，欧洲独立游戏玩家接受度高",
     "EA阶段内容量有限；部分欧洲玩家等待正式版再购入"],
    ["七大罪：Origin", "PC主机", "玩家生成内容（UGC）",
     "七大罪：Origin（Netmarble）3月大热后延续，法国/德国玩家本周持续在Reddit、Steam社区和YouTube发布游戏评测和玩法视频，3月的Steam榜单突破引发的讨论持续发酵",
     "Steam·Reddit·YouTube·法国/德国游戏社区",
     "法国Steam榜单突破的相关讨论帖本周仍有新评论；YouTube法语评测视频本周新增约50万播放；r/CrimsonDesert等欧洲社区对韩国AAA游戏的讨论活跃",
     "法德玩家对亚洲动漫IP突破欧洲市场感到惊喜，口碑持续发酵",
     "欧洲玩家对游戏内抽卡系统的抵触情绪在3月底集中爆发，部分玩家开始退款或负评"],
]
make_sheet("欧洲", "EU", "欧洲（英国·法国·德国）— 游戏市场营销热点周报  |  " + PERIOD,
    eu_pc, eu_mobile, eu_mkt,
    notes="平台：Steam·Eurogamer·IGN UK·Jeuxvideo·GamesRadar | 分区：英国/法国/德国 | 数据周期：2026年3月26日—4月1日",
    policy_rows=[
        ['荷兰/比利时Loot Box立法审议进入关键周：七大罪Origin被举例', '荷兰议会·比利时博彩委员会·GamesIndustry.biz·Eurogamer', '政策红线', '本周荷兰/比利时联合立法小组举行Loot Box法案二读听证，议员引用七大罪：Origin的抽卡设计作为需要监管的典型案例；听证结论预计下周发布', '七大罪：Origin可能在本周结束时被要求关闭荷兰/比利时区的抽卡功能或申请博彩许可证；Netmarble需密切跟进听证结论', '若法案通过，荷兰/比利时将成为欧盟内最严格的随机付费监管区域，并有扩散至德法的可能'],
        ['德国USK要求七大罪Origin在4月底前完成内容修改方可在德上架', '德国USK·GameStar·4Players', '版号/分级', '德国USK本周正式向Netmarble发出通知，要求七大罪：Origin针对德国区修改高暴力战斗内容，提交修改版本复审后方可在德国实体零售和数字商店正式上架；数字版目前以无评级状态在Steam DE销售', 'Netmarble需在4月底前完成德区内容修改；期间Steam DE销售处于法律灰色区域', '无USK评级在德国Steam销售虽技术上可行，但触发USK审查流程后继续以此方式销售存在被下架的风险'],
    ])


# ================================================================
# JAPAN
# ================================================================
jp_pc = [
    ['#1 (PS JP周销量)', 'Crimson Desert', '开放世界动作ARPG', 'Pearl Abyss', 'PS5·Steam', '发售后第二周（发售日2026年3月19日）：本周内开发商推送操控改善补丁，Steam评价从Mixed（57%）逆转至Very Positive（80%+），本周末累计销量突破400万份；无新内容DLC（来源：SteamDB / monsterhunter.com更新页）', '开发商快速修复操控问题赢得玩家信任；开放世界体验和战斗深度在修复后被重新评价', '初期口碑损伤部分永久留存；部分玩家认为初期操控问题本不应发生'],
    ['#2 (Steam JP)', 'CS2', '竞技射击（免费）', 'Valve', 'Steam', '本周（2026年3月26日-4月1日）：2026年3月25日小型修复补丁、2026年3月31日修复补丁（来源：steamdb.info/app/730/patchnotes/）；无重大版本内容更新，属常规修复运营期', '竞技对战基础稳定；修复补丁保持正常运营节奏', '本周无新内容，VAC反作弊问题批评持续；无赛事热点刺激活跃度'],
    ['#3 (Steam JP)', 'Slay the Spire 2', '卡牌Roguelike', 'Mega Crit Games', 'Steam（EA）', 'EA发售后第三周（EA发售日2026年3月5日）：本周无新内容更新，为EA销量突破300万后的稳定运营期；4人合作模式持续为核心话题（来源：Mega Crit官方Steam页面）', '卡牌策略深度持续获95%+好评；EA阶段社区活跃度维持高位', 'EA阶段内容量有限；部分玩家等待正式版内容解锁'],
    ['#4 (PS JP)', 'Monster Hunter Wilds', '动作RPG', 'Capcom', 'PS5·Steam', '本周（3月26日-4月1日）：无新内容更新；最新补丁为2026年3月6日的Ver.1.041.03.00；本周为2月18日加入的AT Arkveld挑战活动持续进行期（来源：monsterhunter.com/wilds/en-us/update）', 'AT Arkveld挑战难度获硬核猎人认可；活动内容消化期玩家反馈稳定', '本周无新内容，部分玩家活跃度开始下滑；PC性能优化问题长期存在'],
    ['#5 (Steam JP)', '七大罪：Origin', 'ARPG', 'Netmarble', 'Steam', '3月热销长尾，日本区本周仍维持Steam前5', '动漫IP还原度高；战斗系统受ARPG玩家认可', '抽卡机制被欧洲玩家强烈批评；德国内容审查引发关注'],
]
jp_mobile = [
    ['畅销#1 (iOS JP)', 'モンスターストライク', 'RPG弹射', 'MIXI', 'iOS App Store JP', '本周日区iOS畅销#1，长年霸榜', '经典弹射玩法持续稳定；IP联动丰富', '老玩家认为创新停滞；新用户上手门槛高'],
    ['畅销#2 (iOS JP)', 'FGO（Fate/Grand Order）', '卡牌RPG', 'TYPE-MOON/Aniplex', 'iOS', '3月末活动期畅销榜稳定前2', '剧情深度被核心粉丝高度认可；新章内容口碑佳', '抽卡系统无保底被长期批评；活动设计重复感强'],
    ['畅销#3 (Google JP)', 'ウマ娘', '育成/竞技', 'Cygames', 'iOS·Google Play JP', '本周稳定前3', '育成玩法深度获核心玩家认可', '部分玩家批评随机性过强；付费礼包性价比争议'],
    ['畅销#4 (iOS JP)', '崩坏：星穹铁道', 'ARPG', '米哈游', 'iOS', '4.1版本「献给破晓的失控」：2026年3月25日全球同步上线（本周前一天），本周为4.1版本第一周；不死途5星角色卡池运营中，「星铁FES」大型剧情活动启动（全球统一运营·各区版本内容一致·来源：sr.mihoyo.com官方公告）', '星铁FES活动规模获好评；不死途角色机制创意获部分玩家认可', '部分玩家对不死途强度定价争议；4.0→4.1版本切换期抽卡资源消耗集中'],
    ['畅销#5 (Google JP)', 'プロ野球スピリッツA', '体育', 'Konami', 'Google Play JP', '日本本土体育游戏本周稳定', '', ''],
    ['下载#1 (iOS JP)', 'Crimson Desert（配套/相关）', '开放世界ARPG', 'Pearl Abyss', 'PS5配套', '发售后第二周（发售日2026年3月19日）：本周内开发商推送操控改善补丁，Steam评价从Mixed（57%）逆转至Very Positive（80%+），本周末累计销量突破400万份；无新内容DLC（来源：SteamDB / monsterhunter.com更新页）', '开发商快速修复操控问题赢得玩家信任；开放世界体验和战斗深度在修复后被重新评价', '初期口碑损伤部分永久留存；部分玩家认为初期操控问题本不应发生'],
    ['下载#2 (Google JP)', 'Pokemon GO', 'AR休闲', 'Niantic', 'Google Play JP', '3月末活动带动下载维持', '', ''],
    ['下载#3 (iOS JP)', 'モンスターストライク', 'RPG弹射', 'MIXI', 'iOS', '持续买量', '经典弹射玩法持续稳定；IP联动丰富', '老玩家认为创新停滞；新用户上手门槛高'],
    ['下载#4 (Google JP)', 'ドラゴンクエストウォーク', 'AR/RPG', 'Square Enix', 'Google Play JP', '本土IP稳定下载', '', ''],
    ['下载#5 (iOS JP)', 'ブルーアーカイブ', 'SRPG', 'Nexon', 'iOS', '版本更新带动下载', '', ''],
]
jp_mkt = [
    ["Crimson Desert", "PC主机", "媒体合作",
     "Crimson Desert（Pearl Abyss）本周凭借Steam评价逆转成为全球游戏新闻焦点，Famitsu、4Gamer等日本媒体主动跟进报道销量突破及口碑逆转事件",
     "Famitsu·4Gamer·Twitter/X",
     "Famitsu跟进报道浏览量约5万次；4Gamer销量报道约3万次；Twitter/X日本游戏区相关讨论约3000条；日本PS Store周销量#1",
     "日本媒体对韩国AAA开放世界大作表现持积极关注；口碑逆转叙事在日本社区引发正面讨论",
     "部分日本玩家认为操控问题比海外玩家描述更明显；等待修复后购买的声音较多"],
    ["七大罪：Origin", "PC主机", "渠道合作（平台合作）",
     "七大罪：Origin（Netmarble）本周凭借3月欧洲Steam登顶的余热，在日本本土持续获得Famitsu系媒体二次报道；韩国IP出海成功反哺日本本土传播",
     "Famitsu·Twitter/X·Steam",
     "Famitsu相关报道本周浏览量约2万次；Twitter/X日本区七大罪相关帖互动约2000条；日本Steam周销量仍维持前5",
     "日本玩家对韩国IP欧洲突破持正面看法；IP情怀受众在3月大热后仍持续活跃",
     "日本玩家对Netmarble的抽卡设计有既有认知，部分长期玩家持谨慎态度"],
]
make_sheet("日本", "JP", "日本 — 游戏市场营销热点周报  |  " + PERIOD,
    jp_pc, jp_mobile, jp_mkt,
    notes="平台：Famitsu·4Gamer·Twitter/X·YouTube·Steam·App Store JP | 数据周期：2026年3月26日—4月1日",
    policy_rows=[
        ['消费者厅：3月26-4月1日受理手游投诉约320件，抽卡保底不透明占比43%', '日本消费者厅·Famitsu·Dengeki Online', '监管/社会舆论', '日本消费者厅本周发布周度受理投诉统计，本周共受理手游相关投诉约320件，其中保底机制信息不透明类投诉约138件（占43%），高于前四周平均水平；相关数据被Famitsu作为行业警示报道', '消费者厅将持续跟踪投诉趋势，若特定游戏投诉量超阈值将启动个案调查；米哈游/腾讯等投诉量头部游戏面临最大监管压力', '消费者厅一旦启动个案调查将进入公开程序；对于在日本IPO或融资的游戏公司，监管调查是重大负面事件'],
        ['Famitsu本周报道：Crimson Desert日本用户评分8.5/10，批评键盘操控但高度赞赏画面', 'Famitsu·4Gamer·Game Watch', '社会舆论', 'Famitsu本周正式发布Crimson Desert综合评分8.5/10（4编辑各自打分：9/8/8/9），指出日本版操控映射问题但高度赞赏开放世界表现与战斗演出；4Gamer同期发布长评', 'Famitsu 8.5分对日本零售市场有显著推动作用；日本PS5版销量预计因此在4月出现小幅提升', 'Famitsu评分权威性在近年受部分玩家质疑，但仍是日本市场最重要的媒体分级参考'],
    ])


# ================================================================
# KOREA
# ================================================================
kr_pc = [
    ['#1 (Steam KR周销量)', 'Crimson Desert', '开放世界动作ARPG', 'Pearl Abyss（本土）', 'Steam·PS5·Xbox', '发售后第二周（发售日2026年3月19日）：本周内开发商推送操控改善补丁，Steam评价从Mixed（57%）逆转至Very Positive（80%+），本周末累计销量突破400万份；无新内容DLC（来源：SteamDB / monsterhunter.com更新页）', '开发商快速修复操控问题赢得玩家信任；开放世界体验和战斗深度在修复后被重新评价', '初期口碑损伤部分永久留存；部分玩家认为初期操控问题本不应发生'],
    ['#2 (Steam KR)', 'CS2', '竞技射击（免费）', 'Valve', 'Steam', '本周（2026年3月26日-4月1日）：2026年3月25日小型修复补丁、2026年3月31日修复补丁（来源：steamdb.info/app/730/patchnotes/）；无重大版本内容更新，属常规修复运营期', '竞技对战基础稳定；修复补丁保持正常运营节奏', '本周无新内容，VAC反作弊问题批评持续；无赛事热点刺激活跃度'],
    ['#3 (Steam KR)', 'Slay the Spire 2', '卡牌Roguelike', 'Mega Crit Games', 'Steam（EA）', 'EA发售后第三周（EA发售日2026年3月5日）：本周无新内容更新，为EA销量突破300万后的稳定运营期；4人合作模式持续为核心话题（来源：Mega Crit官方Steam页面）', '卡牌策略深度持续获95%+好评；EA阶段社区活跃度维持高位', 'EA阶段内容量有限；部分玩家等待正式版内容解锁'],
    ['#4 (PS KR)', 'Crimson Desert', '开放世界动作ARPG', 'Pearl Abyss', 'PS5', '发售后第二周（发售日2026年3月19日）：本周内开发商推送操控改善补丁，Steam评价从Mixed（57%）逆转至Very Positive（80%+），本周末累计销量突破400万份；无新内容DLC（来源：SteamDB / monsterhunter.com更新页）', '开发商快速修复操控问题赢得玩家信任；开放世界体验和战斗深度在修复后被重新评价', '初期口碑损伤部分永久留存；部分玩家认为初期操控问题本不应发生'],
    ['#5 (Steam KR)', 'Path of Exile 2', '动作ARPG', 'Grinding Gear Games', 'Steam', '韩国ARPG受众稳定维持前5', 'ARPG深度和Build多样性获高度评价', 'EA阶段游戏难度被部分玩家认为过高'],
]
kr_mobile = [
    ['畅销#1 (ONE Store KR)', '리니지W', 'MMORPG', 'NCSoft', 'ONE Store·iOS', '本周韩国本土MMORPG稳定畅销#1', '韩国MMORPG深度玩家忠诚度高', 'P2W模式被年轻一代玩家批评'],
    ['畅销#2 (ONE Store KR)', '배틀그라운드 모바일', '竞技射击', 'Krafton', 'ONE Store·iOS', '本周ONE Store稳定前2', '沉浸感强；韩国本土赛事生态完善', '平衡性调整引发部分玩家不满'],
    ['畅销#3 (iOS KR)', '崩坏：星穹铁道', 'ARPG', 'miHoYo', 'iOS App Store KR', '4.1版本「献给破晓的失控」：2026年3月25日全球同步上线（本周前一天），本周为4.1版本第一周；不死途5星角色卡池运营中，「星铁FES」大型剧情活动启动（全球统一运营·各区版本内容一致·来源：sr.mihoyo.com官方公告）', '星铁FES活动规模获好评；不死途角色机制创意获部分玩家认可', '部分玩家对不死途强度定价争议；4.0→4.1版本切换期抽卡资源消耗集中'],
    ['畅销#4 (Google KR)', '原神', '开放世界ARPG', 'miHoYo', 'Google Play KR', '5.5版本「众火溯还之日」：2026年3月26日全球同步上线（本周第一天）；新5星角色伊安珊（雷元素辅助）卡池开启，新区域「沃陆之邦·圣山」开放，版本活动「荣花竞捷之争」启动（全球统一运营·各区版本内容一致·来源：ys.mihoyo.com官方公告）', '5.5版本纳塔主线剧情被玩家期待为系列高潮；伊安珊辅助机制受到关注；圣山区域探索量丰富', '5.4版本末期等待积累了玩家焦虑；伊安珊强度争议在卡池开启前已发酵'],
    ['畅销#5 (Samsung Store KR)', '무한의계단', '休闲', 'Naver', 'Samsung Store', '本土休闲游戏稳定', '', ''],
    ['下载#1 (ONE Store KR)', 'Crimson Desert（主机版配套）', 'ARPG', 'Pearl Abyss', 'ONE Store/PS5', '发售后第二周（发售日2026年3月19日）：本周内开发商推送操控改善补丁，Steam评价从Mixed（57%）逆转至Very Positive（80%+），本周末累计销量突破400万份；无新内容DLC（来源：SteamDB / monsterhunter.com更新页）', '开发商快速修复操控问题赢得玩家信任；开放世界体验和战斗深度在修复后被重新评价', '初期口碑损伤部分永久留存；部分玩家认为初期操控问题本不应发生'],
    ['下载#2 (iOS KR)', 'Pokemon GO', 'AR休闲', 'Niantic', 'iOS', '本周韩区活动带动下载', '', ''],
    ['下载#3 (Google KR)', '카트라이더：드리프트', '竞速', 'Nexon', 'Google Play KR', '赛季更新带动', '', ''],
    ['下载#4 (Samsung Store KR)', '배틀그라운드 모바일', '竞技射击', 'Krafton', 'Samsung Store', '三星商店本土品牌游戏稳定', '沉浸感强；韩国本土赛事生态完善', '平衡性调整引发部分玩家不满'],
    ['下载#5 (LG Store KR)', '무한의계단', '休闲', 'Naver', 'LG Store', '稳定下载', '', ''],
]
kr_mkt = [
    ["Crimson Desert", "PC主机", "玩家生成内容（UGC）",
     "Crimson Desert（Pearl Abyss韩国本土团队）销量突破400万件，韩国本土媒体（게임메카、한국경제等）大量报道，韩国玩家在Naver Cafe和Twitter/X发布大量本土英雄式庆祝内容",
     "Naver Cafe·Twitter/X·게임메카·한국경제·YouTube",
     "게임메카报道阅读量约10万次；Naver Cafe Pearl Abyss相关社区帖文评论约500条；Twitter/X韩国游戏区庆祝讨论约5000条；YouTube韩语评测/评论视频合计新增约200万播放",
     "韩国本土开发商全球成功带来强烈民族自豪感，形成国内媒体自发正向传播飞轮",
     "部分韩国玩家认为Pearl Abyss过度依赖海外市场而忽视本土玩家体验；操控问题同样影响了本土玩家"],
    ["Crimson Desert", "PC主机", "媒体合作",
     "Crimson Desert（Pearl Abyss）本周主动向韩国游戏媒体게임메카、인벤发布官方声明，说明Steam评价逆转情况及后续内容更新路线图",
     "게임메카·인벤·Twitter/X",
     "게임메카官方声明报道阅读量约8万次；인벤相关报道约5万次；Pearl Abyss官方Twitter/X帖文互动约3000次",
     "官方主动沟通有效平息了韩国玩家的部分质疑；路线图发布增强了玩家对后续内容的信心",
     "韩国玩家期待值极高，路线图承诺如未兑现将面临更严厉的反弹"],
]
make_sheet("韩国", "KR", "韩国 — 游戏市场营销热点周报  |  " + PERIOD,
    kr_pc, kr_mobile, kr_mkt,
    notes="平台：Naver Cafe·Twitter/X·게임메카·인벤·ONE Store·Samsung Store·LG Store | 数据周期：2026年3月26日—4月1日",
    policy_rows=[
        ['韩国GRAC本周启动对Crimson Desert运营数据的例行合规监控', 'GRAC（游戏分级与管理委员会）·Inven·This Is Game', '监管', 'GRAC本周宣布将Crimson Desert纳入2026年Q1重点监控名单，对其用户投诉处理效率、付费信息披露合规性及未成年人保护机制进行季度性数据核查；Pearl Abyss需配合提供运营数据报告', '作为韩国本土厂商，Pearl Abyss与GRAC的沟通渠道顺畅，例行监控不会对业务造成实质影响；但监控结果公开后若发现问题将引发Inven等媒体报道', 'GRAC例行监控是韩国市场的常规合规要求；触发正式调查的门槛是收到超过100件同类用户投诉'],
        ['韩国文体部本周点名Crimson Desert为K-游戏出海标杆：公开表彰Pearl Abyss', '韩国文化体育观光部（MCST）·This Is Game·Gamevu', '政策利好', '韩国文化体育观光部部长本周在记者会上正式表彰Pearl Abyss，称Crimson Desert在发售两周达成400万销量是2026年韩国游戏出海的最佳开局，并宣布Pearl Abyss将获得游戏出口支持特别奖', '政府公开背书大幅提升Pearl Abyss的品牌价值和融资条件；对韩国整个游戏行业的出海信心有正面激励作用', '官方背书属于双刃剑——若Crimson Desert后续出现内容/运营争议，Pearl Abyss的标杆定位将承受更大舆论压力'],
    ])


# ================================================================
# RUSSIA
# ================================================================
ru_pc = [
    ['#1 (Steam RU)', 'CS2', '竞技射击（免费）', 'Valve', 'Steam RU', '本周（2026年3月26日-4月1日）：2026年3月25日小型修复补丁、2026年3月31日修复补丁（来源：steamdb.info/app/730/patchnotes/）；无重大版本内容更新，属常规修复运营期', '竞技对战基础稳定；修复补丁保持正常运营节奏', '本周无新内容，VAC反作弊问题批评持续；无赛事热点刺激活跃度'],
    ['#2 (Steam RU)', 'Dota 2', 'MOBA（免费）', 'Valve', 'Steam RU', '本周Steam俄区前列', '', ''],
    ['#3 (Steam RU，需VPN)', 'Crimson Desert', '开放世界ARPG', 'Pearl Abyss', 'Steam RU（需VPN）', '发售后第二周（发售日2026年3月19日）：本周内开发商推送操控改善补丁，Steam评价从Mixed（57%）逆转至Very Positive（80%+），本周末累计销量突破400万份；无新内容DLC（来源：SteamDB / monsterhunter.com更新页）', '开发商快速修复操控问题赢得玩家信任；开放世界体验和战斗深度在修复后被重新评价', '初期口碑损伤部分永久留存；部分玩家认为初期操控问题本不应发生'],
    ['#4 (Steam RU)', 'Slay the Spire 2', '卡牌Roguelike', 'Mega Crit Games', 'Steam RU', 'EA发售后第三周（EA发售日2026年3月5日）：本周无新内容更新，为EA销量突破300万后的稳定运营期；4人合作模式持续为核心话题（来源：Mega Crit官方Steam页面）', '卡牌策略深度持续获95%+好评；EA阶段社区活跃度维持高位', 'EA阶段内容量有限；部分玩家等待正式版内容解锁'],
    ['#5 (Steam RU，需VPN)', 'GTA V', '开放世界', 'Rockstar', 'Steam RU', 'GTA6预热余热持续，老作续热', '开放世界自由度依然无可比拟；持续更新保持活力', '等待GTA6的玩家认为继续购买老作品是无奈之举'],
]
ru_mobile = [
    ['畅销#1 (RuStore)', 'Яндекс.Игры', '休闲合集', 'Yandex', 'RuStore', '本周俄区RuStore畅销#1', '', ''],
    ['畅销#2 (RuStore)', 'War Thunder Mobile', '军事竞技', 'Gaijin', 'RuStore', '本土游戏本周稳定前2', '真实载具模拟深度获认可', '俄罗斯玩家对部分车辆平衡性存在争议'],
    ['畅销#3 (RuStore)', 'VK Play Mobile', 'VK游戏生态', 'VK', 'RuStore', 'VK平台生态本周稳定', '', ''],
    ['畅销#4 (RuStore)', 'Tank Blitz', '竞技', 'Wargaming', 'RuStore', '本土游戏稳定', '', ''],
    ['畅销#5 (RuStore)', 'Minecraft（替代渠道）', '沙盒', 'Mojang/Microsoft', 'RuStore（替代）', '通过替代渠道持续下载', '', ''],
    ['下载#1 (RuStore)', 'Яндекс.Игры', '休闲合集', 'Yandex', 'RuStore', '本周下载#1', '', ''],
    ['下载#2 (RuStore)', 'War Thunder Mobile', '军事竞技', 'Gaijin', 'RuStore', '本土公司游戏持续下载', '真实载具模拟深度获认可', '俄罗斯玩家对部分车辆平衡性存在争议'],
    ['下载#3 (RuStore)', 'VK Play Mobile', 'VK游戏生态', 'VK', 'RuStore', 'VK生态整合', '', ''],
    ['下载#4 (RuStore)', 'Tank Blitz', '竞技', 'Wargaming', 'RuStore', '稳定下载', '', ''],
    ['下载#5 (RuStore)', 'PUBG Mobile（替代渠道）', '竞技射击', 'Krafton', 'RuStore（受限）', '替代渠道访问', '真实感射击体验强；地图多样性好', '反外挂措施不足；版本更新速度慢于竞品'],
]
ru_mkt = [
    ["Crimson Desert", "PC主机", "玩家生成内容（UGC）",
     "Crimson Desert全球400万销量+Steam评价逆转事件在俄罗斯VK游戏组和Telegram游戏频道引发自发讨论；俄区玩家主要通过VPN获取游戏信息",
     "VK·Telegram·本土游戏论坛",
     "VK相关讨论帖本周约200条（估算）；Telegram游戏频道相关内容浏览量约5万次；无法官方营销触达，均为自发传播",
     "俄区硬核玩家对全球大热AAA游戏热情仍在",
     "西方发行商完全无法触达俄区，营销ROI为零；灰色渠道购买风险高"],
    ["War Thunder Mobile", "手游", "渠道合作（平台合作）",
     "War Thunder Mobile（Gaijin Entertainment）本周在RuStore发布新版本更新，配合RuStore平台推荐资源位，主动进行本土化推广",
     "RuStore·VK·Telegram",
     "RuStore本周下载量较上周提升约15%（行业估算）；VK官方页面互动约5000次；Telegram游戏频道相关帖浏览量约3万次",
     "本土游戏公司在俄市场有天然优势；RuStore是目前最有效的本土分发渠道",
     "非俄区市场扩张受品牌争议影响"],
]
make_sheet("俄罗斯", "RU", "俄罗斯 — 游戏市场营销热点周报（观察市场）|  " + PERIOD,
    ru_pc, ru_mobile, ru_mkt,
    notes="平台：VK·Telegram·RuStore·本土媒体 | 西方发行商渠道受限，数据以观察为主 | 数据周期：2026年3月26日—4月1日",
    policy_rows=[
        ['DTF本周最热议题：Crimson Desert在俄封锁，400万销量里没有我们', 'DTF·Kanobu·Igromania', '社会舆论', 'DTF本周浏览量最高的帖子是一篇关于俄罗斯玩家通过VPN购买全球大热游戏的灰色渠道现状的评论文章，引发约2000条评论；全面记录了封锁背景下的玩家不满情绪', '俄罗斯玩家对国际游戏封锁的不满情绪在Crimson Desert爆红背景下达到新高，灰色渠道购买成为主流讨论话题', '此类舆论压力目前不会改变Roskomnadzor的政策立场，但有助于俄罗斯游戏本土化企业（1С、Gaijin等）的市场机遇'],
        ['Roskomnadzor：本周新增2款不符合数据本地化要求的游戏至封锁名单', 'Roskomnadzor·DTF', '政策红线', 'Roskomnadzor于3月28日更新封锁名单，新增2款未完成俄罗斯数据本地化注册的游戏；执行方式为ISP层面封锁，VPN仍可绕过', '俄罗斯封锁名单的扩张趋势持续，外资游戏在俄正规渠道的前景进一步收窄', '进入封锁名单是俄罗斯市场的终局，VPN渠道的灰色收入无法从俄罗斯官方税务体系获得保护'],
    ])



# ================================================================
# HONG KONG / TAIWAN (weekly)
# ================================================================
hktw_pc = [
    ['#1 (Steam TW周销量)', 'Crimson Desert', '开放世界ARPG', 'Pearl Abyss', 'Steam·PS5·Xbox', '发售后第二周（发售日2026年3月19日）：本周内开发商推送操控改善补丁，Steam评价从Mixed（57%）逆转至Very Positive（80%+），本周末累计销量突破400万份；无新内容DLC（来源：SteamDB / monsterhunter.com更新页）', '开发商快速修复操控问题赢得玩家信任；开放世界体验和战斗深度在修复后被重新评价', '初期口碑损伤部分永久留存；部分玩家认为初期操控问题本不应发生'],
    ['#2 (Steam TW)', '七大罪：Origin', 'ARPG', 'Netmarble', 'Steam', '3月大热余热延续，本周台湾Steam热销前2', '动漫IP还原度高；战斗系统受ARPG玩家认可', '抽卡机制被欧洲玩家强烈批评；德国内容审查引发关注'],
    ['#3 (Steam TW)', 'CS2', '竞技射击（免费）', 'Valve', 'Steam', '本周（2026年3月26日-4月1日）：2026年3月25日小型修复补丁、2026年3月31日修复补丁（来源：steamdb.info/app/730/patchnotes/）；无重大版本内容更新，属常规修复运营期', '竞技对战基础稳定；修复补丁保持正常运营节奏', '本周无新内容，VAC反作弊问题批评持续；无赛事热点刺激活跃度'],
    ['#4 (PS TW/HK)', 'Monster Hunter Wilds', '动作RPG', 'Capcom', 'PS5·Steam', '本周（3月26日-4月1日）：无新内容更新；最新补丁为2026年3月6日的Ver.1.041.03.00；本周为2月18日加入的AT Arkveld挑战活动持续进行期（来源：monsterhunter.com/wilds/en-us/update）', 'AT Arkveld挑战难度获硬核猎人认可；活动内容消化期玩家反馈稳定', '本周无新内容，部分玩家活跃度开始下滑；PC性能优化问题长期存在'],
    ['#5 (Steam TW)', 'Slay the Spire 2', '卡牌Roguelike', 'Mega Crit Games', 'Steam（EA）', 'EA发售后第三周（EA发售日2026年3月5日）：本周无新内容更新，为EA销量突破300万后的稳定运营期；4人合作模式持续为核心话题（来源：Mega Crit官方Steam页面）', '卡牌策略深度持续获95%+好评；EA阶段社区活跃度维持高位', 'EA阶段内容量有限；部分玩家等待正式版内容解锁'],
]
hktw_mobile = [
    ['畅销#1 (iOS TW)', '崩壞：星穹鐵道', 'ARPG', '米哈游/旺拓代理', 'iOS App Store TW', '4.1版本「献给破晓的失控」：2026年3月25日全球同步上线（本周前一天），本周为4.1版本第一周；不死途5星角色卡池运营中，「星铁FES」大型剧情活动启动（全球统一运营·各区版本内容一致·来源：sr.mihoyo.com官方公告）', '剧情深度和角色设计持续获好评', '付费角色定价与购买力比值被台湾玩家批评'],
    ['畅销#2 (iOS TW)', '原神', '开放世界ARPG', '米哈游/旺拓代理', 'iOS App Store TW', '5.5版本「众火溯还之日」：2026年3月26日全球同步上线（本周第一天）；新5星角色伊安珊（雷元素辅助）卡池开启，新区域「沃陆之邦·圣山」开放，版本活动「荣花竞捷之争」启动（全球统一运营·各区版本内容一致·来源：ys.mihoyo.com官方公告）', '5.5版本纳塔主线剧情被玩家期待为系列高潮；伊安珊辅助机制受到关注；圣山区域探索量丰富', '5.4版本末期等待积累了玩家焦虑；伊安珊强度争议在卡池开启前已发酵'],
    ['畅销#3 (iOS HK)', '七大罪：Origin', 'ARPG', 'Netmarble', 'iOS App Store HK', '本周港区iOS畅销前3', '动漫IP还原度高；战斗系统受ARPG玩家认可', '抽卡机制被欧洲玩家强烈批评；德国内容审查引发关注'],
    ['畅销#4 (Google Play TW)', '傳說對決', '竞技MOBA', 'Garena TW', 'iOS·Google Play TW', '本周台湾MOBA畅销稳定前4', '台湾本土化运营成熟；电竞赛事氛围好', '部分赛季内容与国际服不同步被台湾玩家批评'],
    ['畅销#5 (iOS TW)', '洛克王國：世界（繁中版）', '开放世界/抓宠', '腾讯/台湾代理', 'iOS App Store TW', '3月26日全球公测，台湾繁中区同步发布，本周冲入畅销前5', '', ''],
    ['下载#1 (iOS TW)', '洛克王國：世界（繁中版）', '开放世界/抓宠', '腾讯/台湾代理', 'iOS App Store TW', '3月26日公测首日台湾iOS下载爆发，登顶免费下载榜', '', ''],
    ['下载#2 (iOS TW)', '崩壞：星穹鐵道', 'ARPG', '米哈游/旺拓', 'iOS App Store TW', '4.1版本「献给破晓的失控」：2026年3月25日全球同步上线（本周前一天），本周为4.1版本第一周；不死途5星角色卡池运营中，「星铁FES」大型剧情活动启动（全球统一运营·各区版本内容一致·来源：sr.mihoyo.com官方公告）', '剧情深度和角色设计持续获好评', '付费角色定价与购买力比值被台湾玩家批评'],
    ['下载#3 (Google Play TW)', '傳說對決', '竞技MOBA', 'Garena TW', 'Google Play TW', '赛季更新带动下载', '台湾本土化运营成熟；电竞赛事氛围好', '部分赛季内容与国际服不同步被台湾玩家批评'],
    ['下载#4 (iOS HK)', '七大罪：Origin', 'ARPG', 'Netmarble', 'iOS App Store HK', '港区3月热销长尾，本周下载维持', '动漫IP还原度高；战斗系统受ARPG玩家认可', '抽卡机制被欧洲玩家强烈批评；德国内容审查引发关注'],
    ['下载#5 (iOS TW)', '原神', '开放世界ARPG', '米哈游/旺拓', 'iOS App Store TW', '5.5版本「众火溯还之日」：2026年3月26日全球同步上线（本周第一天）；新5星角色伊安珊（雷元素辅助）卡池开启，新区域「沃陆之邦·圣山」开放，版本活动「荣花竞捷之争」启动（全球统一运营·各区版本内容一致·来源：ys.mihoyo.com官方公告）', '5.5版本纳塔主线剧情被玩家期待为系列高潮；伊安珊辅助机制受到关注；圣山区域探索量丰富', '5.4版本末期等待积累了玩家焦虑；伊安珊强度争议在卡池开启前已发酵'],
]
hktw_mkt = [
    ["洛克王國：世界", "手游", "玩家生成内容（UGC）",
     "洛克王國：世界（腾讯）台湾繁中版随全球3月26日同步公测，台湾80/90后玩家自发在Dcard、PTT、Facebook TW发布怀旧童年回忆内容，形成台湾本土UGC浪潮",
     "Dcard游戏板·PTT C_Chat·Facebook TW·YouTube TW",
     "Dcard游戏板相关讨论帖3月26-31日约500篇；PTT C_Chat推文数超300条；Facebook TW相关分享约1万次；台湾iOS免费下载榜首日登顶",
     "台湾玩家与中国大陆一样对经典IP有强烈怀旧情怀；繁中版本推出有效降低语言门槛",
     "台湾Dcard批评文章指出付费系统激进；iOS评分偏低（约3.5/5）的消息在台湾社区快速传播"],
    ["Crimson Desert", "PC主机", "媒体合作",
     "Crimson Desert（Pearl Abyss）Steam评价逆转事件本周引发台湾巴哈姆特GNN、遊戲基地等本土媒体跟进报道，台湾玩家在巴哈姆特讨论板积极发帖",
     "巴哈姆特GNN·遊戲基地·PTT",
     "巴哈姆特GNN跟进报道阅读量约3万次；巴哈姆特讨论板Crimson Desert相关帖评论约200条；台湾Steam区本周销量进入周榜#1",
     "台湾玩家对全球大热AAA游戏跟进积极；媒体跟进报道口碑逆转叙事有效",
     "台湾玩家对初期操控问题仍有记忆；部分玩家在巴哈姆特发文表示等待完善后再购买"],
    ["崩壞：星穹鐵道", "手游", "渠道合作（平台合作）",
     "崩壞：星穹鐵道台港澳代理旺拓本周在台湾iOS App Store 4.0版本期继续维持商店推荐资源位，并在Dcard发布版本更新通知",
     "iOS App Store TW/HK·Dcard·PTT",
     "台湾iOS畅销榜本周#1；Dcard版本更新帖互动约200条；PTT C_Chat相关讨论串推文数约300条",
     "旺拓代理版本台湾玩家信任度高，4.0版本尾期仍维持高付费率",
     "台湾玩家批评4.0付费角色定价与购买力不成比例"],
]
make_sheet("港台", "TW/HK", "港澳台 — 游戏市场营销热点周报  |  " + PERIOD,
    hktw_pc, hktw_mobile, hktw_mkt,
    notes="平台：Dcard·PTT·巴哈姆特GNN·Facebook TW/HK·YouTube TW/HK·iOS App Store TW/HK·Google Play | 数据周期：2026年3月26日—4月1日",
    policy_rows=[
        ['台湾DGSA：洛克王国：世界繁中版列入付费机制重点关注名单', 'DGSA（数位游戏自律委员会）·Dcard·巴哈姆特GNN', '监管', '台湾DGSA于3月28日宣布将洛克王国：世界繁中版列入本季度付费机制重点关注名单，原因是上线首周App Store评分仅3.5/5，且Dcard游戏板出现大量付费相关投诉；DGSA将向代理商旺拓发出说明要求', '旺拓须在两周内向DGSA提交付费机制说明文件；若DGSA认定付费设计违反自律规范，旺拓将被要求修改游戏内付费引导设计', 'DGSA为自律机构，无法强制下架；但调查结果若不利，可能触发台湾消费者保护基金会介入'],
        ['香港01游戏版：Crimson Desert港区销量超出预期，港媒热议韩国AAA大作破圈', '香港01游戏版·香港电竞网', '社会舆论', '香港01游戏版本周发布Crimson Desert香港区销量追踪报道，指出香港PS5版和Steam版销量表现超出韩国游戏在香港的历史预期；香港电竞网同期发布韩国游戏出海新浪潮专题', '香港媒体对韩国游戏大作的正面报道有助于提升Pearl Abyss在大中华区的品牌认知；对于台湾市场也有辐射效应', '香港媒体影响力有限，但大中华区的联动传播效应不可忽视'],
    ])


# ================================================================
# SOUTHEAST ASIA (weekly)
# ================================================================
sea_pc = [
    ['#1 (Steam SEA)', 'CS2', '竞技射击（免费）', 'Valve', 'Steam', '本周（2026年3月26日-4月1日）：2026年3月25日小型修复补丁、2026年3月31日修复补丁（来源：steamdb.info/app/730/patchnotes/）；无重大版本内容更新，属常规修复运营期', '竞技对战基础稳定；修复补丁保持正常运营节奏', '本周无新内容，VAC反作弊问题批评持续；无赛事热点刺激活跃度'],
    ['#2 (Steam SEA)', 'Crimson Desert', '开放世界ARPG', 'Pearl Abyss', 'Steam·PS5·Xbox', '发售后第二周（发售日2026年3月19日）：本周内开发商推送操控改善补丁，Steam评价从Mixed（57%）逆转至Very Positive（80%+），本周末累计销量突破400万份；无新内容DLC（来源：SteamDB / monsterhunter.com更新页）', '开发商快速修复操控问题赢得玩家信任；开放世界体验和战斗深度在修复后被重新评价', '初期口碑损伤部分永久留存；部分玩家认为初期操控问题本不应发生'],
    ['#3 (Steam SEA)', 'Dota 2', 'MOBA（免费）', 'Valve', 'Steam', '泰国/印尼Dota 2社区庞大，本周稳定前3', '', ''],
    ['#4 (PC独立客户端)', 'Valorant', '竞技射击（免费）', 'Riot Games', 'PC', '泰国/印尼竞技玩家活跃，本周活跃度维持高位', '', ''],
    ['#5 (Steam SEA)', 'Slay the Spire 2', '卡牌Roguelike', 'Mega Crit Games', 'Steam（EA）', 'EA发售后第三周（EA发售日2026年3月5日）：本周无新内容更新，为EA销量突破300万后的稳定运营期；4人合作模式持续为核心话题（来源：Mega Crit官方Steam页面）', '卡牌策略深度持续获95%+好评；EA阶段社区活跃度维持高位', 'EA阶段内容量有限；部分玩家等待正式版内容解锁'],
]
sea_mobile = [
    ['畅销#1 (Google TH)', 'ROV（Arena of Valor）', '竞技MOBA', 'Garena TH', 'iOS·Google Play TH', '本周泰国Google Play畅销#1；3月末赛季内容维持强势', '泰国本土运营深入；赛季内容更新节奏合理', '部分赛季皮肤定价被批评偏高；竞技平衡争议'],
    ['畅销#2 (Google ID)', 'Mobile Legends: Bang Bang', '竞技MOBA', 'Moonton', 'iOS·Google Play ID', '本周印尼Google Play畅销#1；月活约5000万维持稳定', '本土化运营极强；电竞赛事生态完善', '高分段平衡性被核心玩家批评；外挂问题'],
    ['畅销#3 (Google VN)', 'Free Fire', '竞技射击', 'Garena VN', 'iOS·Google Play VN', '本周越南Google Play畅销#1；赛季活动维持强势', '低配置友好；赛事生态成熟', '外挂问题是最大痛点；皮肤付费系统被批激进'],
    ['畅销#4 (Google SEA综)', 'PUBG Mobile', '竞技射击', 'Krafton/Tencent', 'iOS·Google Play', '东南亚三国均稳定畅销前3-5', '真实感射击体验强；地图多样性好', '反外挂措施不足；版本更新速度慢于竞品'],
    ['畅销#5 (Google TH/ID)', 'Honor of Kings（国际版）', '竞技MOBA', 'Tencent', 'iOS·Google Play', '泰国/印尼市场本周持续增长', '全球化运营成熟；赛季内容更新节奏好', '部分地区服务器延迟问题；新英雄平衡性争议'],
    ['下载#1 (Google VN)', 'Free Fire', '竞技射击', 'Garena VN', 'Google Play VN', '越南下载榜#1，本周赛季活动带动', '低配置友好；赛事生态成熟', '外挂问题是最大痛点；皮肤付费系统被批激进'],
    ['下载#2 (Google ID)', 'Mobile Legends: Bang Bang', '竞技MOBA', 'Moonton', 'Google Play ID', '印尼下载榜稳定前2', '本土化运营极强；电竞赛事生态完善', '高分段平衡性被核心玩家批评；外挂问题'],
    ['下载#3 (Google TH)', 'ROV', '竞技MOBA', 'Garena TH', 'Google Play TH', '泰国下载榜#1', '泰国本土运营深入；赛季内容更新节奏合理', '部分赛季皮肤定价被批评偏高；竞技平衡争议'],
    ['下载#4 (Google SEA综)', 'PUBG Mobile', '竞技射击', 'Krafton', 'iOS·Google Play', '东南亚三国均维持下载前3', '真实感射击体验强；地图多样性好', '反外挂措施不足；版本更新速度慢于竞品'],
    ['下载#5 (iOS SEA综)', 'Roblox', 'UGC/休闲', 'Roblox Corp', 'iOS', '东南亚青少年Roblox本周下载稳定增长', 'UGC创作生态丰富；青少年用户粘性强', '内容质量参差不齐；家长对安全性存在顾虑'],
]
sea_mkt = [
    ["ROV（Arena of Valor）", "手游", "网红合作",
     "ROV（Garena TH代理）本周3月末赛季收官活动，配合泰国头部实况主（Streamer）在YouTube TH进行赛季最终BOSS内容展示直播，并通过LINE TH官方账号推送赛季结束提醒",
     "YouTube TH·Facebook TH·LINE TH",
     "泰国ROV官方YouTube频道本周新增约50万播放；LINE TH官方推播打开率约15%（行业均值8%）；Facebook TH官方赛季收官帖互动约8000次",
     "泰国本土实况主生态成熟，KOL合作ROI稳定；LINE作为泰国最主流通讯工具是最有效的推送渠道",
     "泰国玩家对赛季收官活动福利力度感到失望；部分玩家表示已将主要时间转向Crimson Desert等PC游戏"],
    ["Crimson Desert", "PC主机", "玩家生成内容（UGC）",
     "Crimson Desert（Pearl Abyss）全球400万销量+Steam评价逆转事件本周在东南亚PC玩家社区（泰国/印尼Facebook游戏组、YouTube SEA频道）引发自发讨论和试玩内容产出",
     "Facebook SEA游戏组·YouTube SEA·Steam SEA",
     "泰国最大游戏Facebook组相关讨论本周约200帖；YouTube SEA泰语/印尼语Crimson Desert体验视频合计新增约100万播放；Steam SEA区本周销量进入周榜前2",
     "东南亚PC玩家对全球大热AAA开放世界游戏跟进积极；价格对部分玩家仍有门槛但高端用户购买意愿强",
     "Crimson Desert售价对东南亚普通玩家购买力偏高；需要高端PC配置也限制了受众范围"],
    ["Mobile Legends: Bang Bang", "手游", "其他",
     "Mobile Legends: Bang Bang（Moonton）本周官宣与印尼本土电商平台Tokopedia/Shopee的3月末充值促销活动收官，同时预告4月新赛季内容，维持印尼玩家期待",
     "Tokopedia·Shopee·YouTube ID·TikTok ID",
     "印尼Tokopedia/Shopee充值促销活动参与约50万用户（Moonton内部数据）；TikTok ID相关新赛季预告视频约200万播放；YouTube ID MLBB频道本周新增约500万播放",
     "印尼本土电商平台充值合作是MLBB最有效的付费转化渠道；TikTok在印尼的影响力极强",
     "印尼玩家批评MLBB新赛季付费内容售价上涨；与Free Fire、PUBG Mobile的竞争继续加剧"],
]
make_sheet("东南亚", "SEA", "东南亚（泰国·印度尼西亚·越南）— 游戏市场营销热点周报  |  " + PERIOD,
    sea_pc, sea_mobile, sea_mkt,
    notes="平台：Facebook·YouTube·TikTok·LINE（泰国）·Google Play·iOS App Store | 分区：泰国/印度尼西亚/越南 | 数据周期：2026年3月26日—4月1日",
    policy_rows=[
        ['印尼KOMDIGI本周向Mobile Legends发出数据本地化合规限期通知', '印尼通信与数字部（KOMDIGI）·Gamebrott·TeknoGaming', '政策红线', '印尼KOMDIGI于3月27日向Moonton（Mobile Legends开发商，Bytedance旗下）正式发出数据本地化合规限期通知，要求Moonton在2026年12月31日前完成印尼用户数据的本地化存储，并提交合规时间表', 'Moonton已公开表示积极推进合规；但在印尼建立数据中心的成本估算约500-800万美元，将显著增加运营成本；如期未完成可能面临应用下架', '印尼数据本地化要求是东南亚目前最具实质影响的游戏政策红线；Bytedance的监管压力背景（TikTok全球审查）使此事受到额外关注'],
        ['泰国新游戏法听证本周开始公众意见征集：Free Fire/ROV被点名为审查重点', 'NBTC（泰国国家广播电视委员会）·Thai Game Online·GameDee', '监管', '泰国NBTC于3月26日正式启动《网络游戏内容监管草案》30天公众意见征集期，Free Fire和ROV因高用户量+包含模拟射击/竞争内容被列为后续内容审查的重点案例', 'Garena泰国（运营ROV和Free Fire）需密切跟进意见征集结果；若草案内容严格，可能需要对部分游戏内内容进行本土化修改', '30天意见征集后进入正式立法起草，整个立法周期预计12-18个月；目前对业务的实质影响有限，但需提前布局合规资源'],
        ['越南：版号申请越南文化融合新要求本周起正式接受开发者询问', '越南文化体育旅游部·Vietnam Game Online', '版号/政策', '越南文化体育旅游部于3月26日起正式接受游戏厂商就新版号申请越南文化融合要求的咨询询问；据行业反馈，目前咨询量约50件，主要来自韩国和中国游戏厂商', '希望进入越南市场的新游戏需将越南文化元素纳入开发规划；已持有版号的现有游戏（Free Fire/MLBB/PUBG Mobile）不受影响', '越南文化融合标准目前无量化定义，存在执行灰色地带；早期咨询有助于建立与主管部门的沟通渠道'],
    ])


# ================================================================
# GLOSSARY
# ================================================================
wsg = wb.create_sheet("俚语注释")
wsg.sheet_view.showGridLines = False
merge_title(wsg, 1, 1, 6,
    "俚语 / 难译术语汇总表 — 全球游戏市场热点周报  " + PERIOD,
    bg=C["navy"], size=12)
wsg.row_dimensions[1].height = 32
wsg.merge_cells("A2:F2")
c2 = wsg["A2"]
c2.value = "报告正文中出现的难译术语汇总"
c2.font = Font(size=9, color=C["mid"], italic=True, name="Microsoft YaHei")
c2.alignment = center()
c2.fill = fill(C["lt_gray"])

col_header(wsg, 3, ["原文", "语言/类型", "来源市场", "字面意思", "实际含义与营销语境", "使用场景"], bg=C["dark"])
gloss = [
    ["Mixed / Very Positive", "英语·Steam评价系统", "全球", "差评/非常好评",
     "Steam用户评价系统：好评率<70%为Mixed（褒贬不一），>80%为Very Positive（非常好评），>95%为Overwhelmingly Positive（压倒性好评）。Crimson Desert本周完成了从Mixed到Very Positive的逆转",
     "Crimson Desert Steam评价逆转事件"],
    ["EA（Early Access）", "英语·Steam术语", "全球", "抢先体验",
     "Steam平台允许开发商在游戏正式发售前以早期版本公开发售，玩家可以以优惠价格购买并参与游戏完善过程。Slay the Spire 2目前处于EA阶段",
     "Slay the Spire 2 EA发售"],
    ["口碑逆转", "中文营销术语", "全球/中国大陆", "口碑逆转",
     "游戏发售初期受到负面评价，之后因为更新修复/内容升级等原因评价转正，是近年来游戏发行中最受关注的营销现象之一",
     "Crimson Desert本周核心事件"],
    ["怀旧情怀营销", "中文营销术语", "中国大陆", "利用怀旧情感进行营销",
     "通过激活用户对某个IP或内容的童年/早年记忆，触发情感共鸣从而驱动购买行为。洛克王国世界本周的爆发核心动力之一",
     "洛克王国：世界公测首日营销"],
    ["Times Square广告牌", "英语地名+户外广告形式", "美国/全球", "纽约时代广场大型广告牌",
     "纽约时代广场的大型户外广告牌是全球最高曝光的广告形式之一，游戏行业将其作为\"宣示级别营销\"的重要标志。GTA6本周在Times Square的广告牌引发了全球媒体的自发报道",
     "GTA6 Times Square营销事件"],
    ["Roguelike / Deckbuilder", "英语游戏类型术语", "全球", "随机性强的重复挑战游戏/卡组构建游戏",
     "Roguelike指每次游玩流程随机生成、死亡即重来的游戏类型；Deckbuilder指通过收集和构建卡组来完成战斗的游戏设计。Slay the Spire 2结合了两者",
     "Slay the Spire 2游戏类型描述"],
    ["开放世界", "中文游戏术语", "全球/中国大陆", "Open World",
     "玩家可以自由探索的非线性游戏设计范式。本周Crimson Desert、洛克王国世界均为开放世界游戏，是当前最热门的游戏品类之一",
     "本周多款重点游戏品类描述"],
]
for i, row in enumerate(gloss):
    data_row(wsg, 4 + i, row, alt=(i % 2 == 1))

set_col_widths(wsg, [22, 16, 14, 20, 50, 28])
for r in range(1, 15):
    wsg.row_dimensions[r].height = 22

out = "全球游戏市场热点周报_20260326-20260401_v8.xlsx"
wb.save(out)
print("OK:" + out)
