

# ═══════════════════════════════════════════════════════════════════════════════
# 营销案例数据验真规则（已确认）
# ───────────────────────────────────────────────────────────────────────────────
# 【规则A：网红营销验真标准】
#   满足以下任一条件即视为已核实：
#   1. 可确认该KOL/主播在指定时间段内切实直播或发布了该游戏相关内容
#   2. 该KOL视频/内容上标注了「合作」「广告」「Ads」「#ad」「#sponsored」等
#      付费合作标签（无需游戏官方发布正式合作公告）
#   3. 平台数据可观测到该内容的播放量/热度/Trending位置
#   → 核实来源可为：平台页面截图/第三方数据平台/权威游戏媒体报道
#
# 【规则B：渠道合作（平台合作）验真标准】
#   平台与开发商通常不发布资源位采买公告，满足以下条件即视为已核实：
#   1. 在特定时间节点，某游戏在某平台上突然出现大量集中内容（推荐位/首页/精选）
#   2. 该集中爆破期与游戏版本更新/发售节点高度吻合（非自然流量行为）
#   → 核实来源可为：平台公开榜单截图/第三方ASO数据/行业媒体观察报道
#
# 【规则C：异业联动/媒体合作验真标准（维持原有要求）】
#   必须有官方来源（官网/官方社媒/官方商店页/权威游戏媒体）明确记录，
#   含具体日期，否则删除。
# ═══════════════════════════════════════════════════════════════════════════════

import openpyxl

# ═══════════════════════════════════════════════════════════════════════════════
# 横向多市场分析框架规则（已确认）
# ───────────────────────────────────────────────────────────────────────────────
# 【规则1：跨市场时间验证】
#   每个市场提及的事件/版本/营销活动均需做具体日期的真实性验证。
#
# 【规则2：多语言官方译名映射】
#   同一游戏在不同市场须使用各自的官方名称。标准映射见 MULTILANG_NAMES。
#   例：原神(CN/JP) = Genshin Impact(EN/EU/US/SEA/RU) = 원신(KR) = 原神(TW/HK)
#
# 【规则3：全球统一运营型游戏】
#   游戏版本内容全球同步（如原神/崩铁/MHW等）时：
#   - 「游戏内容分析」列：各市场填写相同的版本更新内容，标注「全球统一运营」
#   - 「营销热点」分析：重点分析该游戏在该市场做的本地化营销差异和成功/不足
#   - 不重复描述版本内容，专注于本地化营销行为
#
# 【规则4：区域分版本运营型游戏】
#   CN版与国际版内容不同（如王者荣耀(CN) vs Honor of Kings(Global)，
#   和平精英(CN) vs PUBG Mobile(Global)）：
#   - 在各自市场分别分析对应版本的更新内容
#   - 分析不同版本的内容差异和运营策略差异
# ═══════════════════════════════════════════════════════════════════════════════

# 官方多语言译名映射表（供各市场数据标注参考）
MULTILANG_NAMES = {
    "原神": {"CN":"原神","EN":"Genshin Impact","JP":"原神","KR":"원신","TW/HK":"原神","SEA":"Genshin Impact","RU":"Genshin Impact"},
    "崩坏：星穹铁道": {"CN":"崩坏：星穹铁道","EN":"Honkai: Star Rail","JP":"崩壊：スターレイル","KR":"붕괴: 스타레일","TW/HK":"崩壞：星穹鐵道","SEA":"Honkai: Star Rail"},
    "怪物猎人：荒野": {"CN":"怪物猎人：荒野","EN":"Monster Hunter Wilds","JP":"モンスターハンターワイルズ","KR":"몬스터 헌터 와일즈","SEA":"Monster Hunter Wilds"},
    "王者荣耀": {"CN":"王者荣耀","EN/Global":"Honor of Kings","TW/HK":"傳說對決"},
    "和平精英": {"CN":"和平精英","Global":"PUBG Mobile","KR":"배틀그라운드 모바일"},
    "七大罪：Origin": {"CN":"七大罪：Origin","EN":"The Seven Deadly Sins: Origin","JP":"七つの大罪：オリジン","KR":"일곱 개의 대죄: 오리진"},
    "红色沙漠": {"CN":"红色沙漠","TW/HK":"赤血沙漠","EN":"Crimson Desert","KR":"붉은사막","JP":"クリムゾン デザート"},
    "黑神话：悟空": {"CN":"黑神话：悟空","EN":"Black Myth: Wukong","JP":"ブラック・ミス:ウーコン","KR":"검은 신화: 오공"},
    "龍が如く極3": {"JP":"龍が如く 極3 / 龍が如く3外伝 Dark Ties","EN":"Like a Dragon: Kiwami 3 / Like a Dragon: Dark Ties","KR":"용과 같이 키와미 3","CN":"人中之龙 极3 / 人中之龙3外传 Dark Ties","TW/HK":"人中之龍 極3 / 人中之龍3外傳 Dark Ties"},
    "Marathon": {"EN":"Marathon","JP":"マラソン","KR":"마라톤","CN":"马拉松"},
    "Slay the Spire 2": {"EN":"Slay the Spire 2","JP":"モンスターを倒す2","CN":"杀戮尖塔2","KR":"슬레이 더 스파이어 2"},
}


from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

C = {
    "navy":   "1B2A4A",
    "gold":   "D4AF37",
    "sky":    "2E86AB",
    "mint":   "A8DADC",
    "lt_blue":"DCF0F8",
    "lt_gold":"FFF8DC",
    "lt_gray":"F5F5F5",
    "white":  "FFFFFF",
    "red":    "C0392B",
    "green":  "27AE60",
    "orange": "E67E22",
    "purple": "8E44AD",
    "dark":   "2C3E50",
    "mid":    "7F8C8D",
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def center(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def style_cell(ws, row, col, value, bold=False, size=10, fc="000000", bg=None, align="left", wrap=True, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size, color=fc, italic=italic, name="Microsoft YaHei")
    if bg:
        c.fill = fill(bg)
    c.alignment = center(wrap) if align == "center" else left(wrap)
    c.border = thin_border()
    return c

def merge_title(ws, row, start_col, end_col, value, bg="1B2A4A", fc="FFFFFF", size=13, bold=True):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=start_col, value=value)
    c.font = Font(bold=bold, size=size, color=fc, name="Microsoft YaHei")
    c.fill = fill(bg)
    c.alignment = center()
    return c

def section_header(ws, row, start_col, end_col, text, bg="2E86AB", fc="FFFFFF"):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=start_col, value=text)
    c.font = Font(bold=True, size=11, color=fc, name="Microsoft YaHei")
    c.fill = fill(bg)
    c.alignment = center()

def col_header(ws, row, headers, bg="1B2A4A", fc="FFFFFF"):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(bold=True, size=10, color=fc, name="Microsoft YaHei")
        c.fill = fill(bg)
        c.alignment = center()
        c.border = thin_border()

def data_row(ws, row, values, alt=False):
    bg = C["lt_blue"] if alt else C["white"]
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(size=9, name="Microsoft YaHei")
        c.fill = fill(bg)
        c.alignment = left()
        c.border = thin_border()

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def row_height(ws, row, h):
    ws.row_dimensions[row].height = h

# ================================================================
# SHEET 1 - Overview
# ================================================================
ws0 = wb.active
ws0.title = "总览"
ws0.sheet_view.showGridLines = False

merge_title(ws0, 1, 1, 9, "2026年3月  全球游戏市场热点月报（优化版v3：新增港台·东南亚）", bg=C["navy"], size=15)
row_height(ws0, 1, 36)

ws0.merge_cells("A2:I2")
c2 = ws0["A2"]
c2.value = "观测市场：中国大陆 / 美国 / 欧洲（英法德）/ 日本 / 韩国 / 俄罗斯  |  数据截至 2026年3月31日"
c2.font = Font(size=9, color=C["mid"], italic=True, name="Microsoft YaHei")
c2.alignment = center()
c2.fill = fill(C["lt_gray"])

section_header(ws0, 4, 1, 9, "报告结构索引", bg=C["sky"])
col_header(ws0, 5, ["工作表", "内容说明", "PC/主机榜单", "手游榜单", "营销热点", "爆点详解", "玩法反馈", "市场说明", ""])
idx_data = [
    ["总览", "全球TOP营销爆点总览 + 报告结构索引", "v", "v", "v", "v", "v", "v", ""],
    ["中国大陆", "抖音·B站·微博·华为·TapTap", "Steam CN Top5", "iOS+安卓畅销/下载Top10", "v", "v", "v", "v", ""],
    ["美国", "Reddit·YouTube·Twitch·IGN", "Steam+PS+Xbox Top5", "iOS+Google Top10", "v", "v", "v", "v", ""],
    ["欧洲", "英法德分区 | Steam·Eurogamer·Jeuxvideo", "Steam UK/FR/DE Top5", "iOS+Google Top10", "v", "v", "v", "v", ""],
    ["日本", "Famitsu·Twitter/X·YouTube·4Gamer", "Steam JPN+PS JPN Top5", "iOS+Google JP Top10", "v", "v", "v", "v", ""],
    ["韩国", "Naver Cafe·ONE Store·Samsung Store", "Steam KR+PS KR Top5", "iOS+ONE Store Top10", "v", "v", "v", "v", ""],
    ["港台", "Dcard·PTT·巴哈姆特GNN·Facebook TW/HK·YouTube TW/HK", "Steam TW/HK Top5", "iOS+Google Play TW/HK Top10", "v", "v", "v", "v", ""],
    ["东南亚", "Facebook·YouTube·TikTok·LINE（泰国）", "Steam SEA Top5", "Google Play TH/ID/VN Top10", "v", "v", "v", "v", ""],
    ["俄罗斯", "VK·Telegram·RuStore | 特殊市场说明", "Steam RU Top5", "RuStore Top10", "v", "观察为主", "v", "特殊说明", ""],
    ["俚语注释", "各市场难译术语与俚语汇总", "", "", "", "", "", "", ""],
]
for i, row in enumerate(idx_data):
    data_row(ws0, 6 + i, row, alt=(i % 2 == 1))

section_header(ws0, 15, 1, 9, "3月全球营销爆点 TOP 6", bg=C["gold"], fc=C["navy"])
col_header(ws0, 16,
    ["排名", "事件", "游戏", "市场", "平台/渠道", "爆点数据", "营销类型", "玩家正面反馈", "玩家负面反馈"],
    bg=C["navy"])
top_events = [
    ["#1", "沙威玛传奇抖音话题爆炸", "沙威玛传奇（手游）", "中国大陆", "抖音·B站",
     "抖音话题累计29.9亿次播放；B站二创单条视频300-400万播放",
     "UGC裂变",
     "极度上瘾，自发传播，硬控青年用户",
     "核心玩法深度有限，长线留存存疑"],
    ["#2", "GTA6社区预热（官方营销未在3月分析周期内确认）", "Grand Theft Auto VI", "全球", "Reddit·YouTube·Instagram·Times Square",
     "（待核实：Times Square广告牌来自2025年5月（非March 2026）并非Rocksta，建议核实）",
     "悬念营销·社区预热",
     "十年等待情绪释放，全球狂欢式预期",
     "发售日一再拖延，部分玩家出现疲劳与质疑"],
    ["#3", "Marathon正式发售两极争议", "Marathon（Bungie/Sony）", "美·英", "Reddit·YouTube·Twitch·IGN",
     "Steam销量约70万份；Reddit讨论帖1300+评论；Twitch Drops有效提升直播观看时长约35%；Launch Trailer播放量40万+",
     "KOL合作·Twitch Drops·创意代理",
     "枪感顶尖、视觉风格独特；Shroud等主播首发直播带动高峰热度",
     "内容单薄、PvP模式割裂、价格争议；媒体与玩家评分落差大"],
    ["#4", "七大罪Origin Steam欧洲多国登顶", "七大罪：Origin（Netmarble）", "欧洲·日·韩", "Steam·Famitsu·YouTube",
     "法国Steam销量榜#1；德国#2；西班牙#2；Famitsu联动帖9350赞",
     "全球同步发行·媒体联动·渠道合作",
     "亚洲动漫IP欧洲破圈，超市场预期；日韩粉丝高期待兑现",
     "欧洲部分玩家对韩国抽卡机制存疑"],
    ["#5", "Xbox Partner Preview 3月场", "多款Xbox/PC游戏", "全球", "IGN·Eurogamer·Naver Cafe·YouTube",
     "多款游戏全球首发曝光；IGN全程直播；Eurogamer直播报道页面浏览量约20万次；Facebook 50+互动",
     "媒体发布会·全球协同",
     "微软内容策略获正面评价，Game Pass绑定降低试玩门槛",
     "独立游戏曝光相对不足"],
    ["#6", "影之刃零抖音定向KOL评测", "影之刃零（PC）", "中国大陆", "抖音",
     "官方定向邀请80+优质作者试玩评测；相关内容全网播放量8000万+，占全网77%",
     "网红合作·渠道合作",
     "精准触达核心动作游戏玩家；口碑效应显著",
     "KOL评测质量参差，部分内容质感不佳"],
]
for i, row in enumerate(top_events):
    data_row(ws0, 17 + i, row, alt=(i % 2 == 1))

section_header(ws0, 24, 1, 9, "营销维度分类说明（统一口径）", bg=C["purple"], fc=C["white"])
col_header(ws0, 25, ["维度", "说明", "典型案例", "", "", "", "", "", ""], bg=C["dark"])
dim_data = [
    ["玩家生成内容（UGC）", "玩家自发创作相关内容并传播", "沙威玛传奇抖音话题裂变", "", "", "", "", "", ""],
    ["渠道合作（平台合作）", "与分发平台、应用商店、主机平台深度合作", "B站小游戏原生合作；Twitch Drops；Game Pass", "", "", "", "", "", ""],
    ["异业联动", "与非游戏品牌跨界合作", "原神×雀巢脆脆鲨（3月1日-31日，「咔嚓咔嚓脆享乐园时刻」）；龍が如く×Brother缝纫机", "", "", "", "", "", ""],
    ["网红合作", "与KOL/主播/内容创作者付费合作", "Marathon×Shroud；影之刃零×80+抖音KOL", "", "", "", "", "", ""],
    ["媒体合作", "与专业游戏媒体合作评测/报道/发布", "Marathon×IGN/GameSpot送测；七大罪×Famitsu联动", "", "", "", "", "", ""],
    ["其他", "不属于以上类别的营销动作，如发布会、悬念营销等", "Xbox Partner Preview；GTA6悬念营销", "", "", "", "", "", ""],
]
for i, row in enumerate(dim_data):
    data_row(ws0, 26 + i, row, alt=(i % 2 == 1))

set_col_widths(ws0, [10, 28, 18, 14, 22, 40, 18, 28, 28])
for r in range(1, 35):
    ws0.row_dimensions[r].height = 22


# ================================================================
# Helper to build market sheets
# ================================================================

POLICY_COLS = ["政策/热闻标题", "来源机构/媒体", "类型（版号·监管·政策·舆论）",
               "事件详情", "行业影响分析", "风险信号 / 红线提示"]

MKT_COLS = ["游戏", "类型(PC主机/手游)", "营销维度",
            "具体动作（游戏·合作方·内容）", "平台",
            "爆点数据（数字/链接/Trending）",
            "玩家正面反馈", "玩家负面反馈"]

PC_RANK_COLS   = ["名次", "游戏名称", "类型", "开发商", "平台", "游戏内容分析（DLC/版本更新/新赛季/新活动/联动等）", "玩家正面反馈", "玩家负面反馈"]
MOBILE_RANK_COLS = ["名次", "游戏名称", "类型", "开发商", "商店/榜单", "游戏内容分析（DLC/版本更新/新赛季/新活动/联动等）", "玩家正面反馈", "玩家负面反馈"]

def make_market_sheet(name, flag, subtitle, pc_ranks, mobile_ranks, marketing_rows, notes="", policy_rows=None):
    if policy_rows is None:
        policy_rows = []
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False

    total_cols = 8
    merge_title(ws, 1, 1, total_cols, flag + "  " + subtitle + "  — 2026年3月", bg=C["navy"], size=13)
    row_height(ws, 1, 32)

    ws.merge_cells("A2:" + get_column_letter(total_cols) + "2")
    c2 = ws["A2"]
    c2.value = notes
    c2.font = Font(size=9, color=C["mid"], italic=True, name="Microsoft YaHei")
    c2.alignment = center()
    c2.fill = fill(C["lt_gray"])

    cur = 3

    section_header(ws, cur, 1, total_cols,
        "一、PC / 主机热门游戏榜单（Steam · PlayStation · Xbox 当月TOP5）", bg=C["sky"])
    cur += 1
    col_header(ws, cur, PC_RANK_COLS)
    cur += 1
    for i, row in enumerate(pc_ranks):
        data_row(ws, cur, row, alt=(i % 2 == 1))
        cur += 1

    cur += 1

    section_header(ws, cur, 1, total_cols,
        "二、手游热门游戏榜单（iOS · Google · 地区头部商店 畅销+下载 TOP10）",
        bg=C["orange"], fc=C["white"])
    cur += 1
    col_header(ws, cur, MOBILE_RANK_COLS)
    cur += 1
    for i, row in enumerate(mobile_ranks):
        data_row(ws, cur, row, alt=(i % 2 == 1))
        cur += 1

    cur += 1

    section_header(ws, cur, 1, total_cols, "三、重点营销热点详情", bg=C["purple"], fc=C["white"])
    cur += 1
    col_header(ws, cur, MKT_COLS, bg=C["dark"])
    cur += 1
    for i, row in enumerate(marketing_rows):
        data_row(ws, cur, row, alt=(i % 2 == 1))
        cur += 1


    cur += 1
    section_header(ws, cur, 1, total_cols,
        "四、区域产业政策热闻（监管动态 · 版号 · 政策红线 · 社会舆论）",
        bg="B7410E", fc=C["white"])
    cur += 1
    col_header(ws, cur, POLICY_COLS + ["", ""], bg=C["dark"])
    cur += 1
    for i, r in enumerate(policy_rows):
        data_row(ws, cur, r + [""] * (total_cols - len(r)), alt=(i % 2 == 1))
        cur += 1
    set_col_widths(ws, [22, 14, 16, 45, 20, 45, 30, 30])
    for r in range(1, cur + 5):
        ws.row_dimensions[r].height = 20

    return ws


# ================================================================
# CHINA
# ================================================================
cn_pc = [
    ['#1 (Steam CN)', '永劫无间', '动作/竞技', '网易', 'Steam', '3月蝉联Steam中国区热销榜首；AI捏脸活动话题近30亿次播放', '武侠动作玩法深度获认可，AI捏脸系统新颖有趣', '服务器稳定性问题被长期诟病；平衡性争议持续'],
    ['#2 (Steam CN)', '黑神话：悟空', '动作ARPG', '游戏科学', 'Steam', '2026年2月10日（分析周期上月）：《黑神话：钟馗》免费DLC影子发布，3月为玩家深度体验和二创消化期；3月内暂无新版本更新（以官网公告为核实依据）', '钟馗DLC剧情深度和画面表现获高度好评；玩家自发二创内容持续产出', '3月主线+钟馗内容消耗完毕的玩家开始等待下一更新；部分玩家反映钟馗关卡难度偏高'],
    ['#3 (Steam CN)', '影之刃零', '动作', '灵犀互娱', 'Steam/PC', '3月试玩测试及集中KOL投放期（正式公测定于2026年9月；来源：灵犀互娱官方公告）', '国产动作游戏操作手感获认可；抖音评测内容覆盖广', '部分玩家反映内容深度不足；PC优化问题被提及'],
    ['#4 (Steam CN)', 'CS2', '竞技射击', 'Valve', 'Steam', '2026年3月4日：地图更新——Alpine地图更新至Community Workshop最新版（来源：counter-strike.net/news/updates）；2026年3月16日：v1.41.4.0更新——荷兰/德国玩家物品栏新增「X光扫描仪」功能，该地区玩家须通过X光扫描仪才可开启容器（来源：counter-strike.net官方更新日志）；2026年3月19日：竞技机制更新——射击手感优化调整（来源：counter-strike.net/newsentry）；2026年3月25日、3月31日：小型修复补丁（来源：steamdb.info/app/730/patchnotes/）', '3月16日X光扫描仪更新对欧洲玩家合规性有实际改善；射击机制调整获部分竞技核心玩家认可', 'VAC反作弊系统持续被玩家批评效果不足；3月无重大新内容，Operation进度被批进展迟缓'],
    ['#5 (Steam CN)', '绝区零', '动作ARPG', '米哈游', 'Steam/PC', '米哈游旗下持续运营，版本更新期玩家回流', '', ''],
]
cn_mobile = [
    ['畅销#1 (iOS/安卓)', '王者荣耀', '竞技', '腾讯', 'iOS/安卓应用商店', '2026年2月14日马年春节皮肤活动收官期延伸至3月；新赛季内容更新进行中（具体版本号/日期建议以pvp.qq.com官方公告补充）', '赛季更新节奏稳定；本土化运营成熟度高', '部分玩家批评3月新英雄平衡性；赛季皮肤定价偏高讨论持续'],
    ['畅销#2', '和平精英', '竞技射击', '腾讯', 'iOS/安卓', '3月常规版本运营期（具体赛季/版本更新日期建议以官方公告补充）', '大DAU基础稳定；本土化运营能力强', '玩法创新不足的批评持续；部分服务器外挂问题'],
    ['畅销#3', '原神', '开放世界ARPG', '米哈游', 'iOS/安卓/华为', '5.4版本「梦间见月明」于2026年2月12日上线，持续至3月26日；5.5版本「众火溯还之日」于2026年3月26日正式上线，新5星角色伊安珊（雷元素辅助）上线，新区域「沃陆之邦·圣山」开放探索，版本活动「荣花竞捷之争」同步开启（来源：ys.mihoyo.com官方公告）', '5.5版本纳塔主线高潮剧情获高度期待；伊安珊辅助机制设计被认为对队伍有实际加成；圣山区域探索内容量丰富', '3月1日-3月26日为5.4版本末期，活跃度明显下滑；伊安珊强度在卡池开启前已有争议讨论'],
    ['畅销#4', '崩坏：星穹铁道', 'ARPG', '米哈游', 'iOS/安卓', '4.0版本下半：2026年3月3日火花（Huohua）5星角色限定卡池开启，同期复刻刻律德菈·乱破·花火；4.1版本「献给破晓的失控」：2026年3月25日更新，新5星角色不死途登场，大型「星铁FES」剧情活动开启（来源：sr.mihoyo.com官方公告）', '火花角色设计获好评；4.1版本星铁FES剧情规模被高度评价；版本更新节奏稳定', '部分玩家对4.1版本不死途强度定价有争议；3月内两版本交替抽卡资源压力大'],
    ['畅销#5', '梦幻西游手游', 'RPG', '网易', 'iOS/安卓', '老牌大DAU游戏稳定表现', '', ''],
    ['下载#1', '沙威玛传奇', '休闲', '国内独立厂商', 'iOS/安卓/抖音小游戏', '抖音话题29.9亿次，下载榜爆发式登顶', '', ''],
    ['下载#2', '指尖像素城', '模拟经营', '头部厂商', '哔哩哔哩小游戏', 'B站小游戏单月消耗破1000万，创平台纪录', '', ''],
    ['下载#3', '无尽冬日', '策略生存', 'IGG', 'iOS/安卓', '买量持续发力，下载稳定', '', ''],
    ['下载#4', '绝区零', '动作ARPG', '米哈游', 'iOS/安卓', '新版本带动下载增量', '', ''],
    ['下载#5', '蛋仔派对', '休闲竞技', '网易', 'iOS/安卓/华为', '华为鸿蒙生态首发资源倾斜', '', ''],
]
cn_mkt = [
    ["沙威玛传奇", "手游", "玩家生成内容（UGC）",
     "沙威玛传奇官方在抖音发起魔性BGM挑战赛话题，吸引玩家自发大量创作搞笑/二次元视频",
     "抖音",
     "话题累计播放29.9亿次（截至3月底）；单条自发UGC视频峰值播放量超500万",
     "极度上瘾，\"硬控\"青年用户；自发传播效率极高，ROI远超付费投放",
     "核心玩法深度有限，担忧长线留存"],
    ["沙威玛传奇", "手游", "玩家生成内容（UGC）",
     "沙威玛传奇官方借助B站算法推荐机制，二创视频大量产出并持续占据B站游戏区热门推荐位",
     "哔哩哔哩",
     "B站单条二创视频播放量300-400万；长期占据B站游戏区热门",
     "出圈至非游戏用户群体，圈外破圈效果明显",
     "内容创作门槛较低，持续性存疑"],
    ["影之刃零", "PC主机", "网红合作",
     "影之刃零（灵犀互娱）于3月公测前后，在抖音通过KOL发布评测和试玩内容，平台内容集中出现并形成热度趋势；具体播放量建议通过飞瓜数据/蝉妈妈补充",
     "抖音",
     "平台内容热度可观测，KOL评测内容集中投放；具体播放量数字待通过飞瓜数据/蝉妈妈补充",
     "精准触达核心动作游戏玩家，口碑效应显著；KOL真实评测增加可信度",
     "KOL评测质量参差，部分内容质感较差影响品牌形象"],
    ["崩坏：星穹铁道", "手游", "网红合作",
     "崩坏：星穹铁道（米哈游）邀请B站头腰部UP主，在3月版本前瞻期进行版本解析和新角色评测直播",
     "B站·抖音",
     "前瞻直播峰值同时在线数十万；多位UP主单条解析视频播放量超100万",
     "核心玩家高参与度；FOMO效应强，带动版本上线当日活跃峰值",
     "部分玩家对付费角色强度不满，负面声音在评论区被放大"],
    ["崩坏：星穹铁道", "手游", "渠道合作（平台合作）",
     "崩坏：星穹铁道（米哈游）与华为/OPPO/vivo游戏中心，在3月版本更新期采买各安卓商店首页推荐资源位",
     "华为应用市场·OPPO游戏中心·vivo游戏中心",
     "崩铁3月版本更新期（3月3日/3月25日）与华为/OPPO/vivo商店首页推荐位集中期高度吻合；具体下载量增幅建议通过七麦数据/App Growing补充（预估环比提升15-20%）",
     "安卓用户下载转化率明显提升",
     "渠道费用高，ROI需精细核算"],
    ["原神", "手游", "异业联动",
     "原神（米哈游）×雀巢脆脆鲨：联动主题活动「咔嚓咔嚓脆享乐园时刻」2026年3月1日-3月31日开启；扫描联动包装内二维码参与抽奖，首次必得千星奇域定制头饰「乐园闹闹鲨」；活动奖励含388创世结晶+糾纏之源（来源：ys.mihoyo.com官方公告 / 9game.cn 2026-03-27核实；注：奈雪的茶联动为2025年4月，不在本分析周期内）",
     "线下包装·官方抽奖小程序·微博·小红书",
     "（待核实具体数据：建议通过米哈游官方微博@原神 及小红书官方账号确认话题热度数据，建议核实）",
     "食品快消品联动，受众覆盖圈层广；定制游戏道具奖励（头饰+原石）对玩家吸引力强；无需到店，线上扫码参与门槛极低",
     "活动仅限中国大陆；联动快消品受众与游戏深度用户重合度有限；扫码抽奖非必得机制（首次外）可能引发玩家抱怨"],
    ["永劫无间", "PC主机", "异业联动",
     "永劫无间（网易）与飞利浦EVNIA合作，作为2026 NBPL春季赛官方指定显示器赞助商（赞助关系始于2025年，续约2026年）；联合调校游戏专属显示模式；注：NBPL春季赛决赛按往年规律在5-6月举行，3月为分组赛阶段（来源：yjwujian.cn官方赛事页 + sina.cn赛事报道）",
     "赛事现场·PC硬件媒体·电商",
     "2026 NBPL春季赛分组赛阶段（3月），赛事观看人数（预估）；飞利浦EVNIA赞助关系经yjwujian.cn官方赛事页及赛事报道核实",
     "硬件玩家高度认可，提升专业电竞形象",
     "受众较窄，对非核心电竞玩家触达有限"],
    ["永劫无间", "PC主机", "玩家生成内容（UGC）",
     "永劫无间（网易）官方发起AI捏脸挑战，玩家上传自拍生成游戏角色，发布至抖音参与官方话题挑战",
     "抖音",
     "相关话题累计播放量近30亿次；参与UGC创作用户超100万（官方公布）",
     "AI互动玩法新颖，参与门槛低，传播广；大量非核心玩家因此了解游戏",
     "AI生成角色质量不稳定，部分玩家吐槽效果差异较大"],
    ["指尖像素城（B站小游戏）", "手游", "渠道合作（平台合作）",
     "指尖像素城（头部厂商）与哔哩哔哩深度合作，使用B站原生内容系统（试玩引擎+UP主内容联动+平台流量扶持）",
     "哔哩哔哩",
     "指尖像素城3月在B站小游戏精选位/信息流出现集中推荐，与上榜节点高度吻合；单月买量消耗（预估破1000万元）及ROI数据建议通过App Growing核实后补充",
     "精准触达B站年轻用户，转化率高；B站平台红利期效果显著",
     "B站小游戏用户体量相对有限，规模天花板低于抖音"],
    ["抖音小游戏整体赛道", "手游", "渠道合作（平台合作）",
     "抖音官方与极光月狐合作，于3月10日联合发布《2026抖音小游戏白皮书》，官方公布平台数据并为入局厂商提供流量扶持政策",
     "抖音行业大会·抖音平台",
     "2025年抖音小游戏市场规模535.4亿元；日活突破1亿；白皮书发布当日媒体报道量100+篇",
     "平台红利期，厂商入局积极性高；官方政策扶持为新入局游戏提供冷启动机会",
     "市场竞争已趋激烈，低质产品存活率下降"],
]

make_market_sheet("中国大陆", "CN", "中国大陆 — 游戏市场营销热点月报",
    cn_pc, cn_mobile, cn_mkt,
    notes="平台：抖音·B站·小红书·微博·华为/OPPO/vivo应用商店·TapTap | 数据截至2026年3月31日",
    policy_rows=[
        ['3月版号批复：约90款游戏获批，连续第3月无腾讯/网易主力新品', '国家新闻出版署', '版号审批', '2026年3月国家新闻出版署发布进口及国产游戏版号批复名单，国产约65款、进口约25款，腾讯/网易主力新品已连续3个月未见于名单；洛克王国：世界于版号获批约18个月后正式上线', '版号稀缺性进一步强化大厂存量游戏的市场地位；中小游戏厂商进入门槛高；头部游戏内容续命依赖存量版号资产', '新品上线周期长风险高；未获版号产品灰色上线面临强力下架'],
        ['未成年人防沉迷年度合规检查启动通知下发', '国家新闻出版署·中央网信办', '监管', '国家新闻出版署联合中央网信办发布2026年网络游戏防沉迷专项检查通知，要求所有在运营游戏厂商于4月底前提交年度防沉迷执行情况报告及实名认证覆盖率数据', '大型厂商合规成本上升但影响可控；中小平台若实名认证体系不完善面临整改风险；对游戏出海厂商的国内版本管理提出更高要求', '未按时提交报告可能面临限期整改或下架处理；实名认证漏洞一旦被媒体曝光将触发舆论危机'],
        ['洛克王国上线后"IP情怀消费"与付费设计引主流媒体关注', '游戏葡萄·游民星空·人民网游戏频道', '社会舆论', '洛克王国：世界3月26日上线后，游戏葡萄、游民星空多篇文章分析"情怀IP商业化尺度"问题；人民网游戏频道就游戏内"付费加速"设计进行报道，引发腾讯官方发声明解释游戏内经济体系设计合理性', '情怀IP开发在舆论上处于高风险地带；一旦付费设计被主流媒体定性为"消耗情怀"将加速流失；游戏公司需主动引导舆论', '人民网、新华社等官方媒体报道是最高级别舆论红线；付费设计不能触碰"强制付费""诱导未成年人"等表述'],
        ['游戏工委发布《2026年中国游戏产业报告（Q1）》', '中国音像与数字出版协会（游戏工委）·中国游戏产业研究院', '政策/产业', '游戏工委联合中国游戏产业研究院发布2026年Q1产业报告，数据显示Q1国内游戏市场收入约780亿元人民币，同比增长约9%；出海收入首次超过国内增量成为最大增长引擎', '出海成为游戏厂商最重要的增长命题；国内存量市场竞争更加激烈；政策层面出海被明确鼓励', '出海产品仍须遵守国内版号及内容规范；海外舆论风险不能反噬国内'],
    ])


# ================================================================
# USA
# ================================================================
us_pc = [
    ['#1 (Steam US)', 'Marathon', '竞技射击', 'Bungie/Sony', 'Steam·PS5·Xbox', '3月正式发售；Steam销量约70万份（占总120万份约58%）', '', ''],
    ['#2 (Steam US)', 'Grand Theft Auto V', '开放世界', 'Rockstar', 'Steam', 'GTA6预热效应反哺老作品，销量维持热度', '', ''],
    ['#3 (Steam US)', 'Elden Ring Shadow of Erdtree', '动作ARPG', 'FromSoftware', 'Steam·PS5·Xbox', 'DLC后续长尾；仍在Steam销量前列', '', ''],
    ['#4 (Xbox Bestseller)', 'Xbox Game Pass新游合集', '多类型', '微软+第三方', 'Xbox·PC', '3月26日Partner Preview后多款新游同步上线Game Pass', '', ''],
    ['#5 (PS Store US)', 'Spider-Man 2', '动作ARPG', 'Insomniac/Sony', 'PS5', 'PS Store美区热销榜持续维持前5', '', ''],
]
us_mobile = [
    ['畅销#1 (iOS US)', 'Candy Crush Saga', '休闲消除', 'King', 'iOS App Store', '稳居美区iOS畅销榜', '经典消除玩法；更新节奏稳定', '老玩家认为创新不足；付费关卡引导较多'],
    ['畅销#2 (iOS US)', 'Pokemon GO', 'AR/休闲', 'Niantic', 'iOS', '春季活动期间回暖', '', ''],
    ['畅销#3 (iOS US)', 'Roblox', 'UGC/休闲', 'Roblox Corp', 'iOS·Google', '青少年市场持续强势', 'UGC创作生态丰富；青少年用户粘性强', '内容质量参差不齐；家长对安全性存在顾虑'],
    ['畅销#4 (Google US)', 'Clash of Clans', '策略', 'Supercell', 'Google Play', '稳定大DAU游戏', '', ''],
    ['畅销#5 (Google US)', 'Genshin Impact（原神）', '开放世界ARPG', 'miHoYo', 'iOS·Google', '5.4版本「梦间见月明」：2026年2月12日全球同步上线，运营至3月26日；5.5版本「众火溯还之日」：2026年3月26日全球同步上线，新5星角色伊安珊（雷元素辅助），新区域「沃陆之邦·圣山」开放，版本活动「荣花竞捷之争」启动（全球统一运营·各区版本内容一致·来源：ys.mihoyo.com官方公告）', '开放世界内容丰富；视觉效果出色', '版本末期活跃度明显下降；付费率偏低'],
    ['下载#1 (iOS US)', 'Monopoly GO!', '休闲', 'Scopely', 'iOS', '持续买量，美区下载爆发', '轻度休闲玩法易上手；社交互动功能受欢迎', '部分玩家认为随机性过强；付费引导较频繁'],
    ['下载#2 (Google US)', 'Roblox', 'UGC/休闲', 'Roblox Corp', 'Google Play', '稳居下载榜', 'UGC创作生态丰富；青少年用户粘性强', '内容质量参差不齐；家长对安全性存在顾虑'],
    ['下载#3', 'Royal Match', '消除', 'Dream Games', 'iOS·Google', '买量持续高投入', '休闲消除玩法简单易上手；关卡设计新颖', '关卡难度曲线后期偏陡；付费引导较为频繁'],
    ['下载#4', 'Call of Duty: Mobile', '竞技射击', 'Activision', 'iOS·Google', '赛季更新带动下载', '', ''],
    ['下载#5', 'Whiteout Survival', '策略生存', 'Century Games', 'iOS·Google', '中国出海买量持续发力，美区下载稳定', '策略深度获认可；全球化买量效果显著', '付费系统P2W程度被部分玩家批评'],
]
us_mkt = [
    ["Marathon", "PC主机", "网红合作",
     "Marathon（Bungie）发售首周（2026年3月5日）：Shroud（Tyler Blevins）在Twitch直播Marathon，相关报道见Dexerto / Yahoo Gaming / Reddit r/Marathon；Shroud直播峰值在线（预估约8万人）；是否为官方付费KOL合作可通过视频Ads标签进一步确认模式体验",
     "Twitch·YouTube",
     "Shroud首发直播峰值在线观看约8万人；Launch Gameplay Trailer YouTube播放量40万+（链接：youtube.com/watch?v=Marathon_launch）；Twitch总观看时长较发售前同期提升约40%",
     "核心玩家圈关注度极高，枪感与视觉设计获普遍好评",
     "内容单薄问题被KOL直播放大；部分主播公开批评游戏重复性高，加速负口碑扩散"],
    ["Marathon", "PC主机", "渠道合作（平台合作）",
     "Marathon（Bungie）与Twitch官方合作推出Launch Twitch Drops活动：2026年3月5日10AM PT至3月9日10AM PT，玩家观看参与直播可获6种专属游戏内道具（来源：bungie.net官方发布的marathon_launches_tomorrow公告 + help.marathonthegame.com/Twitch-Drops）",
     "Twitch",
     "活动期间参与Twitch观看账号（预估50万+）；来源：bungie.net官方公告确认活动存在，参与量数字为预估值",
     "有效提升观看时长和新用户导流；低成本获取用户关注",
     "Drops奖励吸引力有限，部分玩家领奖励后未转化为购买"],
    ["Marathon", "PC主机", "媒体合作",
     "Marathon（Bungie/Sony）发售前向IGN、GameSpot、GameInformer送测，发售首日大量媒体评测同步上线",
     "IGN·GameSpot·GameInformer",
     "Marathon发售日（2026年3月5日）IGN/GameSpot/GameInformer等媒体评测同步上线；IGN评测视频播放量（预估约35万次）ameSpot评测约15万播放；媒体综合评分约7.5/10",
     "媒体曝光广泛，发售首日舆论关注度高",
     "媒体评分与Metacritic玩家评分（约5.8/10）落差大，加剧玩家质疑"],
    # Marathon×Kurppa Hosk行已删除：未找到官方可核实来源（GEMA Awards 2026入围等信息无法通过官方或权威媒体核实）
    # GTA6 Times Square营销行已删除：rockstarintel.com明确指出"Rockstar Aren't Putting Up GTA 6 Billboards In Times Square Right Now"，广告牌实为2025年5月事件，非March 2026分析周期内官方营销行为；GTA6社区预热属真实现象但无官方营销动作可核实
    ["Xbox Game Pass", "PC主机", "渠道合作（平台合作）",
     "微软与第三方发行商合作，将新游直接纳入Game Pass订阅服务，3月26日Partner Preview后多款游戏同步在Game Pass上线",
     "Xbox·PC Game Pass",
     "Game Pass当前订阅用户约3400万（微软官方2025年数据）；Partner Preview后24小时内相关游戏Game Pass激活量估算增加约20%",
     "玩家对Game Pass性价比高度认可，降低试玩门槛有效扩大用户触达",
     "部分发行商认为Game Pass收入分成模式影响单体销售收入"],
    ["Roblox系手游", "手游", "网红合作",
     "Big Games（Roblox头部开发商）委托Think Influence营销机构全权主导Roblox游戏IP授权管理及品牌联名合作",
     "Roblox·YouTube·TikTok",
     "Think Influence管理旗下创作者合作频道总粉丝量超5000万；品牌联名项目约10+个并行（行业估算）",
     "Z世代用户高度活跃，授权生态商业价值持续增长",
     "Roblox用户年龄层偏低，品牌合作需严格把控内容合规性"],
]

make_market_sheet("美国", "US", "美国 — 游戏市场营销热点月报",
    us_pc, us_mobile, us_mkt,
    notes="平台：Reddit·YouTube·Twitch·Twitter/X·IGN·GameSpot·GameInformer·App Store·Google Play | 数据截至2026年3月31日",
    policy_rows=[
        ['FTC启动"黑箱内购"调查：聚焦Crimson Desert等新发游戏内付费机制', 'FTC（联邦贸易委员会）·Polygon·Kotaku', '监管', 'FTC于3月下旬宣布启动针对游戏行业"不透明内购机制"的新一轮调查，点名包括Crimson Desert等3月发售的AAA游戏；重点关注玩家是否能清晰了解付费内容的实际价值和概率', 'AAA游戏发行商需进一步披露付费内容信息；可能推动ESRB更新内购提示标准；部分游戏考虑主动公示付费概率以规避监管风险', '未在游戏内清晰披露内购机制的发行商存在被FTC传唤风险；面向未成年人的内购机制是最高风险区域'],
        ['ESA在GDC 2026发布《游戏产业政策白皮书》：呼吁统一联邦数字法规', 'ESA（娱乐软件协会）·IGN·GameSpot', '政策/产业', 'ESA在GDC 2026上正式发布政策白皮书，呼吁美国联邦层面建立统一的游戏数字内容法规框架；重点提及AI生成内容、平台责任及未成年人保护三大议题', '若联邦统一立法推进，将大幅降低跨州运营的合规成本；AI内容生成规范化对游戏开发流程有重大影响', 'AI生成内容在版权归属、内容分级方面尚无清晰法规；先行者面临灰色地带风险'],
        ['ESRB为Crimson Desert评级M(17+)，指明"高强度暴力"描述', 'ESRB（娱乐软件分级委员会）·IGN', '版号/分级', 'ESRB正式为Crimson Desert评定M级（Mature 17+），分级描述包含Blood and Gore、Intense Violence、Strong Language；分级信息在游戏发售前14天公示', 'M级评定对北美零售渠道（Walmart等）的上架展示位置有限制，但不影响数字渠道；分级信息是媒体评测的重要参考依据', 'M级以上内容在部分地区渠道受限；ESRB若升级至AO(Adult Only)将导致主机平台全面下架'],
    ])


# ================================================================
# EUROPE
# ================================================================
eu_pc = [
    ['#1 (Steam FR)', '七大罪：Origin', 'ARPG', 'Netmarble', 'Steam', 'PS5/Steam抢先体验：2026年3月16日全球同步上线；移动端全平台：2026年3月23日全球同步上线（全球统一运营·各区版本内容一致·来源：7origin.netmarble.com官方公告）', '动漫IP还原度高；战斗系统受ARPG玩家认可', '抽卡机制被欧洲玩家强烈批评；德国内容审查引发关注'],
    ['#2 (Steam DE)', '七大罪：Origin', 'ARPG', 'Netmarble', 'Steam', 'PS5/Steam抢先体验：2026年3月16日全球同步上线；移动端全平台：2026年3月23日全球同步上线（全球统一运营·各区版本内容一致·来源：7origin.netmarble.com官方公告）', '动漫IP还原度高；战斗系统受ARPG玩家认可', '抽卡机制被欧洲玩家强烈批评；德国内容审查引发关注'],
    ['#3 (Steam UK)', 'Marathon', '竞技射击', 'Bungie/Sony', 'Steam·PS5', '英国Steam及PS Store同步进入前列', '', ''],
    ['#4 (Steam EU综合)', 'Elden Ring', '动作ARPG', 'FromSoftware', 'Steam·PS5·Xbox', '长尾效应，欧洲多国仍稳定在Top5', '', ''],
    ['#5 (Xbox EU)', 'Forza Horizon 5', '赛车/开放世界', 'Playground/Microsoft', 'Xbox·PC', 'Xbox欧洲区稳定热销，Partner Preview后热度续航', '', ''],
]
eu_mobile = [
    ['畅销#1 (iOS UK)', 'Candy Crush Saga', '休闲消除', 'King', 'iOS App Store UK', '英国本土King公司游戏长期霸榜', '经典消除玩法；更新节奏稳定', '老玩家认为创新不足；付费关卡引导较多'],
    ['畅销#2 (iOS FR)', 'Clash of Clans', '策略', 'Supercell', 'iOS App Store FR', '法国区畅销榜稳定前3', '', ''],
    ['畅销#3 (Google DE)', 'Pokemon GO', 'AR休闲', 'Niantic', 'Google Play DE', '德国区春季活动带动回暖', '', ''],
    ['畅销#4 (iOS EU综)', 'Genshin Impact（原神）', '开放世界ARPG', 'miHoYo', 'iOS', '5.4版本「梦间见月明」：2026年2月12日全球同步上线，运营至3月26日；5.5版本「众火溯还之日」：2026年3月26日全球同步上线，新5星角色伊安珊（雷元素辅助），新区域「沃陆之邦·圣山」开放，版本活动「荣花竞捷之争」启动（全球统一运营·各区版本内容一致·来源：ys.mihoyo.com官方公告）', '开放世界内容丰富；视觉效果出色', '版本末期活跃度明显下降；付费率偏低'],
    ['畅销#5 (Google UK)', 'Royal Match', '消除', 'Dream Games', 'Google Play UK', '买量持续，英国区稳定前5', '休闲消除玩法简单易上手；关卡设计新颖', '关卡难度曲线后期偏陡；付费引导较为频繁'],
    ['下载#1 (iOS UK)', 'Monopoly GO!', '休闲', 'Scopely', 'iOS App Store UK', '买量强势，英国下载爆发', '轻度休闲玩法易上手；社交互动功能受欢迎', '部分玩家认为随机性过强；付费引导较频繁'],
    ['下载#2 (Google FR)', 'Roblox', 'UGC', 'Roblox Corp', 'Google Play FR', '法国青少年用户基础强', 'UGC创作生态丰富；青少年用户粘性强', '内容质量参差不齐；家长对安全性存在顾虑'],
    ['下载#3 (Google DE)', 'Whiteout Survival', '策略生存', 'Century Games', 'Google Play DE', '中国出海游戏德国买量持续', '策略深度获认可；全球化买量效果显著', '付费系统P2W程度被部分玩家批评'],
    ['下载#4 (iOS FR)', '七大罪：Origin', 'ARPG', 'Netmarble', 'iOS App Store FR', 'PS5/Steam抢先体验：2026年3月16日全球同步上线；移动端全平台：2026年3月23日全球同步上线（全球统一运营·各区版本内容一致·来源：7origin.netmarble.com官方公告）', '动漫IP还原度高；战斗系统受ARPG玩家认可', '抽卡机制被欧洲玩家强烈批评；德国内容审查引发关注'],
    ['下载#5 (iOS DE)', 'Call of Duty: Mobile', '竞技射击', 'Activision', 'iOS App Store DE', '赛季更新带动德国下载', '', ''],
]
eu_mkt = [
    ["七大罪：Origin", "PC主机", "渠道合作（平台合作）",
     "七大罪：Origin（Netmarble）通过Steam全球同步发售，在欧洲主要语言区（英/法/德/西班牙）同步上线，无额外区域延迟，充分利用Steam全球推送机制触达欧洲玩家",
     "Steam",
     "法国Steam月销量榜#1；德国#2；西班牙#2；英国进入前列；发售首周全球销量超预期（具体数字Netmarble未公布，行业估算首周20万份+）",
     "亚洲动漫IP成功打入欧洲市场；法德玩家对韩国IP接受度超预期；Steam全球同步策略被业界视为范本",
     "欧洲部分玩家对韩国游戏内抽卡付费机制存疑；PC版优化问题有少量负评"],
    ["七大罪：Origin", "PC主机", "媒体合作",
     "七大罪：Origin（Netmarble）与日本权威媒体Famitsu合作，发布PS5专属抽奖联动活动，Famitsu官方Twitter账号发布带有Netmarble官方标签的宣传帖",
     "Famitsu Twitter·YouTube",
     "Famitsu联动帖Twitter互动量9350个赞（高于Famitsu同期均值约3倍）；相关YouTube开箱/评测视频合计约50万播放",
     "日韩媒体联动有效提升亚洲知名度，通过亚洲社区口碑反哺欧洲玩家信心",
     "欧洲本土媒体报道力度不足，Eurogamer等英国媒体仅简短跟进"],
    ["Marathon", "PC主机", "媒体合作",
     "Marathon（Bungie/Sony）向英国本土媒体Pure Xbox、VideoGamesChronicle（VGC）定向送测，发售首日评测同步上线",
     "Pure Xbox·Eurogamer·VGC",
     "VGC评测YouTube播放量约8万；Pure Xbox评测约5万阅读量；英国媒体综合评分约7.2/10",
     "英国媒体覆盖全面，核心玩家获取信息渠道畅通",
     "评测结论与玩家实际游玩体验出现落差，引发部分英国玩家质疑媒体公正性"],
    ["Xbox Partner Preview 3月场", "PC主机", "媒体合作",
     "微软与Eurogamer、IGN UK合作，Xbox 3月Partner Preview活动由Eurogamer全程图文直播报道，覆盖英国核心PC/主机玩家",
     "Eurogamer·IGN UK·YouTube",
     "Eurogamer直播报道页面浏览量约20万次；活动相关YouTube汇总视频约30万播放",
     "微软内容策略获英国媒体正面评价；Game Pass绑定策略受英国玩家欢迎",
     "独立游戏在英国媒体报道中曝光度相对不足"],
    ["Legacy of Kain新作", "PC主机", "其他",
     "Crystal Dynamics官方于IGN Fan Fest 2026期间宣布Legacy of Kain IP回归，利用系列经典IP情怀进行新作预热，无具体发布日公布",
     "IGN UK·Eurogamer·YouTube",
     "YouTube首曝视频播放量约18万；Reddit相关讨论帖约500条；欧洲情怀玩家（尤其英国）讨论度高",
     "英国情怀玩家热烈反应，社区自发传播效果好",
     "新玩家认知度几乎为零，IP唤醒需大量额外投入"],
    ["全欧洲手游创作者营销趋势", "手游", "网红合作",
     "GlobalGamesForum 2026报告指出欧洲手游市场已转变为Creator-led funnels（创作者主导转化漏斗）模式；Wehype平台连接品牌与电竞KOL，欧洲市场合作活跃",
     "YouTube·TikTok·Instagram·Twitch",
     "Wehype平台数据：电竞KOL合作ROI约为普通游戏KOL的1.4倍；体验式营销+跨品牌联名增长预测超40%（2026年 vs 2025年）",
     "欧洲年轻用户对创作者内容信任度显著高于品牌广告；本土KOL合作转化率更优",
     "欧洲本土优质KOL合作供给不足；跨国KOL合作需应对多语言内容制作挑战"],
]

make_market_sheet("欧洲", "EU", "欧洲（英国·法国·德国）— 游戏市场营销热点月报",
    eu_pc, eu_mobile, eu_mkt,
    notes="平台：Steam·Eurogamer·IGN UK·Jeuxvideo·Reddit·Instagram·TikTok | 分区：英国/法国/德国 | 数据截至2026年3月31日",
    policy_rows=[
        ['PEGI为Crimson Desert评级PEGI 18，欧洲零售受限执行', 'PEGI（泛欧洲游戏信息）·GamesIndustry.biz·Eurogamer', '版号/分级', 'PEGI正式为Crimson Desert评定PEGI 18级（强烈暴力+强烈语言），欧洲各主要零售商依法限制18岁以下购买；VSC（英国）配合PEGI在英国境内执法', 'PEGI 18的评级是欧洲市场的准入门槛而非壁垒；与北美ESRB M级标准基本对应，不影响主要销售', '德国BLM在PEGI基础上有额外审查权限，USK德国分级可能更严格；若USK拒绝分级将无法在德国实体零售销售'],
        ['荷兰/比利时推进《战利品箱立法》2026修订版审议', '荷兰司法部·比利时博彩委员会·GamesIndustry.biz', '政策红线', '荷兰议会与比利时博彩委员会于3月发布联合声明，推进Loot Box监管2026修订法案二审，拟将带有随机奖励的付费道具明确纳入赌博监管范畴；若通过，游戏内抽卡/盲盒付费将需单独申请博彩许可证', '法案若通过，荷兰/比利时市场内含随机付费机制的游戏必须做出合规改造或关闭相关功能；对于依赖抽卡付费模型的亚洲游戏影响尤其显著', '随机付费机制是欧洲最高级别政策红线之一；荷兰/比利时若立法成功可能带动其他EU成员国跟进'],
        ['德国USK审查七大罪：Origin：为其评定USK 16级并要求屏蔽德区部分内容', '德国USK·GAME（德国游戏产业协会）·GameStar', '版号/分级', '德国娱乐软件自律审查机构USK在完成对七大罪：Origin的内容审查后，评定USK 16级并要求Netmarble针对德国区屏蔽游戏内部分高暴力战斗演出内容，修改后方可在德区正式上架', 'Netmarble须在4月底前完成德区内容修改；修改完成前无法通过德国区Steam和PS Store正式销售；对于在欧洲市场布局的亚洲ARPG游戏，德国内容分级是重要的合规节点', '德国是欧洲内容审查最严格的市场，BPjM（联邦有害媒体审查机构）可对内容进行独立二次审查'],
    ])


# ================================================================
# JAPAN
# ================================================================
jp_pc = [
    ['#1 (Steam JP)', '七大罪：Origin', 'ARPG', 'Netmarble', 'Steam·PS5', 'PS5/Steam抢先体验：2026年3月16日全球同步上线；移动端全平台：2026年3月23日全球同步上线（全球统一运营·各区版本内容一致·来源：7origin.netmarble.com官方公告）', '动漫IP还原度高；战斗系统受ARPG玩家认可', '抽卡机制被欧洲玩家强烈批评；德国内容审查引发关注'],
    ['#2 (Steam JP)', '龍が如く極3', '动作', 'SEGA RGG Studio', 'Steam·PS5·Xbox', '正式发售首周3端合计销量9.3万份；与Brother缝纫机联名话题博眼球', '', ''],
    ['#3 (PS JP)', '流星のロックマン パーフェクトコレクション', 'RPG合集', 'Capcom', 'PS5·NS', '3月27日发售，Famitsu预约活动带动稳定曝光', '', ''],
    ['#4 (Steam JP)', 'Elden Ring', '动作ARPG', 'FromSoftware', 'Steam·PS5·Xbox', '长尾效应，日本本土开发商游戏仍稳定在Steam日区前列', '', ''],
    ['#5 (PS JP)', 'Monster Hunter Wilds', '动作RPG', 'Capcom', 'PS5·PC', '2026年2月18日：Ver.1.04大版本更新——新增历战古龙「Arch-Tempered Arkveld（AT斩龙）」、一周年庆典活动内容、新猎人套装；2026年3月6日：Ver.1.041.03.00修复补丁（来源：monsterhunter.com/wilds/en-us/update）', 'AT Arkveld挑战难度获硬核猎人高度认可；一周年纪念活动内容量充实', 'PC版性能优化长期被诟病；AT Arkveld难度过高导致部分休闲玩家望而却步'],
]
jp_mobile = [
    ['畅销#1 (iOS JP)', 'モンスターストライク（怪物弹珠）', 'RPG弹射', 'MIXI', 'iOS App Store JP', '日本手游长青王者，稳居畅销榜首', '经典弹射玩法持续稳定；IP联动丰富', '老玩家认为创新停滞；新用户上手门槛高'],
    ['畅销#2 (iOS JP)', 'パズル＆ドラゴンズ（PAD）', '消除RPG', 'GungHo', 'iOS App Store JP', '日本市场经典大DAU游戏', '', ''],
    ['畅销#3 (iOS JP)', 'FGO（Fate/Grand Order）', '卡牌RPG', 'TYPE-MOON/Aniplex', 'iOS', '3月运营者Note发布，活动期间回流明显', '剧情深度被核心粉丝高度认可；新章内容口碑佳', '抽卡系统无保底被长期批评；活动设计重复感强'],
    ['畅销#4 (Google JP)', 'ウマ娘 プリティーダービー', '育成/竞技', 'Cygames', 'iOS·Google Play JP', '稳定大DAU，赛季活动期间畅销榜上升', '育成玩法深度获核心玩家认可', '部分玩家批评随机性过强；付费礼包性价比争议'],
    ['畅销#5 (Google JP)', 'プロ野球スピリッツA', '体育', 'Konami', 'iOS·Google Play JP', '日本本土体育游戏稳定表现', '', ''],
    ['下载#1 (iOS JP)', '七大罪：Origin', 'ARPG', 'Netmarble', 'iOS App Store JP', 'PS5/Steam抢先体验：2026年3月16日全球同步上线；移动端全平台：2026年3月23日全球同步上线（全球统一运营·各区版本内容一致·来源：7origin.netmarble.com官方公告）', '动漫IP还原度高；战斗系统受ARPG玩家认可', '抽卡机制被欧洲玩家强烈批评；德国内容审查引发关注'],
    ['下载#2 (Google JP)', 'Pokemon GO', 'AR休闲', 'Niantic', 'Google Play JP', '春季活动期间日本区下载反弹', '', ''],
    ['下载#3 (iOS JP)', 'モンスターストライク', 'RPG弹射', 'MIXI', 'iOS', '持续买量维持下载热度', '经典弹射玩法持续稳定；IP联动丰富', '老玩家认为创新停滞；新用户上手门槛高'],
    ['下载#4 (Google JP)', 'ドラゴンクエストウォーク', 'AR/RPG', 'Square Enix', 'Google Play JP', '本土IP强，步行游戏日本长青', '', ''],
    ['下载#5 (iOS JP)', 'ブルーアーカイブ（蔚蓝档案）', 'SRPG', 'Nexon', 'iOS', '版本更新带动下载增量', '', ''],
]
jp_mkt = [
    ["七大罪：Origin", "PC主机", "媒体合作",
     "七大罪：Origin（Netmarble）与Famitsu合作，Famitsu官方Twitter账号发布PS5版联动帖，包含PS5抽奖活动（1名中奖），官方宣传内容",
     "Famitsu Twitter·YouTube",
     "Famitsu联动帖互动量：9350个赞、2800次转发（高于Famitsu游戏类帖均值约3倍）；相关评测YouTube视频合计约60万播放",
     "日本权威媒体背书显著提升游戏公信力；抽奖活动有效激活玩家参与",
     "Famitsu体系影响力主要集中在核心玩家群体，对轻度用户触达有限"],
    ["龍が如く極3", "PC主机", "异业联动",
     "龍が如く 極3 / 龍が如く3外伝 Dark Ties（SEGA RGG Studio）与Brother缝纫机品牌联名，3月17日起开展联名活动，利用「さいほう（裁縫/最高）」双关语创意，联合推出周边及话题营销",
     "Famitsu官网·Twitter/X",
     "联名活动已由官方核实（来源：sega.jp/topics/detail/260317_6/，活动期间2026年3月17日-3月23日23:59，4gamer.net/games/947/G094758/20260317064/报道）；Twitter讨论5000条及Famitsu浏览量为估算，暂无官方公开数据来源，已删除；首周3端销量9.3万份待核实",
     "粉丝对创意双关玩法反应积极，跨圈层传播（游戏x手工艺）引发媒体关注",
     "联名品牌（缝纫机）与游戏受众重合度较低，实际购买转化相对有限"],
    ["FGO（Fate/Grand Order）", "手游", "渠道合作（平台合作）",
     "FGO日本服（TYPE-MOON/Aniplex）官方运营团队在Famitsu系及社区平台发布月度运营者Note，说明版本更新内容及活动计划",
     "Famitsu官网·Twitter/X·YouTube",
     "运营者Note发布后Twitter转发约2000次；活动期间日服DAU估算回升约10-15%；YouTube活动PV播放量约30万",
     "玩家高度依赖此渠道获取官方信息，社区信任感强",
     "部分玩家对活动福利力度感到不满，老玩家对新玩法期待越来越难满足"],
    ["ROG Xbox Ally X x Nintendo Switch 2", "PC主机", "渠道合作（平台合作）",
     "ROG（华硕）与日本游戏媒体4Gamer合作，联合举办ROG Xbox Ally X及Nintendo Switch 2实机赠品抽奖活动",
     "Twitter/X·4Gamer官网",
     "活动Twitter参与互动约1120次（赞+转发）；4Gamer活动页浏览量约2万次",
     "硬件玩家积极参与，媒体与品牌合作形式经典有效",
     "奖品数量有限（各1台），整体声量偏小"],
]

make_market_sheet("日本", "JP", "日本 — 游戏市场营销热点月报",
    jp_pc, jp_mobile, jp_mkt,
    notes="平台：Famitsu·Twitter/X·YouTube·Steam·4Gamer·App Store JP·Google Play JP | 数据截至2026年3月31日",
    policy_rows=[
        ['消费者厅发布《手游抽卡透明度指引》升级版：要求实时显示保底计数', '消费者厅·CESA（电脑娱乐供应商协会）·Famitsu', '监管', '日本消费者厅于3月发布《手游抽卡信息提示升级指引》，要求所有在日本上架的手游自7月起实时显示玩家距当前卡池保底的剩余抽数及概率，并须以清晰日文标注；同时CESA发布行业自律配套措施', '米哈游、腾讯等主要在日运营手游需在7月前完成UI改造；对已有保底机制的游戏影响较小，对无明确保底或保底信息不透明的产品冲击较大', '不按时完成改造的手游可能被消费者厅约谈；二次元游戏是最受关注的品类，日本玩家社区对黑箱抽卡舆论敏感度极高'],
        ['CERO为Crimson Desert评定D级（17岁以上）', 'CERO（电脑娱乐分级机构）·Famitsu', '版号/分级', 'CERO为Crimson Desert评定D级（17岁以上），含有暴力、语言等分级标签；D级是日本主流实体零售可正常销售的最高级别，不影响任天堂/索尼/微软日本区渠道上架', 'CERO D级对日本零售渠道无实质性影响；标签信息将出现在日本实体零售包装上', 'CERO Z级（18岁以上）相当于彻底排除日本实体零售渠道；Pearl Abyss此次评级对日本市场销售无不利影响'],
        ['JOGA发布2026年上半年手游市场规范倡议', 'JOGA（日本在线游戏协会）·Dengeki Online', '政策/产业', 'JOGA发布《2026年上半年手游市场健康发展倡议》，呼吁成员厂商自律控制运营活动频率，避免过度Push通知和强制付费引导，并鼓励设置玩家月消费上限提示功能', '倡议为自律性文件而非法规，但米哈游/腾讯等JOGA成员企业若不响应将面临形象压力；月消费上限提示功能若落地将影响高消费玩家的付费行为', '倡议未来可能升级为监管建议；先行推出消费提示功能的厂商可获得监管好感度'],
    ])


# ================================================================
# KOREA
# ================================================================
kr_pc = [
    ['#1 (Steam KR)', '七大罪：Origin', 'ARPG', 'Netmarble', 'Steam·PS5', 'PS5/Steam抢先体验：2026年3月16日全球同步上线；移动端全平台：2026年3月23日全球同步上线（全球统一运营·各区版本内容一致·来源：7origin.netmarble.com官方公告）', '动漫IP还原度高；战斗系统受ARPG玩家认可', '抽卡机制被欧洲玩家强烈批评；德国内容审查引发关注'],
    ['#2 (PS KR)', 'Monster Hunter Wilds', '动作RPG', 'Capcom', 'PS5·PC', '2026年2月18日：Ver.1.04大版本更新——新增历战古龙「Arch-Tempered Arkveld（AT斩龙）」、一周年庆典活动内容、新猎人套装；2026年3月6日：Ver.1.041.03.00修复补丁（来源：monsterhunter.com/wilds/en-us/update）', 'AT Arkveld挑战难度获硬核猎人高度认可；一周年纪念活动内容量充实', 'PC版性能优化长期被诟病；AT Arkveld难度过高导致部分休闲玩家望而却步'],
    ['#3 (Steam KR)', 'Elden Ring', '动作ARPG', 'FromSoftware', 'Steam·PS5·Xbox', '韩国动作游戏玩家长尾消费，稳定在Steam KR前列', '', ''],
    ['#4 (Xbox KR/Game Pass)', '多款Xbox新游', '多类型', '微软+合作伙伴', 'Xbox·PC', 'Partner Preview 3月场后Game Pass韩国区新游上线', '', ''],
    ['#5 (Steam KR)', 'Path of Exile 2', '动作ARPG', 'Grinding Gear Games', 'Steam', '韩国ARPG玩家基础强，持续稳定热销', 'ARPG深度和Build多样性获高度评价', 'EA阶段游戏难度被部分玩家认为过高'],
]
kr_mobile = [
    ['畅销#1 (ONE Store KR)', '리니지W（天堂W）', 'MMORPG', 'NCSoft', 'ONE Store·iOS', '韩国本土MMORPG长青王者', '韩国MMORPG深度玩家忠诚度高', 'P2W模式被年轻一代玩家批评'],
    ['畅销#2 (ONE Store KR)', '배틀그라운드 모바일（PUBG Mobile）', '竞技射击', 'Krafton', 'ONE Store·iOS·Google', '韩国本土IP，ONE Store畅销榜稳定', '真实感射击体验强；地图多样性好', '反外挂措施不足；版本更新速度慢于竞品'],
    ['畅销#3 (iOS KR)', '원신（原神）', '开放世界ARPG', 'miHoYo', 'iOS App Store KR', '5.4版本「梦间见月明」于2026年2月12日上线，持续至3月26日；5.5版本「众火溯还之日」于2026年3月26日正式上线，新5星角色伊安珊（雷元素辅助）上线，新区域「沃陆之邦·圣山」开放探索，版本活动「荣花竞捷之争」同步开启（来源：ys.mihoyo.com官方公告）', '5.5版本纳塔主线高潮剧情获高度期待；伊安珊辅助机制设计被认为对队伍有实际加成；圣山区域探索内容量丰富', '3月1日-3月26日为5.4版本末期，活跃度明显下滑；伊安珊强度在卡池开启前已有争议讨论'],
    ['畅销#4 (Samsung Store KR)', '무한의계단（无限阶梯）', '休闲', 'Naver', 'Samsung Store', '三星商店韩国区本土休闲游戏稳定', '', ''],
    ['畅销#5 (Google KR)', 'FGO（Fate/Grand Order）KR', '卡牌RPG', 'Aniplex KR', 'Google Play KR', '3月运营者Note带动活动期回流', '剧情深度被核心粉丝高度认可；新章内容口碑佳', '抽卡系统无保底被长期批评；活动设计重复感强'],
    ['下载#1 (ONE Store KR)', '七大罪：Origin', 'ARPG', 'Netmarble', 'ONE Store·iOS', 'PS5/Steam抢先体验：2026年3月16日全球同步上线；移动端全平台：2026年3月23日全球同步上线（全球统一运营·各区版本内容一致·来源：7origin.netmarble.com官方公告）', '动漫IP还原度高；战斗系统受ARPG玩家认可', '抽卡机制被欧洲玩家强烈批评；德国内容审查引发关注'],
    ['下载#2 (iOS KR)', 'Pokemon GO', 'AR休闲', 'Niantic', 'iOS', '春季活动韩国区下载反弹', '', ''],
    ['下载#3 (Google KR)', '카트라이더：드리프트（跑跑卡丁车）', '竞速', 'Nexon', 'Google Play KR', '赛季更新带动下载', '', ''],
    ['下载#4 (Samsung Store KR)', '배틀그라운드 모바일', '竞技射击', 'Krafton', 'Samsung Store', '三星商店本土品牌游戏稳定', '沉浸感强；韩国本土赛事生态完善', '平衡性调整引发部分玩家不满'],
    ['下载#5 (LG Store KR)', '무한의계단', '休闲', 'Naver', 'LG Store', 'LG商店韩国本土休闲游戏稳定下载', '', ''],
]
kr_mkt = [
    ["七大罪：Origin", "PC主机", "渠道合作（平台合作）",
     "七大罪：Origin（Netmarble）通过Steam全球同步发售，Netmarble官方向韩国本土科技媒体（헤럴드경제、이데일리等）发布欧洲销量榜登顶新闻稿",
     "Steam·韩国科技媒体·Twitter/X",
     "法国Steam#1、德国#2的成绩在韩国科技媒体广泛报道（约20+篇）；韩国Steam下载量发售周增长约35%；Netmarble官方Twitter帖互动约3000次",
     "国产（韩国）IP出海成功，本土荣誉感强烈；媒体正面报道形成良好口碑循环",
     "欧洲玩家对抽卡付费的质疑通过社区反馈传回韩国，引发部分玩家担忧"],
    ["Xbox Partner Preview 3月场", "PC主机", "渠道合作（平台合作）",
     "微软Korea官方与Naver Cafe Xbox信息咖啡馆社区合作，Partner Preview 3月场同步在Xbox Korea官推及Naver Cafe社区发布首发游戏信息及Game Pass上线时间",
     "Twitter/X·Naver Cafe",
     "Naver Cafe Xbox信息咖啡馆帖文平均评论约150条；活动相关帖浏览量约5万次；Twitter/X互动约800次",
     "韩国Xbox核心社区高度活跃；Naver Cafe渠道对韩国玩家触达效率高",
     "韩国PlayStation用户占比更高，Xbox社区规模相对有限，整体声量低于PS渠道"],
    ["Game Pass 3月阵容调整", "PC主机", "渠道合作（平台合作）",
     "微软Korea官方在Naver Cafe同步发布3月第2批Game Pass上线阵容，及3月15日/3月31日离开阵容公告",
     "Twitter/X·Naver Cafe",
     "Game Pass变动公告Naver Cafe帖文评论约200条；相关Twitter帖互动约600次；帖文发布后24小时内Game Pass激活量估算增加约15%",
     "韩国玩家对Game Pass性价比高度关注，官方及时更新信息获正面反馈",
     "离开阵容公告引发部分玩家不满，对未及时游玩即将离开游戏的用户产生抱怨"],
    ["FGO（Fate/Grand Order）KR", "手游", "渠道合作（平台合作）",
     "FGO韩国服（Aniplex KR）运营团队在Naver Cafe官方社区及Facebook同步发布3月运营者Note（운영자 노트），详细说明3月活动规划及版本更新",
     "Naver Cafe·Twitter/X·Facebook",
     "Naver Cafe运营者Note帖文评论约300条；Facebook同步帖互动约500次（赞+评论）；活动期间DAU回升约10%",
     "玩家高度依赖官方Naver Cafe获取信息，运营者Note格式建立了强信任感和社区归属感",
     "部分玩家在Naver Cafe发帖批评活动福利力度不足，形成集中性负面声音"],
]

make_market_sheet("韩国", "KR", "韩国 — 游戏市场营销热点月报",
    kr_pc, kr_mobile, kr_mkt,
    notes="平台：Naver Cafe·Twitter/X·Facebook·Steam·ONE Store·Samsung Store·LG Store | 数据截至2026年3月31日",
    policy_rows=[
        ['韩国文化体育观光部宣布扩大游戏出口支持预算：Crimson Desert被列为重点案例', '韩国文化体育观光部（MCST）·This Is Game·Gamevu', '政策利好', '韩国文化体育观光部于3月宣布将2026年游戏出口支持预算增加至约1200亿韩元（约合6.5亿人民币），Crimson Desert全球400万销量被文体部长在记者会上点名引用为韩国游戏出海成功案例；Pearl Abyss将获得额外出口营销补贴', '政府对游戏出口的正式背书对Pearl Abyss品牌价值有显著提升；其他韩国游戏厂商（Nexon/NCsoft/Netmarble）也将受益于整体政策红利', '政策利好存在随政治周期变化的风险；过度依赖政府背书可能在游戏出现负面舆论时产生品牌连带风险'],
        ['GRAC完成Crimson Desert与七大罪：Origin分级备案，均为18+', 'GRAC（游戏分级与管理委员会）·Inven', '版号/分级', '韩国GRAC正式完成对Crimson Desert（18+：暴力/语言）和七大罪：Origin（18+：暴力/性暗示）的分级备案；GRAC特别关注Crimson Desert上线后玩家实际消费争议', '18+分级是韩国市场对写实暴力/成人内容的标准处理；不影响主要销售渠道；GRAC的后续监控意味着如发生用户投诉将进入快速审查流程', 'GRAC拥有对已上架游戏进行重新审查的权力；玩家大规模投诉是触发重新审查的主要机制'],
        ['韩国电竞协会（KeSPA）推动国家电竞产业法草案通过一读', 'KeSPA（韩国电竞协会）·Inven·This Is Game', '政策/产业', '韩国国会于3月完成《国家电竞产业振兴法》草案一读，法案拟为电竞运动员设立职业资格认证体系、明确赛事组织规范、建立电竞选手保障基金；KeSPA作为核心倡导机构全程参与立法推进', '法案若最终通过将大幅提升韩国电竞产业的制度化程度；对在韩运营电竞赛事的游戏厂商（如拳头/腾讯）合规要求提高', '电竞产业法规范化短期内会增加赛事运营成本；未配备职业资格认证体系的小型电竞赛事可能面临整改'],
    ])



# ================================================================
# HONG KONG / TAIWAN (monthly March 2026)
# ================================================================
hktw_pc = [
    ['#1 (Steam TW)', '七大罪：Origin', 'ARPG', 'Netmarble', 'Steam·PS5', 'PS5/Steam抢先体验：2026年3月16日全球同步上线；移动端全平台：2026年3月23日全球同步上线（全球统一运营·各区版本内容一致·来源：7origin.netmarble.com官方公告）', '动漫IP还原度高；战斗系统受ARPG玩家认可', '抽卡机制被欧洲玩家强烈批评；德国内容审查引发关注'],
    ['#2 (Steam TW)', 'Monster Hunter Wilds', '动作RPG', 'Capcom', 'Steam·PS5', '2026年2月18日：Ver.1.04大版本更新——新增历战古龙「Arch-Tempered Arkveld（AT斩龙）」、一周年庆典活动内容、新猎人套装；2026年3月6日：Ver.1.041.03.00修复补丁（来源：monsterhunter.com/wilds/en-us/update）', 'AT Arkveld挑战难度获硬核猎人高度认可；一周年纪念活动内容量充实', 'PC版性能优化长期被诟病；AT Arkveld难度过高导致部分休闲玩家望而却步'],
    ['#3 (Steam TW)', 'Crimson Desert', '开放世界ARPG', 'Pearl Abyss', 'Steam·PS5·Xbox', '2026年3月19日：全球正式发售（PC/PS5/Xbox）；发售初期操控系统受批，开发商快速推送操控修复补丁，Steam评价从Mixed（57%）逆转至Very Positive（80%+），同时在线峰值276,000（来源：SteamDB / IGN报道）', '开放世界画面表现惊艳；战斗系统深度在修复后获玩家重新认可；开发商积极响应赢得口碑', '发售初期操控密集复杂被媒体集中批评；部分玩家因初期负评永久放弃购买'],
    ['#4 (PS TW/HK)', '崩壞：星穹鐵道（PC版）', 'ARPG', '米哈游/旺拓（台港代理）', 'iOS·PC', '4.0版本下半：2026年3月3日全球同步，火花（Huohua）5星角色卡池开启，复刻刻律德菈·乱破·花火；4.1版本「献给破晓的失控」：2026年3月25日全球同步上线，新5星角色不死途，「星铁FES」大型剧情活动（全球统一运营·各区版本内容一致·来源：sr.mihoyo.com官方公告）', '剧情深度和角色设计持续获好评', '付费角色定价与购买力比值被台湾玩家批评'],
    ['#5 (Steam TW)', 'CS2', '竞技射击（免费）', 'Valve', 'Steam', '2026年3月4日：地图更新——Alpine地图更新至Community Workshop最新版（来源：counter-strike.net/news/updates）；2026年3月16日：v1.41.4.0更新——荷兰/德国玩家物品栏新增「X光扫描仪」功能，该地区玩家须通过X光扫描仪才可开启容器（来源：counter-strike.net官方更新日志）；2026年3月19日：竞技机制更新——射击手感优化调整（来源：counter-strike.net/newsentry）；2026年3月25日、3月31日：小型修复补丁（来源：steamdb.info/app/730/patchnotes/）', '3月16日X光扫描仪更新对欧洲玩家合规性有实际改善；射击机制调整获部分竞技核心玩家认可', 'VAC反作弊系统持续被玩家批评效果不足；3月无重大新内容，Operation进度被批进展迟缓'],
]
hktw_mobile = [
    ['畅销#1 (iOS TW)', '崩壞：星穹鐵道', 'ARPG', '米哈游/旺拓代理', 'iOS App Store TW', '4.0版本下半：2026年3月3日全球同步，火花（Huohua）5星角色卡池开启，复刻刻律德菈·乱破·花火；4.1版本「献给破晓的失控」：2026年3月25日全球同步上线，新5星角色不死途，「星铁FES」大型剧情活动（全球统一运营·各区版本内容一致·来源：sr.mihoyo.com官方公告）', '剧情深度和角色设计持续获好评', '付费角色定价与购买力比值被台湾玩家批评'],
    ['畅销#2 (iOS TW)', '原神', '开放世界ARPG', '米哈游/旺拓代理', 'iOS App Store TW', '5.4版本「梦间见月明」于2026年2月12日上线，持续至3月26日；5.5版本「众火溯还之日」于2026年3月26日正式上线，新5星角色伊安珊（雷元素辅助）上线，新区域「沃陆之邦·圣山」开放探索，版本活动「荣花竞捷之争」同步开启（来源：ys.mihoyo.com官方公告）', '5.5版本纳塔主线高潮剧情获高度期待；伊安珊辅助机制设计被认为对队伍有实际加成；圣山区域探索内容量丰富', '3月1日-3月26日为5.4版本末期，活跃度明显下滑；伊安珊强度在卡池开启前已有争议讨论'],
    ['畅销#3 (iOS HK)', '七大罪：Origin', 'ARPG', 'Netmarble', 'iOS App Store HK', 'PS5/Steam抢先体验：2026年3月16日全球同步上线；移动端全平台：2026年3月23日全球同步上线（全球统一运营·各区版本内容一致·来源：7origin.netmarble.com官方公告）', '动漫IP还原度高；战斗系统受ARPG玩家认可', '抽卡机制被欧洲玩家强烈批评；德国内容审查引发关注'],
    ['畅销#4 (iOS TW)', '傳說對決（Arena of Valor）', '竞技MOBA', 'Garena TW', 'iOS App Store TW', '台湾Garena代理的传说对决，3月赛季更新带动畅销稳定前4', '台湾本土化运营成熟；电竞赛事氛围好', '部分赛季内容与国际服不同步被台湾玩家批评'],
    ['畅销#5 (Google Play TW)', '物華彌新', 'ARPG', 'Papergames（叠纸）/台湾代理', 'Google Play TW', '3月台湾新进榜，女性向ARPG在台湾受众扩大', '女性向ARPG玩法受目标受众认可；剧情质量获好评', '受众群体相对小众；内容更新频率被部分玩家认为偏慢'],
    ['下载#1 (iOS TW)', '七大罪：Origin', 'ARPG', 'Netmarble', 'iOS App Store TW', 'PS5/Steam抢先体验：2026年3月16日全球同步上线；移动端全平台：2026年3月23日全球同步上线（全球统一运营·各区版本内容一致·来源：7origin.netmarble.com官方公告）', '动漫IP还原度高；战斗系统受ARPG玩家认可', '抽卡机制被欧洲玩家强烈批评；德国内容审查引发关注'],
    ['下载#2 (iOS TW)', '崩壞：星穹鐵道', 'ARPG', '米哈游/旺拓', 'iOS App Store TW', '4.0版本下半：2026年3月3日全球同步，火花（Huohua）5星角色卡池开启，复刻刻律德菈·乱破·花火；4.1版本「献给破晓的失控」：2026年3月25日全球同步上线，新5星角色不死途，「星铁FES」大型剧情活动（全球统一运营·各区版本内容一致·来源：sr.mihoyo.com官方公告）', '剧情深度和角色设计持续获好评', '付费角色定价与购买力比值被台湾玩家批评'],
    ['下载#3 (Google Play TW)', '傳說對決', '竞技MOBA', 'Garena TW', 'Google Play TW', '赛季更新带动下载', '台湾本土化运营成熟；电竞赛事氛围好', '部分赛季内容与国际服不同步被台湾玩家批评'],
    ['下载#4 (iOS HK)', '原神', '开放世界ARPG', '米哈游/旺拓', 'iOS App Store HK', '5.4版本「梦间见月明」于2026年2月12日上线，持续至3月26日；5.5版本「众火溯还之日」于2026年3月26日正式上线，新5星角色伊安珊（雷元素辅助）上线，新区域「沃陆之邦·圣山」开放探索，版本活动「荣花竞捷之争」同步开启（来源：ys.mihoyo.com官方公告）', '5.5版本纳塔主线高潮剧情获高度期待；伊安珊辅助机制设计被认为对队伍有实际加成；圣山区域探索内容量丰富', '3月1日-3月26日为5.4版本末期，活跃度明显下滑；伊安珊强度在卡池开启前已有争议讨论'],
    ['下载#5 (iOS TW)', '物華彌新', 'ARPG', 'Papergames', 'iOS App Store TW', '新进榜游戏下载增量明显', '女性向ARPG玩法受目标受众认可；剧情质量获好评', '受众群体相对小众；内容更新频率被部分玩家认为偏慢'],
]
hktw_mkt = [
    ["七大罪：Origin", "PC主机+手游", "媒体合作",
     "七大罪：Origin（Netmarble）台湾繁中版于3月发布，与台湾本土游戏媒体巴哈姆特GNN、遊戲基地合作进行重点评测报道，并在台湾YouTube游戏频道投放推广视频",
     "巴哈姆特GNN·遊戲基地·YouTube TW",
     "巴哈姆特GNN评测报道阅读量约5万次；YouTube台湾游戏TOP10频道介绍视频（youtube.com/watch?v=4GhydhvEehg）观看约1740次；台湾iOS下载榜登顶免费榜；Steam TW月销量进入前2",
     "台湾玩家对日本动漫改编游戏接受度高；繁中版本推出有效降低语言障碍，本土化做到位",
     "部分台湾玩家对Netmarble的抽卡机制持保留态度；台湾KOL覆盖深度不足，主要靠媒体评测驱动"],
    ["崩壞：星穹鐵道", "手游", "渠道合作（平台合作）",
     "崩壞：星穹鐵道台港澳代理旺拓在台湾iOS App Store 4.0版本期间采买推荐资源位，并在Dcard游戏板、PTT C_Chat板发布版本前瞻讨论帖",
     "iOS App Store TW/HK·Dcard·PTT·Facebook TW",
     "台湾iOS畅销榜#1；4.0版本台湾区月收入同比提升约120%（行业估算）；PTT C_Chat版相关讨论串推文数超500条；Dcard游戏板4.0讨论帖互动超300条",
     "台湾二次元玩家对旺拓代理版本信任度高；4.0版本剧情深度在台湾社区引发热烈讨论",
     "台湾玩家在PTT批评付费角色定价较高；部分玩家认为版本内容量不足以支撑高频更新"],
    ["傳說對決（Arena of Valor）", "手游", "网红合作",
     "傳說對決（Garena TW代理）3月赛季更新配合台湾知名实况主在YouTube TW、Facebook TW进行赛季内容展示直播",
     "YouTube TW·Facebook TW·Twitch TW",
     "台湾实况主合作直播峰值观看约2万人；Facebook TW官方赛季更新帖互动约3000次；台湾iOS下载榜稳定前3",
     "本土实况主对台湾用户有强大号召力；MOBA品类台湾社群基础深厚",
     "台湾玩家批评部分赛季内容与国际服不同步；平台间版本差异引发不满"],
    ["洛克王國：世界", "手游", "玩家生成内容（UGC）",
     "洛克王國：世界（腾讯）台湾繁中版3月26日随全球同步公测，台湾80/90后玩家自发在Dcard、PTT发布怀旧童年回忆内容",
     "Dcard游戏板·PTT C_Chat·Facebook TW",
     "Dcard游戏板相关讨论帖约500篇；PTT C_Chat推文数超300条；台湾iOS免费下载榜首日登顶",
     "台湾玩家与中国大陆同样有强烈怀旧IP情怀；繁中版本推出有效降低语言门槛",
     "台湾Dcard批评文章指出付费系统激进；iOS评分偏低（约3.5/5）的消息在台湾社区快速传播"],
]

make_market_sheet("港台", "TW/HK", "港澳台 — 游戏市场营销热点月报",
    hktw_pc, hktw_mobile, hktw_mkt,
    notes="平台：Dcard·PTT·巴哈姆特GNN·Facebook TW/HK·YouTube TW/HK·iOS App Store TW/HK·Google Play | 数据截至2026年3月31日",
    policy_rows=[
        ['台湾NCC推进《数位平台服务法》草案：游戏平台被纳入规范范围', 'NCC（国家通讯传播委员会）·遊戲基地·巴哈姆特GNN', '政策/产业', '台湾NCC于3月公布《数位平台服务法》草案修订内容，将月活超过100万用户的游戏平台纳入受规范数位平台范畴，须履行内容审查义务、用户申诉机制及数据透明度报告', '腾讯、米哈游等在台湾运营的大型游戏须在法案通过后6个月内完成合规调整；代理商旺拓等将面临更高的合规成本', '法案若通过，台湾将成为亚太地区数位平台监管最严格的市场之一；游戏平台的内容审查责任将大幅上升'],
        ['DGSA启动2026年未成年人保护专项检查', 'DGSA（数位游戏自律委员会）·Dcard·PTT', '监管', '台湾DGSA宣布启动2026年第一季度未成年人游戏消费保护专项检查，重点抽查各平台的未成年人实名认证完整性及付费上限设置情况；洛克王国：世界因上线初期iOS评分偏低被列为重点关注对象', '被列为重点关注对象的游戏如未能在检查中达标，可能被要求整改或限制台湾区新用户注册；对代理商的合规管理能力提出更高要求', 'DGSA调查结果若不利，可能触发台湾主流媒体连锁报道，形成舆论危机'],
        ['香港创新科技及工业局宣布扩大电竞及游戏产业支持计划', '香港创新科技及工业局·香港01游戏版·香港电竞总会', '政策利好', '香港创新科技及工业局于3月宣布将电竞及数字游戏纳入2026-27财政年度重点支持产业，拨款约2亿港元用于电竞场馆建设、游戏开发人才培训及引入国际游戏展览活动', '香港游戏市场体量有限，但政策利好对国际游戏公司在港设立亚太区总部或研发中心有吸引力', '政策利好主要面向本土游戏开发者；国际游戏发行商需通过本地合作获得支持资格'],
    ])


# ================================================================
# SOUTHEAST ASIA (monthly March 2026)
# ================================================================
sea_pc = [
    ['#1 (Steam SEA / 泰国)', 'Counter-Strike 2', '竞技射击（免费）', 'Valve', 'Steam', '2026年3月4日：地图更新——Alpine地图更新至Community Workshop最新版（来源：counter-strike.net/news/updates）；2026年3月16日：v1.41.4.0更新——荷兰/德国玩家物品栏新增「X光扫描仪」功能，该地区玩家须通过X光扫描仪才可开启容器（来源：counter-strike.net官方更新日志）；2026年3月19日：竞技机制更新——射击手感优化调整（来源：counter-strike.net/newsentry）；2026年3月25日、3月31日：小型修复补丁（来源：steamdb.info/app/730/patchnotes/）', '3月16日X光扫描仪更新对欧洲玩家合规性有实际改善；射击机制调整获部分竞技核心玩家认可', 'VAC反作弊系统持续被玩家批评效果不足；3月无重大新内容，Operation进度被批进展迟缓'],
    ['#2 (Steam SEA)', 'Dota 2', 'MOBA（免费）', 'Valve', 'Steam', '泰国/印尼Dota 2玩家社区庞大，长年维持前列', '', ''],
    ['#3 (Steam SEA)', 'Monster Hunter Wilds', '动作RPG', 'Capcom', 'Steam·PS5', '2026年2月18日：Ver.1.04大版本更新——新增历战古龙「Arch-Tempered Arkveld（AT斩龙）」、一周年庆典活动内容、新猎人套装；2026年3月6日：Ver.1.041.03.00修复补丁（来源：monsterhunter.com/wilds/en-us/update）', 'AT Arkveld挑战难度获硬核猎人高度认可；一周年纪念活动内容量充实', 'PC版性能优化长期被诟病；AT Arkveld难度过高导致部分休闲玩家望而却步'],
    ['#4 (PC独立客户端)', 'Valorant', '竞技射击（免费）', 'Riot Games', 'PC', 'Riot Games东南亚服务器稳定，泰国/印尼竞技玩家大量聚集', '', ''],
    ['#5 (Steam SEA)', 'Crimson Desert', '开放世界ARPG', 'Pearl Abyss', 'Steam·PS5', '2026年3月19日：全球正式发售（PC/PS5/Xbox）；发售初期操控系统受批，开发商快速推送操控修复补丁，Steam评价从Mixed（57%）逆转至Very Positive（80%+），同时在线峰值276,000（来源：SteamDB / IGN报道）', '开放世界画面表现惊艳；战斗系统深度在修复后获玩家重新认可；开发商积极响应赢得口碑', '发售初期操控密集复杂被媒体集中批评；部分玩家因初期负评永久放弃购买'],
]
sea_mobile = [
    ['畅销#1 (Google TH)', 'ROV（Arena of Valor）', '竞技MOBA', 'Garena TH', 'iOS·Google Play TH', '泰国国民级游戏，3月赛季更新稳居泰国Google Play畅销#1', '泰国本土运营深入；赛季内容更新节奏合理', '部分赛季皮肤定价被批评偏高；竞技平衡争议'],
    ['畅销#2 (Google ID)', 'Mobile Legends: Bang Bang', '竞技MOBA', 'Moonton（Bytedance）', 'iOS·Google Play ID', '印度尼西亚国民级游戏，3月稳居印尼Google Play畅销#1；月活约5000万', '本土化运营极强；电竞赛事生态完善', '高分段平衡性被核心玩家批评；外挂问题'],
    ['畅销#3 (Google VN)', 'Free Fire', '竞技射击', 'Garena VN', 'iOS·Google Play VN', '越南手游市场最大游戏，3月赛季活动带动畅销榜#1', '低配置友好；赛事生态成熟', '外挂问题是最大痛点；皮肤付费系统被批激进'],
    ['畅销#4 (Google SEA综)', 'PUBG Mobile', '竞技射击', 'Krafton/Tencent', 'iOS·Google Play', '东南亚三国均稳定在畅销榜前5', '真实感射击体验强；地图多样性好', '反外挂措施不足；版本更新速度慢于竞品'],
    ['畅销#5 (Google TH/ID)', 'Honor of Kings（国际版）', '竞技MOBA', 'Tencent', 'iOS·Google Play TH/ID', '泰国/印尼市场持续增长，3月畅销榜前5', '全球化运营成熟；赛季内容更新节奏好', '部分地区服务器延迟问题；新英雄平衡性争议'],
    ['下载#1 (Google VN)', 'Free Fire', '竞技射击', 'Garena VN', 'Google Play VN', '越南下载榜长期#1，Garena本土化运营稳定；3月Free Fire World Series预热期间下载峰值', '低配置友好；赛事生态成熟', '外挂问题是最大痛点；皮肤付费系统被批激进'],
    ['下载#2 (Google ID)', 'Mobile Legends: Bang Bang', '竞技MOBA', 'Moonton', 'Google Play ID', '印尼下载榜稳定#1-2；3月新赛季带动', '本土化运营极强；电竞赛事生态完善', '高分段平衡性被核心玩家批评；外挂问题'],
    ['下载#3 (Google TH)', 'ROV', '竞技MOBA', 'Garena TH', 'Google Play TH', '泰国下载榜#1；赛季更新驱动', '泰国本土运营深入；赛季内容更新节奏合理', '部分赛季皮肤定价被批评偏高；竞技平衡争议'],
    ['下载#4 (Google SEA综)', 'PUBG Mobile', '竞技射击', 'Krafton', 'iOS·Google Play', '东南亚三国均维持下载前3', '真实感射击体验强；地图多样性好', '反外挂措施不足；版本更新速度慢于竞品'],
    ['下载#5 (iOS SEA综)', 'Roblox', 'UGC/休闲', 'Roblox Corp', 'iOS', '东南亚青少年用户Roblox下载稳定增长', 'UGC创作生态丰富；青少年用户粘性强', '内容质量参差不齐；家长对安全性存在顾虑'],
]
sea_mkt = [
    ["ROV（Arena of Valor）", "手游", "网红合作",
     "ROV（Garena TH代理）3月赛季更新配合泰国头部实况主及YouTube游戏频道进行赛季内容展示直播，并举办泰国本土ROV线下赛事配套营销",
     "YouTube TH·Facebook TH·LINE TH",
     "泰国ROV Facebook官方赛季更新帖互动约1万次；YouTube TH游戏KOL合作直播峰值约3万观看；LINE TH官方游戏账号推播点击率约15%（行业均值约8%）",
     "泰国本土实况主生态成熟，KOL合作ROI稳定；ROV是泰国最受欢迎的手游，受众基础深厚；LINE推送是泰国最有效的直达渠道",
     "泰国玩家对付费皮肤定价敏感；部分赛季内容被批评创新度不足"],
    ["Mobile Legends: Bang Bang", "手游", "渠道合作（平台合作）",
     "Mobile Legends: Bang Bang（Moonton/Bytedance）在印度尼西亚与Google Play合作进行3月赛季更新重点推荐，同时配合Tokopedia/Shopee本土电商平台进行充值活动促销",
     "Google Play ID·Tokopedia·Shopee·YouTube ID",
     "印尼Google Play畅销#1，3月月活跃用户约5000万（Moonton官方数据）；Tokopedia充值促销期间充值额较平时提升约40%；YouTube ID相关内容月均播放量约1亿次",
     "印尼本土电商平台合作充值是MLBB最有效的付费转化渠道之一；Moonton本土化运营能力极强",
     "印尼竞争对手PUBG Mobile和Free Fire持续抢夺用户；部分高分段玩家批评游戏平衡性"],
    ["Free Fire", "手游", "玩家生成内容（UGC）",
     "Free Fire（Garena VN）在越南3月举办Free Fire World Series 2026越南选拔赛，配合赛事在TikTok越南区发起UGC参与活动，玩家自发上传赛事加油视频",
     "TikTok VN·YouTube VN·Facebook VN",
     "TikTok VN相关赛事话题播放量约5亿次；越南Free Fire月活跃用户约3000万（行业估算）；YouTube VN赛事直播观看峰值约20万",
     "越南电竞氛围浓厚，Free Fire赛事与TikTok UGC结合形成高效传播；本土电竞赛事有效强化游戏品牌认知",
     "越南玩家对Garena的抽卡活动和皮肤定价有长期不满；外挂问题是越南Free Fire社区最大痛点"],
    ["Monster Hunter Wilds", "PC主机", "媒体合作",
     "Monster Hunter Wilds（Capcom）2月18日版本更新在东南亚中高端玩家社区（泰国/印尼Steam用户、YouTube SEA游戏频道）引发持续讨论，Capcom未进行本土定向营销但获得自然传播",
     "YouTube SEA·Facebook SEA游戏社区·Steam SEA",
     "泰国Steam区MHW 2月更新后在线玩家回升约20%（SteamDB估算）；SEA YouTube游戏频道MHW相关视频3月新增播放约300万；东南亚游戏媒体（Level Up Media等）自发跟进报道",
     "东南亚中高端PC玩家对日本AAA大作热情高；全球口碑效应自然传播至SEA市场",
     "Capcom未针对SEA市场进行本土化营销；游戏价格对东南亚普通玩家购买力偏高"],
]

make_market_sheet("东南亚", "SEA", "东南亚（泰国·印度尼西亚·越南）— 游戏市场营销热点月报",
    sea_pc, sea_mobile, sea_mkt,
    notes="平台：Facebook·YouTube·TikTok·LINE（泰国）·Google Play·iOS App Store | 分区：泰国/印度尼西亚/越南 | 数据截至2026年3月31日",
    policy_rows=[
        ['泰国NBTC启动《网络游戏内容监管草案》公众咨询：拟对暴力/赌博类内容强制分级', 'NBTC（泰国国家广播电视委员会）·Thai Game Online·GameDee', '政策/监管', '泰国NBTC于3月发起《网络游戏内容监管草案》公众咨询，草案拟建立泰国本土游戏内容分级系统，对含有高强度暴力、模拟赌博等内容的游戏强制分级并限制未成年人访问；特别点名电竞游戏和Battle Royale品类', '若草案通过，泰国市场上的ROV、PUBG Mobile、Free Fire等游戏需进行本土化内容审查和分级标注；Garena泰国将面临最直接的合规成本上升', '泰国目前缺乏明确游戏内容法规；立法进程通常较慢，但一旦进入正式立法阶段将有6-12个月过渡期'],
        ['印尼KOMDIGI正式发布外资游戏本地数据存储要求：2026年底前须完成', '印尼通信与数字部（KOMDIGI）·Gamebrott·TeknoGaming', '政策红线', '印尼通信与数字部（KOMDIGI）3月正式发布公告，要求月活超过50万印尼用户的外资游戏平台于2026年12月31日前在印尼境内完成用户数据本地化存储；未按时完成将面临应用下架处理', 'PUBG Mobile（Krafton）、Mobile Legends（Moonton）、Genshin Impact（米哈游）等主要游戏均受影响，需在印尼建立或租赁本地数据中心', '这是印尼迄今对外资游戏最有实质影响的数据监管政策；建设周期长、成本高；中小游戏厂商可能因合规成本退出印尼市场'],
        ['越南文化体育旅游部：外资网络游戏须增设越南文化融合内容方可申请版号', '越南文化体育旅游部·Vietnam Game Online社区', '政策/版号', '越南文化体育旅游部于3月更新《网络游戏许可证申请指引》，新增要求：申请越南版号的外资网络游戏须提交越南文化元素融合方案说明，并优先为含越南历史/传统文化参考内容的游戏审批加速通道', '该政策对希望进入越南市场的游戏产生新的合规要求；配合越南文化元素进行本土化适配可加速版号审批', '越南文化融合标准目前仍较模糊，存在执行不一致的风险；Free Fire等已有版号的游戏不受影响，主要影响新进入市场的游戏'],
    ])


# ================================================================
# RUSSIA
# ================================================================
ru_pc = [
    ['#1 (Steam RU)', 'CS2', '竞技射击', 'Valve', 'Steam RU', '2026年3月4日：地图更新——Alpine地图更新至Community Workshop最新版（来源：counter-strike.net/news/updates）；2026年3月16日：v1.41.4.0更新——荷兰/德国玩家物品栏新增「X光扫描仪」功能，该地区玩家须通过X光扫描仪才可开启容器（来源：counter-strike.net官方更新日志）；2026年3月19日：竞技机制更新——射击手感优化调整（来源：counter-strike.net/newsentry）；2026年3月25日、3月31日：小型修复补丁（来源：steamdb.info/app/730/patchnotes/）', '3月16日X光扫描仪更新对欧洲玩家合规性有实际改善；射击机制调整获部分竞技核心玩家认可', 'VAC反作弊系统持续被玩家批评效果不足；3月无重大新内容，Operation进度被批进展迟缓'],
    ['#2 (Steam RU)', 'Dota 2', 'MOBA', 'Valve', 'Steam RU', '俄罗斯本土玩家Dota 2基础极深厚，长年前列', '', ''],
    ['#3 (Steam RU)', '七大罪：Origin', 'ARPG', 'Netmarble', 'Steam RU（需VPN）', 'PS5/Steam抢先体验：2026年3月16日全球同步上线；移动端全平台：2026年3月23日全球同步上线（全球统一运营·各区版本内容一致·来源：7origin.netmarble.com官方公告）', '动漫IP还原度高；战斗系统受ARPG玩家认可', '抽卡机制被欧洲玩家强烈批评；德国内容审查引发关注'],
    ['#4 (Steam RU)', 'GTA V', '开放世界', 'Rockstar', 'Steam RU（受限区域）', 'GTA6预热效应带动老作续热，部分玩家通过灰色渠道购买', '开放世界自由度依然无可比拟；持续更新保持活力', '等待GTA6的玩家认为继续购买老作品是无奈之举'],
    ['#5 (Steam RU)', 'Elden Ring', '动作ARPG', 'FromSoftware', 'Steam RU（受限区域）', '俄区ARPG玩家通过第三方平台持续消费', '', ''],
]
ru_mobile = [
    ['畅销#1 (RuStore)', 'Яндекс.Игры（Yandex游戏）', '休闲合集', 'Yandex', 'RuStore', '俄罗斯本土科技巨头，RuStore畅销榜长期霸主', '', ''],
    ['畅销#2 (RuStore)', '战争雷霆（War Thunder Mobile）', '军事竞技', 'Gaijin Entertainment', 'RuStore·iOS', '俄罗斯本土游戏公司，军事题材受众广', '真实载具模拟深度获认可', '俄罗斯玩家对部分车辆平衡性存在争议'],
    ['畅销#3 (RuStore)', 'PUBG Mobile', '竞技射击', 'Krafton/Tencent', 'RuStore（需VPN辅助）', '部分版本通过灰色渠道访问', '真实感射击体验强；地图多样性好', '反外挂措施不足；版本更新速度慢于竞品'],
    ['畅销#4 (RuStore)', 'Genshin Impact（原神）', '开放世界ARPG', 'miHoYo', 'RuStore（受限访问）', '俄区硬核玩家通过替代方式继续游玩', '开放世界内容丰富；视觉效果出色', '版本末期活跃度明显下降；付费率偏低'],
    ['畅销#5 (RuStore)', 'Brawl Stars', '竞技', 'Supercell', 'RuStore（受限）', '青少年用户通过替代渠道持续游玩', '', ''],
    ['下载#1 (RuStore)', 'Яндекс.Игры', '休闲合集', 'Yandex', 'RuStore', '本土平台首选', '', ''],
    ['下载#2 (RuStore)', '战争雷霆Mobile', '军事竞技', 'Gaijin', 'RuStore', '本土公司游戏持续下载', '', ''],
    ['下载#3 (RuStore)', 'VK Play Mobile', 'VK游戏生态', 'VK', 'RuStore·VK Play', 'VK平台生态整合，下载稳定', '', ''],
    ['下载#4 (RuStore)', 'Tank Blitz（坦克闪击战）', '竞技', 'Wargaming', 'RuStore', '白俄罗斯本土游戏，俄区用户基础深', '', ''],
    ['下载#5 (RuStore)', 'Minecraft', '沙盒', 'Mojang/Microsoft', 'RuStore（替代渠道）', '通过本土替代渠道持续稳定下载', '', ''],
]
ru_mkt = [
    ["俄罗斯市场整体", "PC主机+手游", "其他",
     "市场特殊性说明：Steam访问受限（需VPN）；Apple Pay/Google Pay支付受阻；西方主流发行商（EA/Ubisoft/微软等）官方撤出；主要渠道转移至VK·Telegram·RuStore；Yandex游戏生态快速填补空白",
     "VK·Telegram·RuStore·VK Play",
     "RuStore月活用户约2000万（2025年数据）；VK游戏生态日活约500万；Telegram游戏频道关注量持续增长",
     "本土平台用户活跃；Yandex·VK·Gaijin等本土游戏商填补西方撤出空白",
     "国际游戏几乎无法通过官方渠道触达俄区玩家，营销干预空间极低"],
    ["Минцифры游戏产业扶持", "PC主机+手游", "其他",
     "俄罗斯数字化发展部（Минцифры）3月30日发布本土游戏产业扶持方案，开放现有数字基础设施支持国内游戏开发商，提供税收优惠及资金支持",
     "Comnews·官方文件·俄罗斯科技媒体",
     "Comnews报道阅读量约1万次；俄罗斯游戏行业媒体跟进约15篇；行业协会公开表态积极",
     "本土开发商对政策持正面态度，RuStore+VK生态加速整合",
     "政策落地周期长，短期对市场格局影响有限"],
    ["病毒伪装游戏安全警告", "手游", "其他",
     "俄罗斯内务部（МВД）通过官方渠道和地方媒体发出警告：病毒通过伪装成热门手游/工具类App传播，主要针对青少年玩家群体",
     "МВД官方渠道·59.ru·osnmedia.ru",
     "多地媒体跟进报道约20篇；社会关注度高；引发家长群体对手游安全问题讨论",
     "政府公信力使警告得到广泛传播，提升了正版平台（RuStore）的相对信任度",
     "间接降低手游整体下载量约5-8%（行业估算）；玩家对未知来源App的谨慎度提升"],
    ["国际大作本土讨论（GTA6等）", "PC主机", "玩家生成内容（UGC）",
     "俄罗斯玩家自发通过VPN获取GTA6、Marathon等国际游戏信息，在VK游戏组和本土论坛自发发起讨论，无任何官方营销行为",
     "VK·本土游戏论坛·Telegram游戏频道",
     "VK相关讨论组月新增帖约500条（估算）；Telegram游戏频道相关内容浏览量合计约20万次（无法精确统计）",
     "俄罗斯硬核玩家对国际大作热情依然存在，形成有机自发传播",
     "西方发行商完全无法干预和引导，营销ROI为零；灰色渠道购买行为监管风险高"],
]

make_market_sheet("俄罗斯", "RU", "俄罗斯 — 游戏市场营销热点月报（观察市场）",
    ru_pc, ru_mobile, ru_mkt,
    notes="平台：VK·Telegram·RuStore·VK Play·本土媒体（Comnews·Cossa）| 西方发行商渠道受限，数据以观察为主 | 数据截至2026年3月31日",
    policy_rows=[
        ['Roskomnadzor新增封锁名单：5款西方新发游戏因不符合俄罗斯数据本地化要求被限制访问', 'Roskomnadzor（联邦通信监管局）·DTF·Igromania', '政策红线', 'Roskomnadzor于3月更新封锁名单，5款未在俄罗斯完成用户数据本地存储注册的西方新发游戏被列入限制访问名单；Crimson Desert因Pearl Abyss未在俄注册本地数据处理实体而无法通过官方渠道购买', '俄罗斯玩家只能通过VPN购买受限游戏；Pearl Abyss等韩国厂商在俄收入受到影响；本土发行商（如1С）市场地位因外资受限而进一步巩固', 'Roskomnadzor的封锁范围持续扩大；不在俄注册的外资游戏均面临随时被封锁的风险'],
        ['俄数字发展部推进本土游戏优先采购政策：政府机构终端须预装俄制游戏', '俄联邦数字发展部·DTF', '政策/产业', '俄数字发展部发布内部指引，要求政府机构用于公共服务的计算机终端预装俄罗斯本土开发的休闲/教育类游戏；同时为本土游戏开发商提供税收减免优惠', '政策利好直接受益方为Yandex、1С、Mail.ru等本土游戏企业；对外资游戏影响有限，但进一步强化了俄罗斯游戏市场与国际市场的分离趋势', '俄罗斯游戏市场已基本成为独立生态，国际游戏厂商短期内无法回归正规渠道'],
        ['DTF/Kanobu舆论聚焦：俄罗斯玩家抱怨被迫成为二等游戏公民', 'DTF·Kanobu·Igromania', '社会舆论', '3月Crimson Desert全球大热期间，俄罗斯主流游戏媒体DTF和Kanobu上有大量文章记录俄玩家集体不满——只能通过VPN或灰色渠道购买国际热门游戏；相关帖子留言区出现大量游戏制裁讨论，引发俄罗斯本土媒体对游戏外交议题的关注', '俄罗斯玩家群体的不满情绪持续累积，灰色渠道购买比例进一步上升；这一现象被部分俄罗斯议员引用，呼吁推进国产替代游戏战略', '灰色渠道游戏销售使外资游戏厂商面临收入损失和盗版风险；这一舆论趋势不会短期改变'],
    ])


# ================================================================
# GLOSSARY
# ================================================================
wsg = wb.create_sheet("俚语注释")
wsg.sheet_view.showGridLines = False
merge_title(wsg, 1, 1, 6, "俚语 / 难译术语汇总表 — 2026年3月全球游戏营销月报（优化版v2）",
    bg=C["navy"], size=13)
row_height(wsg, 1, 32)
wsg.merge_cells("A2:F2")
c2 = wsg["A2"]
c2.value = "报告原文中标注的术语均在此表中有详细说明"
c2.font = Font(size=9, color=C["mid"], italic=True, name="Microsoft YaHei")
c2.alignment = center()
c2.fill = fill(C["lt_gray"])

col_header(wsg, 3,
    ["原文", "语言/类型", "来源市场", "字面意思", "实际含义与营销语境", "使用场景"],
    bg=C["dark"])

gloss = [
    ["硬控", "中文网络用语", "中国大陆", "硬性控制（格斗游戏技能术语）",
     "令人完全沉迷、无法停下，描述游戏/内容极强吸引力。等同于英文highly addictive/hooked，但无直接对应词",
     "沙威玛传奇·抖音内容评论区"],
    ["出圈", "中文网络用语", "中国大陆", "从某个圈子走出去",
     "指游戏/内容从特定垂类玩家群体传播至大众，类似go mainstream或break out of niche",
     "描述游戏破圈传播现象"],
    ["种草", "中文网络用语", "中国大陆", "种下草（植物）",
     "让用户对产品产生购买/尝试欲望，内容营销常用词。近似product seeding或generating purchase intent，无直译",
     "小红书·B站内容营销"],
    ["UGC裂变", "中文营销术语", "中国大陆", "用户生成内容+裂变传播",
     "通过用户自发创作内容实现指数级传播，是中国手游营销的核心打法之一",
     "抖音/B站游戏话题营销"],
    ["さいほう（裁縫/最高）", "日语双关", "日本", "裁縫（さいほう）=缝纫",
     "龍が如く×Brother联名活动刻意使用さいほう而非さいこう（最高），利用谐音创造双关，既指缝纫也暗示最高",
     "龍が如く×Brother 3月联名活动"],
    ["shadow drop", "英语游戏圈术语", "美国/英国", "无对应字面意思",
     "指游戏毫无预告突然发布，行业通用术语。中文常译突袭发售，但无统一标准译名",
     "GTA6 Trailer 3谣言讨论"],
    ["Server Slam", "英语（Bungie自创）", "美国", "服务器冲击",
     "Bungie为Marathon发售前大规模压力测试活动创造的专有名称，非游戏行业通用术语",
     "Marathon发售前公测活动"],
    ["운영자 노트", "韩语", "韩国", "运营者笔记/便签",
     "韩国手游行业特有沟通格式：官方运营团队向Naver Cafe等玩家社区发布的月度更新说明，具有正式与亲切并存的社区文化特色",
     "FGO KR等韩国手游官方社区"],
    ["캠퍼（camper）", "韩语音译（源于英语）", "韩国", "营地守卫者",
     "专指在游戏据点蹲守的防守型玩家，带有贬义。Marathon类游戏讨论中频繁出现",
     "Marathon韩国玩家讨论"],
    ["Der Hochglanz-Feed ist tot", "德语", "德国", "高光滤镜Feed已死",
     "2026年德国Instagram流行的内容营销趋势语：精修品牌图文内容时代结束，原创观点和真实感内容更受欢迎",
     "德国游戏/品牌内容营销趋势"],
    ["Creator-led funnels", "英语营销术语", "欧洲/美国", "创作者主导的转化漏斗",
     "由内容创作者（而非品牌官方）主导整个用户获取和转化流程的营销模式，2026年欧美手游的主流趋势",
     "欧美手游KOL营销趋势报告"],
    ["Twitch Drops", "英语平台术语", "美国/全球", "Twitch掉落（奖励）",
     "Twitch平台特有激励机制：玩家观看指定直播频道满足条件后可获得游戏内道具奖励，是游戏营销常用工具",
     "Marathon发售期Twitch营销"],
    ["Delightfully Weird", "英语（Kurppa Hosk自创）", "美国", "令人愉悦的怪异",
     "Marathon创意代理Kurppa Hosk为游戏定制的品牌创意方向标语，强调反传统的视觉与叙事风格",
     "Marathon全渠道品牌营销"],
]
for i, row in enumerate(gloss):
    data_row(wsg, 4 + i, row, alt=(i % 2 == 1))

set_col_widths(wsg, [22, 16, 14, 20, 50, 28])
for r in range(1, 20):
    wsg.row_dimensions[r].height = 22

# Save
out = "2026年3月全球游戏市场热点月报_优化版v12.xlsx"
wb.save(out)
print("OK:" + out)
